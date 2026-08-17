# ===========================================================================
# AGENTE 1 — Extração determinística de execuções fiscais (PGMS / HERA)
# Versão 8.0
#
# MUDANÇAS EM RELAÇÃO A promptV7_1.py:
#   1. [feedback 4] Removida a camada de juízo APTO / NÃO APTO. O Agente 1
#      agora SÓ EXTRAI e REPORTA fatos (datas, citação, penhora, sinais
#      processuais, entidades). Não emite mais veredito jurídico — isso evita
#      o "overread" apontado e mantém a decisão legal com o humano / Agente 2.
#   2. [feedback 1] O JSON de saída agora inclui TODOS os processos (sem filtro
#      por decisão) e TODOS os campos que aparecem na planilha.
#   3. Removida a dependência da OpenAI. O Agente 1 é 100% determinístico e
#      offline (ver bloco "LLM / GEMINI — NOTA DE ARQUITETURA" abaixo).
#   4. Imports pesados (pdfplumber/pandas/openpyxl/pytesseract/pdf2image) são
#      carregados sob demanda (lazy) para permitir testes unitários das
#      funções puras sem exigir OCR/Poppler/Tesseract.
#
# As funções de extração (datas, citação, penhora, entidades, sinais) foram
# mantidas VERBATIM do V7_1 — lógica já validada, para não introduzir regressão.
# ===========================================================================

import os
import re
import unicodedata
from datetime import datetime
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed

# Dependências pesadas são importadas DENTRO das funções que as usam (lazy
# import). Assim o módulo pode ser importado para testes das funções puras de
# extração sem exigir OCR/Poppler/Tesseract/pandas instalados — e, se faltar
# alguma, o erro aparece de forma visível no ponto de uso, dizendo qual falta.


# Função para normalizar texto (remover acentos, convertir a minúsculas)
def normalizar(text):
    if not text:
        return ""
    return unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII').lower()

# Función para extraer fechas de un texto normalizado
def _extraer_fechas_de_texto(text_norm):
    # Patrones para extraer fechas (tanto con meses completos como abreviados)
    _MESES_COMPLETOS = {
        "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
        "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
        "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
    }
    _MESES_ABREV = {
        "jan": 1, "fev": 2, "mar": 3, "abr": 4,
        "mai": 5, "jun": 6, "jul": 7, "ago": 8,
        "set": 9, "out": 10, "nov": 11, "dez": 12
    }
    _PATRON_COMPLETO = r"(\d{1,2}) de (janeiro|fevereiro|marco|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro) de (\d{4})"
    _PATRON_ABREV    = r"(\d{1,2})\s+(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\.?\s+(\d{4})"
    fechas = []
    for dia, mes_txt, anio in re.findall(_PATRON_COMPLETO, text_norm):
        fechas.append(datetime(int(anio), _MESES_COMPLETOS[mes_txt], int(dia)))
    for dia, mes_txt, anio in re.findall(_PATRON_ABREV, text_norm):
        if mes_txt in _MESES_ABREV:
            fechas.append(datetime(int(anio), _MESES_ABREV[mes_txt], int(dia)))
    return fechas


# Esto silencia los logs internos de la librería que genera esos mensajes
logging.getLogger('pdfminer').setLevel(logging.ERROR)
logging.basicConfig(level=logging.INFO)

# Directorio de entrada y salida
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

# Las tres carpetas son sobreescribibles por variable de entorno para que la
# infraestructura (o cada lote) pueda dar su propio espacio aislado.
input_directory  = os.environ.get("PASTA_ENTRADA")    or os.path.join(CURRENT_DIR, "processos pra analiser")
PASTA_JSON       = os.environ.get("PASTA_JSON")       or os.path.join(CURRENT_DIR, "JSON")
PASTA_RESULTADOS = os.environ.get("PASTA_RESULTADOS") or os.path.join(CURRENT_DIR, "resultados")


# --- Configuración de OCR (fallback para páginas escaneadas) ---
OCR_DPI = 300
# [AJUSTABLE] DPI más bajo = OCR más rápido. Medido: 300dpi≈1.27s/página,
# 200dpi≈0.73s/página, 150dpi≈0.52s/página. Antes de bajarlo en producción,
# comparar la "Confiança OCR (%)" del Excel contra una muestra real escaneada.
OCR_MIN_CHARS = 20      # umbral empírico — ajustable según casos reales
OCR_IDIOMA = 'por'

# Agrupamiento de páginas OCR en lotes, para reducir llamadas a Poppler.
OCR_CLUSTER_GAP = 5     # páginas separadas por <= N páginas se fusionan en un lote
OCR_MAX_LOTE = 40       # tamaño máximo de un lote
OCR_MAX_WORKERS = 4     # lotes procesados en paralelo


def _pagina_necesita_ocr(texto_pagina, min_chars=OCR_MIN_CHARS):
    """
    pdfplumber devuelve None o una cadena muy corta cuando la página
    es una imagen escaneada sin capa de texto digital.
    """
    if texto_pagina is None:
        return True
    return len(texto_pagina.strip()) < min_chars


def _agrupar_paginas_en_lotes(paginas_ocr, gap_maximo=OCR_CLUSTER_GAP, max_lote=OCR_MAX_LOTE):
    """
    Agrupa números de página que necesitan OCR en rangos (inicio, fin)
    inclusive, para minimizar la cantidad de llamadas a Poppler.
    """
    if not paginas_ocr:
        return []
    paginas_ordenadas = sorted(paginas_ocr)
    lotes_brutos = []
    inicio = anterior = paginas_ordenadas[0]
    for p in paginas_ordenadas[1:]:
        if p - anterior > gap_maximo:
            lotes_brutos.append((inicio, anterior))
            inicio = p
        anterior = p
    lotes_brutos.append((inicio, anterior))

    lotes_finales = []
    for ini, fin in lotes_brutos:
        cursor = ini
        while cursor <= fin:
            sub_fin = min(cursor + max_lote - 1, fin)
            lotes_finales.append((cursor, sub_fin))
            cursor = sub_fin + 1
    return lotes_finales


def _ocr_lote(pdf_path, inicio, fin, paginas_necesarias, dpi=OCR_DPI, idioma=OCR_IDIOMA):
    """
    Renderiza el rango [inicio, fin] en UNA sola llamada a Poppler y aplica OCR
    solo a las páginas de `paginas_necesarias` dentro de ese rango.
    Devuelve {numero_pagina: (texto_extraido, confianza_0_a_100)}.
    """
    # Import lazy + falla visible si falta OCR: no interrumpe el lote entero,
    # devuelve texto vacío para esas páginas y registra el motivo en el log.
    try:
        import pytesseract
        from pytesseract import Output
        from pdf2image import convert_from_path
    except ImportError as e:
        logging.error(
            f"Dependência de OCR ausente ({e}). Página(s) {sorted(paginas_necesarias)} "
            f"de {os.path.basename(pdf_path)} ficarão sem texto. "
            f"Instale pytesseract/pdf2image + tesseract-ocr-por + poppler-utils.",
            exc_info=True,
        )
        return {p: ("", 0.0) for p in paginas_necesarias}

    resultados = {}
    try:
        imagenes = convert_from_path(pdf_path, dpi=dpi, first_page=inicio, last_page=fin)
    except Exception as e:
        logging.error(f"Error renderizando lote {inicio}-{fin} de {pdf_path}: {e}", exc_info=True)
        for p in paginas_necesarias:
            resultados[p] = ("", 0.0)
        return resultados

    for offset, imagen in enumerate(imagenes):
        numero_pagina = inicio + offset
        if numero_pagina not in paginas_necesarias:
            continue  # esta página del rango ya tenía texto digital
        try:
            datos = pytesseract.image_to_data(imagen, lang=idioma, output_type=Output.DICT)
            palabras, confianzas = [], []
            for texto, conf in zip(datos['text'], datos['conf']):
                if texto.strip():
                    palabras.append(texto)
                    if int(conf) >= 0:   # tesseract usa -1 cuando no calcula confianza
                        confianzas.append(int(conf))
            texto_final = ' '.join(palabras)
            confianza = sum(confianzas) / len(confianzas) if confianzas else 0.0
            resultados[numero_pagina] = (texto_final, confianza)
        except Exception as e:
            logging.error(f"Error OCR en página {numero_pagina} de {pdf_path}: {e}", exc_info=True)
            resultados[numero_pagina] = ("", 0.0)
    return resultados


# --- Filtro de tipo de documento: ¿es una execução fiscal de dívida ativa? ---
KEYWORDS_CLASSE_EXECUCAO_FISCAL = [
    "execucao fiscal",
    "execucoes fiscais",
    "cobranca da divida ativa",
    "cobranca de divida ativa",
    "execucao da divida ativa",
]

KEYWORDS_CLASSE_NAO_FISCAL = [
    "obrigacao de fazer",
    "mandado de seguranca",
    "acao civil publica",
    "fornecimento de medicamento",
    "sem registro na anvisa",
    "indenizacao por dano",
    "alvara judicial",
    "usucapiao",
    "interdicao",
    "acao popular",
    "desapropriacao",
]

KEYWORDS_TEXTO_EXECUCAO_FISCAL = [
    "execucao fiscal",
    "exequente",
    "certidao de divida ativa",
    "cda no", "cda n.", "cda no.",
    "lei 6.830", "lei n. 6.830", "lei no 6.830",
    "divida ativa",
    "embargos a execucao fiscal",
    "penhora",
]

KEYWORDS_TEXTO_NAO_FISCAL = [
    "obrigacao de fazer",
    "fornecimento de medicamento",
    "tratamento medico",
    "anvisa",
    "sistema unico de saude",
    "alimentos",
    "guarda do menor",
    "uniao estavel",
    "usucapiao",
    "mandado de seguranca",
]


def _extraer_classe_assunto(text_norm):
    """
    Busca el campo 'Classe - Assunto:' típico de las capas administrativas del PJe.
    """
    m = re.search(r"classe\s*[-/]?\s*assunto\s*:?\s*(.{0,300})", text_norm, re.DOTALL)
    if not m:
        return None
    fragmento = re.sub(r"\s+", " ", m.group(1))
    corte = re.split(
        r"reclamante|requerente|autor|orgao julgador|exequente|executado|reclamado",
        fragmento
    )[0]
    return corte.strip()


def detectar_tipo_processo(text):
    """
    Determina si el texto corresponde a una execução fiscal de dívida ativa.
    Devuelve dict con es_execucao_fiscal, confianza, motivo, classe_assunto.
    """
    text_norm = normalizar(text)
    classe_assunto = _extraer_classe_assunto(text_norm)

    if classe_assunto:
        if any(normalizar(k) in classe_assunto for k in KEYWORDS_CLASSE_NAO_FISCAL):
            return {
                "es_execucao_fiscal": False,
                "confianza": "alta",
                "motivo": f"Classe-Assunto indica outro tipo de ação: '{classe_assunto[:150]}'",
                "classe_assunto": classe_assunto,
            }
        if any(normalizar(k) in classe_assunto for k in KEYWORDS_CLASSE_EXECUCAO_FISCAL):
            return {
                "es_execucao_fiscal": True,
                "confianza": "alta",
                "motivo": f"Classe-Assunto confirma execução fiscal: '{classe_assunto[:150]}'",
                "classe_assunto": classe_assunto,
            }
        # Campo encontrado pero ambiguo: cae a la heurística por conteo

    score_fiscal = sum(1 for k in KEYWORDS_TEXTO_EXECUCAO_FISCAL if normalizar(k) in text_norm)
    score_nao_fiscal = sum(1 for k in KEYWORDS_TEXTO_NAO_FISCAL if normalizar(k) in text_norm)

    if score_fiscal == 0 and score_nao_fiscal == 0:
        return {
            "es_execucao_fiscal": True,
            "confianza": "baixa",
            "motivo": "Nenhum marcador de tipo de ação encontrado — recomenda-se revisão manual",
            "classe_assunto": classe_assunto,
        }

    es_fiscal = score_fiscal >= score_nao_fiscal
    return {
        "es_execucao_fiscal": es_fiscal,
        "confianza": "media",
        "motivo": f"Heurística por contagem de palavras-chave: fiscal={score_fiscal}, não-fiscal={score_nao_fiscal}",
        "classe_assunto": classe_assunto,
    }


# 1. Extraer texto de PDF por páginas
def extract_text_by_page(pdf_path):
    """
    Extrae texto de cada página. Páginas sin texto digital suficiente se agrupan
    en lotes y se procesan con OCR en paralelo.
    Devuelve (list[str], metadata_dict).
    """
    import pdfplumber  # lazy import — falla visible aquí si no está instalado

    textos = {}
    with pdfplumber.open(pdf_path) as pdf:
        total_paginas = len(pdf.pages)
        paginas_que_necesitan_ocr = []
        for i, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text()
            if _pagina_necesita_ocr(texto):
                paginas_que_necesitan_ocr.append(i)
                textos[i] = None  # se completa en la fase de OCR
            else:
                textos[i] = texto

    metadata = {"paginas_ocr": [], "confianza_ocr": {}}

    if paginas_que_necesitan_ocr:
        lotes = _agrupar_paginas_en_lotes(paginas_que_necesitan_ocr)
        necesarias_por_lote = {
            (ini, fin): set(p for p in paginas_que_necesitan_ocr if ini <= p <= fin)
            for ini, fin in lotes
        }
        logging.info(
            f"  {os.path.basename(pdf_path)}: {len(paginas_que_necesitan_ocr)} página(s) "
            f"necesitam OCR, agrupadas em {len(lotes)} lote(s) (até {OCR_MAX_WORKERS} em paralelo)"
        )
        with ThreadPoolExecutor(max_workers=OCR_MAX_WORKERS) as executor:
            futuros = {
                executor.submit(_ocr_lote, pdf_path, ini, fin, necesarias_por_lote[(ini, fin)]): (ini, fin)
                for ini, fin in lotes
            }
            concluidos = 0
            for futuro in as_completed(futuros):
                ini, fin = futuros[futuro]
                concluidos += 1
                try:
                    resultado_lote = futuro.result()
                    for numero_pagina, (texto_ocr, confianza) in resultado_lote.items():
                        textos[numero_pagina] = texto_ocr
                        metadata["paginas_ocr"].append(numero_pagina)
                        metadata["confianza_ocr"][numero_pagina] = round(confianza, 1)
                    logging.info(f"  Lote {concluidos}/{len(lotes)} concluído (páginas {ini}-{fin})")
                except Exception as e:
                    logging.error(f"Lote {ini}-{fin} de {pdf_path} falhou: {e}", exc_info=True)

    metadata["paginas_ocr"].sort()
    lista_textos = [textos[i] for i in range(1, total_paginas + 1)]
    return lista_textos, metadata


# 2. Filtrar texto relevante con palabras clave
def filter_text_by_keywords(text, keywords):
    relevant_lines = []
    norm_keywords = [normalizar(k) for k in keywords]
    for line in text.splitlines():
        if any(kw in normalizar(line) for kw in norm_keywords):
            relevant_lines.append(line)
    return " ".join(relevant_lines)


# 3. Obtener la fecha más reciente en el texto
KEYWORDS_MOVIMENTACAO_PROCESSUAL = [
    "despacho", "decisao interlocutoria", "decisao",
    "sentenca", "acordao",
    "conferi.", "digitei, eu",
    "certidao de publicacao de relacao",
    "certidao de remessa da intimacao",
    "certidao de intimacao",
    "ciencia da intimacao",
    "certidao de publicacao",
    "pede deferimento",
    "pede juntada",
    "vem requerer",
    "vem expor e requerer",
    "vem, por seu procurador",
    "vem, perante",
    "vem, respeitosamente",
    "nestes termos, pede deferimento",
    "requer a v. exa",
    "requer a vossa excelencia",
    "por seu procurador infrafirmado",
    "por sua procuradora infrafirmada",
    "por seu procurador ao fim assinado",
    "vem aduzir e requerer",
    "reiterar pedido",
    "certifico que",
    "certifico, para os devidos fins",
    "o referido e verdade e dou fe",
    "o referido e verdade",
    "ato ordinatorio",
    "cumpra-se",
    "publique-se. intime-se",
    "salvador (ba),",
    "salvador, ba,",
    "data da intimacao",
    "encaminhado para intimacao no portal eletronico",
]

def fecha_ultima_movimentacao(text):
    """
    Solo considera fechas cercanas a keywords de movimentación processual,
    ignorando fechas de fichas cadastrais, extratos y consultas CNPJ.
    """
    text_norm = normalizar(text)
    fechas = []

    MARCADORES_EXCLUSION = [
        "ficha cadastral",
        "extrato fiscal",
        "consulta de dados via cpf",
        "dados do cnpj",
        "dados de empresa via cnpj",
        "posicao de debito",
        "certidao de divida ativa",
        "termo de confissao de divida",
    ]
    excl_norm = [normalizar(m) for m in MARCADORES_EXCLUSION]

    for keyword in KEYWORDS_MOVIMENTACAO_PROCESSUAL:
        keyword_norm = normalizar(keyword)
        for match in re.finditer(re.escape(keyword_norm), text_norm):
            start = max(0, match.start() - 300)
            end   = match.end() + 300
            fragmento_norm = text_norm[start:end]

            if any(marker in fragmento_norm for marker in excl_norm):
                continue

            fechas_encontradas = _extraer_fechas_de_texto(fragmento_norm)
            if fechas_encontradas:
                fechas.extend(fechas_encontradas)

    if not fechas:
        return _fecha_mas_reciente_fallback(text)

    return max(fechas)


def _fecha_mas_reciente_fallback(text):
    """
    Fallback: escanea todo el texto pero excluye secciones de fichas cadastrais,
    extratos y consultas CNPJ.
    """
    MARCADORES_EXCLUSION = [
        "ficha cadastral",
        "extrato fiscal",
        "consulta de dados via cpf",
        "dados do cnpj",
        "dados de empresa via cnpj",
        "posicao de debito",
        "certidao de divida ativa",
        "termo de confissao de divida",
    ]

    lineas = text.split("\n")
    lineas_filtradas = []
    excluir = False

    for linea in lineas:
        linea_norm = normalizar(linea)
        if any(normalizar(m) in linea_norm for m in MARCADORES_EXCLUSION):
            excluir = True
        if any(kw in linea_norm for kw in ["poder judiciario", "comarca de salvador", "despacho", "decisao", "certidao"]):
            excluir = False
        if not excluir:
            lineas_filtradas.append(linea)

    texto_filtrado = "\n".join(lineas_filtradas)
    fechas = _extraer_fechas_de_texto(normalizar(texto_filtrado))
    return max(fechas) if fechas else None


def _limpar_referencias_legais(text_norm):
    """
    Remove datas de referencias legais (decretos, leis, resolucoes) e datas de
    assinatura de CDAs para evitar que sejam capturadas como datas de citacao.
    """
    padroes = [
        r'decreto\s+\w+\s+n\w*\s*\d+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'decreto\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'lei\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'resolucao\s+n\w*\s*\d+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'instrucao\s+normativa\s+\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'portaria\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'certifico que se acha inscrito[^S]{0,400}?salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'estes sao os elementos contidos[^S]{0,200}?salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}\s+raimundo\s+cordeiro',
        r'salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}\s+\w+\s+cordeiro',
        r'coordenador\s+da\s+cda',
    ]
    result = text_norm
    for pat in padroes:
        result = re.sub(pat, ' [REF_LEGAL] ', result)
    return result


def extraer_fecha_cercana(text, keywords, ventana=500, prefer='latest', min_year=1990):
    """
    Busca fechas cerca de las ocurrencias de `keywords`.
    prefer='latest' devuelve la más reciente; 'earliest' la más antigua.
    """
    text_norm = _limpar_referencias_legais(normalizar(text))
    fechas = []
    for keyword in keywords:
        kw_norm = normalizar(keyword)
        for match in re.finditer(re.escape(kw_norm), text_norm):
            start = max(0, match.start() - ventana)
            end = match.end() + ventana
            fechas_local = _extraer_fechas_de_texto(text_norm[start:end])
            fechas_local = [f for f in fechas_local if f.year >= min_year]
            if fechas_local:
                fechas.extend(fechas_local)
    if not fechas:
        return None
    if prefer == 'earliest':
        return min(fechas)
    return max(fechas)


def extraer_fechas_citacion(text):
    """
    Extrae fecha_orden, fecha_intento y fecha_efectiva relacionadas a la citación.
    """
    KEYWORDS_ORDEN_ESPECIFICAS = [
        "determino a citação",
        "ordeno a citação",
        "expeça-se o competente mandado de citação",
        "expeça-se carta de citação",
        "expeça-se mandado de citação",
        "proceda-se à citação",
        "promova-se a citação",
        "expedir carta de citação",
        "diligencie-se para citação",
        "ao oficial de justiça, cite",
        "cite-se por edital", "citação por edital",
        "converta-se em mandado de citação",
    ]
    KEYWORDS_ORDEN_GENERICAS = [
        "expeça-se citação",
        "seja citado", "sejam citados",
        "citem-se",
        "cite-se",
        "cite(m)-se",
        "proceda-se a citacao",
        "proceda-se a citação",
        "o presente despacho servira como mandado",
        "registre-se.",
        "publique-se. registre-se",
    ]
    KEYWORDS_ORDEN = KEYWORDS_ORDEN_ESPECIFICAS + KEYWORDS_ORDEN_GENERICAS

    KEYWORDS_INTENTO_APOS = [
        "conferi.",
        "digitei, eu",
    ]
    KEYWORDS_INTENTO = [
        "carta com ar", "aviso de recebimento", "tentativa",
        "aviso de recebimento negativo",
    ]

    KEYWORDS_EFECTIVA = [
        "recebido", "assinado", "assinatura", "assinatura do recebedor", "recebido em", "recebido com"
    ]

    fecha_orden = extraer_fecha_cercana(text, KEYWORDS_ORDEN, ventana=800, prefer='earliest')

    import re as _re
    def _fecha_intento_apos(text_norm, keywords, min_year=1990):
        """Busca data apenas APÓS o keyword (ventana_antes=10)."""
        fechas = []
        for kw in keywords:
            kw_norm = normalizar(kw)
            for m in _re.finditer(_re.escape(kw_norm), text_norm):
                trecho = text_norm[max(0, m.start()-10): m.end()+300]
                fs = [f for f in _extraer_fechas_de_texto(trecho) if f.year >= min_year]
                fechas.extend(fs)
        return min(fechas) if fechas else None

    text_norm_clean = _limpar_referencias_legais(normalizar(text))
    fecha_intento = _fecha_intento_apos(text_norm_clean, KEYWORDS_INTENTO_APOS)
    if not fecha_intento:
        fecha_intento = extraer_fecha_cercana(text, KEYWORDS_INTENTO, ventana=500, prefer='earliest')

    fecha_efectiva = extraer_fecha_cercana(text, KEYWORDS_EFECTIVA, ventana=400, prefer='earliest')

    logging.debug(f"  fecha_orden (primera):    {fecha_orden}")
    logging.debug(f"  fecha_intento (primera):  {fecha_intento}")
    logging.debug(f"  fecha_efectiva (posible): {fecha_efectiva}")

    return fecha_orden, fecha_intento, fecha_efectiva


# 4. Extraer estado de citación
def extract_citacion(text):
    KEYWORDS_CITACION_OK = [
        "certifico que procedi a citacao",
        "certifico que o executado foi citado",
        "certifico que citei",
        "certifico ter realizado a citacao",
        "fica citado",
        "devidamente citado",
        "ar positivo",
        "certidao de decurso de prazo",
        "decorreu o prazo legal sem qualquer manifestacao",
        "decorreu o prazo legal",
        "decurso de prazo",
        "nao se manifestou quanto ao pagamento",
        "apresentou embargos",
        "embargos foram opostos",
        "citacao valida",
        "instrumento de confissao de divida e compromisso de pagamento parcelado",
        "instrumento de confissao de divida",
        "confissao de divida e compromisso",
        "exarou o ciente",
        "aceitou a contrafe que lhe foi oferecida",
        "ele aceitou a contrafe",
        "citado nos autos",
        "parcelamento de debitos",
    ]

    KEYWORDS_CITACION_NAO_OK = [
        "aviso de recebimento negativo",
        "intime-se a fazenda publica para que adote as providencias cabiveis",
        "o reu nao foi citado",
        "executado nao foi citado",
        "sem citacao do executado",
        "nao houve citacao",
        "ausencia de citacao",
        "nao logrando exito na citacao",
        "nao foi possivel realizar a citacao",
        "a parte executada nao foi citada",
        "deixei de proceder a citacao",
        "deixei de citar",
        "nao encontrado o executado",
        "nao foi localizado o executado",
        "nao reside no endereco",
        "nao mora no endereco",
        "nao conhece o executado",
        "nao sabe informar o seu paradeiro",
        "motivos de devolucao",
        "mudou-se",
        "nao procurado",
        "nao existe o numero",
        "endereco insuficiente",
        "desconhecido",
        "falecido",
    ]

    text_norm = normalizar(text)

    if any(normalizar(k) in text_norm for k in KEYWORDS_CITACION_NAO_OK):
        for k_ok in KEYWORDS_CITACION_OK:
            if normalizar(k_ok) in text_norm:
                return "HOUVE CITAÇÃO"
        return "NÃO HOUVE ou TENTATIVA FALHA"

    if any(normalizar(k) in text_norm for k in KEYWORDS_CITACION_OK):
        return "HOUVE CITAÇÃO"

    return "Citação não encontrado"


# 5. Extraer resultado de la penhora
def extract_penhora(text):
    KEYWORDS_BACENJUD_POSITIVO = [
        "bloqueio bacenjud", "penhora bacenjud",
        "bloqueio sisbajud", "penhora sisbajud",
        "penhora on-line", "penhora online",
    ]
    KEYWORDS_BACENJUD_SOLICITADO = [
        "via sistema sisbajud",
        "via sistema bacenjud",
        "via bacenjud",
        "via sisbajud",
        "sistema sisbajud",
        "sistema bacenjud",
        "bloqueio de dinheiro/ativos financeiros",
        "bloqueio de ativos financeiros",
        "requer o bloqueio de dinheiro",
        "bacen jud",
        "sis bajud",
        "nos moldes do bacen jud",
        "nos moldes do bacenjud",
    ]
    KEYWORDS_BACENJUD_CONFIRMADO = [
        "valores bloqueados",
        "bloqueio efetuado",
        "bloqueio realizado",
        "dinheiro bloqueado",
        "penhora on-line efetuada",
        "penhora online efetuada",
        "extrato de bloqueio",
        "certidao de bloqueio",
        "comprovante de bloqueio",
    ]
    KEYWORDS_BACENJUD_NEGATIVO = [
        "resultado negativo da diligencia bacenjud",
        "resultado negativo da diligencia sisbajud",
        "suspensao pelo art. 40 da lef",
        "sem saldo positivo",
        "nao ha saldo",
        "bacenjud sem exito",
        "sisbajud sem exito",
        "diligencia bacenjud restou infrutifera",
        "diligencia sisbajud restou infrutifera",
        "bloqueio desbloqueado",
        "desbloqueio do valor",
        "nao possui relacionamento com as instituicoes financeiras",
        "cpf indicado nao possui relacionamento",
        "cnpj indicado nao possui relacionamento",
        "negativa bacenjud",
        "ar negativo e/ou negativa bacenjud",
        "negativa do bacenjud",
        "bacenjud negativo",
        "resultado bacenjud negativo",
        "nao possui relacionamentos com",
        "cpf nao possui relacionamento",
        "tentativa de penhora on-line",
    ]

    KEYWORDS_RENAJUD_ATIVO = [
        "comprovante de inclusao de restricao veicular",
        "insercao de restricao veicular",
        "bloqueio renajud",
        "restricao renajud",
        "penhora de veiculo",
        "auto de penhora de veiculo",
    ]
    KEYWORDS_RENAJUD_NEGATIVO = [
        "renajud sem exito",
        "restricao renajud cancelada",
        "nao foram localizados veiculos",
        "pesquisa renajud negativa",
    ]

    KEYWORDS_PENHORA_IMOVEL = [
        "penhora de imovel",
        "penhora do imovel",
        "registro de penhora",
        "matricula do imovel penhorado",
        "imovel penhorado",
        "termo de penhora de imovel",
        "penhora sobre imovel",
    ]

    KEYWORDS_PENHORA_FATURAMENTO = [
        "penhora de faturamento",
        "penhora sobre faturamento",
        "penhora sobre o faturamento",
        "deposito de percentual do faturamento",
    ]

    KEYWORDS_INDISPONIBILIDADE = [
        "indisponibilidade de bens",
        "decretada a indisponibilidade",
        "cnib",
        "cadastro nacional de indisponibilidade",
    ]

    KEYWORDS_PENHORA_QUOTAS = [
        "penhora de quotas",
        "penhora de cotas",
        "penhora de participacao societaria",
        "penhora de acoes",
    ]

    KEYWORDS_PENHORA_CREDITOS = [
        "penhora de creditos",
        "penhora de precatorio",
        "penhora de direitos",
        "penhora de direitos hereditarios",
        "penhora de aplicacao financeira",
    ]

    KEYWORDS_TENTATIVA_PENHORA = [
        "termo de penhora",
        "auto de penhora e avaliacao",
        "diligencia de penhora",
        "tentativa de penhora",
        "intimado da penhora",
        "intimados da penhora",
        "oficial de justica nao localizou bens",
        "nao foram localizados bens",
        "nao encontrou bens",
        "sem bens a penhorar",
        "bens insuficientes",
    ]
    text_norm = normalizar(text)

    if any(normalizar(k) in text_norm for k in KEYWORDS_RENAJUD_ATIVO):
        if any(normalizar(k) in text_norm for k in KEYWORDS_RENAJUD_NEGATIVO):
            return "tentativa de penhora renajud (negativa)"
        return "bloqueio renajud ativo"

    if any(normalizar(k) in text_norm for k in KEYWORDS_INDISPONIBILIDADE):
        return "indisponibilidade de bens (CNIB)"

    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_IMOVEL):
        return "penhora de imóvel"

    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_FATURAMENTO):
        return "penhora de faturamento"

    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_QUOTAS):
        return "penhora de quotas/ações"

    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_CREDITOS):
        return "penhora de créditos/direitos"

    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_NEGATIVO):
        return "tentativa de penhora bacenjud (negativa)"

    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_POSITIVO):
        return "penhora bacenjud/sisbajud"

    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_SOLICITADO):
        if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_NEGATIVO):
            return "tentativa de penhora bacenjud (negativa)"
        if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_CONFIRMADO):
            return "penhora bacenjud/sisbajud"
        return "sisbajud/bacenjud solicitado (resultado desconhecido)"

    if any(normalizar(k) in text_norm for k in KEYWORDS_TENTATIVA_PENHORA):
        return "tentativa de penhora (sem resultado)"

    if "penhora nao realizada" in text_norm:
        return "penhora não realizada"

    return "Penhora não encontrado"


# --- Detección de parcelamento / suspensión ---
KEYWORDS_PARCELAMENTO_ATIVO = [
    "defiro o pedido de suspensao",
    "defiro a suspensao",
    "determino a suspensao do feito",
    "determino a suspensao da execucao",
    "suspendo o feito",
    "suspendo o curso do feito",
    "suspendo/mantenho suspenso",
    "suspensao do feito",
    "suspensao da execucao",
    "suspensao do processo",
    "requerer a suspensao do feito",
    "requer a suspensao do feito",
    "requerer a suspensao da execucao",
    "requer a suspensao da execucao",
    "suspensao pelo parcelamento",
    "pleiteou administrativamente o parcelamento",
    "adesao ao parcelamento",
    "aderiu ao parcelamento",
    "aderiu o executado",
    "parcelamento a que aderiu",
    "parcelamento do debito exequendo",
    "parcelamento do credito tributario",
    "parcelamento do debito tributario",
    "parcelamento administrativo de debitos",
    "parcelamento administrativo",
    "parcelamento em",
    "parcelas mensais e sucessivas",
    "em 38 parcelas",
    "em 12 parcelas",
    "em 24 parcelas",
    "em 36 parcelas",
    "em 48 parcelas",
    "em 60 parcelas",
    "adesao ao pad",
    "pad homologado",
    "bloq pad",
    "pad n",
    "pad no",
    "pad nr",
    "parcelamento pad",
    "pad internet",
    "instrumento de confissao de divida e compromisso de pagamento parcelado",
    "compromisso de pagamento parcelado",
    "instrumento de confissao",
    "art. 151, vi, do ctn",
    "art. 151, inc. vi",
    "art. 151, inciso vi",
    "art. 151, inc. vi,",
    "art. 151, i, do ctn",
    "art. 151, inc. i, do ctn",
    "art. 151, inc. i,",
    "art. 151, inciso i",
    "art. 151, inc. 1,",
    "art. 151, 1, do ctn",
    "art. 151, inc. 1, do",
    "151, inc. 1,",
    "art. 313, ii, do cpc",
    "art. 265, inc. ii, do cpc",
    "art. 265, inc. ii",
]

KEYWORDS_SUSPENSAO_ART40 = [
    "resultado negativo da diligencia bacenjud",
    "resultado negativo da diligencia sisbajud",
    "suspendo a presente execucao fiscal",
    "art. 40, caput, da lei n. 6.830",
    "arquive-se o feito nos termos do art. 40",
    "suspensao da execucao fiscal pelo prazo de 01",
]

def extract_suspensao_art40(text):
    text_norm = normalizar(text)
    if any(normalizar(k) in text_norm for k in KEYWORDS_SUSPENSAO_ART40):
        return "processo suspenso — art. 40 LEF (bens não localizados)"
    return None


def extract_extincao(text):
    """
    Detecta se o processo foi extinto/sentenciado. Retorna string ou None.
    Camadas: 1) reversão (extinção reformada por acórdão) → None;
    2) keywords fortes (sozinhas bastam); 3) keywords fracas (só com forte).
    """
    KEYWORDS_REVERSAO = [
        "dar provimento",
        "dou provimento",
        "da-se provimento",
        "recurso provido",
        "provimento ao recurso",
        "reforma a sentenca",
        "reforma-se a sentenca",
        "reformando a sentenca",
        "anulando a sentenca",
        "anulo a sentenca",
        "retorno dos autos ao primeiro grau",
        "retornem os autos ao juizo de origem",
        "retornem-se os autos",
        "retorno ao juizo de origem",
        "regular tramitacao",
        "para regular tramitacao",
        "redirecionamento da execucao",
        "redirecionar a execucao",
        "desconsideracao da personalidade juridica",
        "incluir o socio",
        "citar o socio",
    ]

    KEYWORDS_FORTES = [
        "processo ja se encontra sentenciado",
        "ja se encontra sentenciado",
        "processo sentenciado",
        "extingo o processo com resolucao do merito",
        "declaro a prescricao",
        "declaro extinto o processo",
        "julgo extinta a execucao",
        "julgo extinto o feito",
        "extingo a execucao",
        "extincao da execucao",
        "extincao do processo",
        "sentenca de extincao",
        "extinto por prescricao",
        "extinta por prescricao",
        "processo extinto",
    ]

    KEYWORDS_FRACAS = [
        "sentenca transitada",
        "transitada em julgado",
        "transito em julgado",
        "processo arquivado",
        "arquivem-se",
        "dando-se baixa",
        "extinção da execução",
        "extinção do processo",
    ]

    text_norm = normalizar(text)

    if any(normalizar(k) in text_norm for k in KEYWORDS_REVERSAO):
        return None

    tem_forte = any(normalizar(k) in text_norm for k in KEYWORDS_FORTES)
    if tem_forte:
        return "processo extinto/sentenciado"

    return None


# ===========================================================================
# EXTRACCIÓN DE ENTIDADES (datos estructurados del processo)
# ===========================================================================


def _so_digitos(s):
    return re.sub(r"\D", "", s or "")
 
 
def validar_cpf(cpf):
    """True se `cpf` (com ou sem máscara) tem 11 dígitos e DVs válidos."""
    cpf = _so_digitos(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    for fim in (9, 10):
        soma = sum(int(cpf[i]) * (fim + 1 - i) for i in range(fim))
        dv = (soma * 10) % 11
        if dv == 10:
            dv = 0
        if dv != int(cpf[fim]):
            return False
    return True
 
 
def validar_cnpj(cnpj):
    """True se `cnpj` (com ou sem máscara) tem 14 dígitos e DVs válidos."""
    cnpj = _so_digitos(cnpj)
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False
    pesos = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    for fim in (12, 13):
        w = pesos[1:] if fim == 12 else pesos
        r = sum(int(cnpj[i]) * w[i] for i in range(fim)) % 11
        dv = 0 if r < 2 else 11 - r
        if dv != int(cnpj[fim]):
            return False
    return True
 
 
def _fmt_cpf(d):
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
 
 
def _fmt_cnpj(d):
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
 
 
# Rótulos que indicam que o número pertence ao EXECUTADO / contribuinte / sócio.
_CTX_FORTE_EXECUTADO = (
    "executad", "contra ", "devedor", "reu", "contribuinte", "cpf/cnpj",
    "cnpj/cpf", "cpf do responsavel", "nome do responsavel", "socio",
    "responsavel", "cpf:", "cnpj:", "cpf ", "cnpj ",
)
# Rótulos que indicam EXEQUENTE / assinante / procurador -> NÃO é o executado.
_CTX_REJEITAR = (
    "cnpj/mf", "exequente", "credor",
    "municipio do salvador", "municipio de salvador",
    "signed by", "assinado por", "procurador ", "oab",
)
 
# Token candidato: começa e termina em dígito, só admite . - / no meio
# (não cruza espaços, para não fundir "CEP 40230731 - 713", telefone, etc.).
_RE_CANDIDATO = re.compile(r"\d[\d./-]{9,16}\d")
_CTX_JANELA = 80  # nº de caracteres de contexto anterior analisados
 
 
def _rotulo_mais_proximo(ctx):
    """
    Decide pelo rótulo MAIS PRÓXIMO do número (maior posição em `ctx`):
    'rejeitar' se o rótulo colado ao número é de exequente/assinatura;
    'forte' se é de executado/contribuinte/sócio; None se não há rótulo.
    Evita rejeitar um CPF legítimo só porque um 'cnpj/mf' de outra entidade
    aparece longe, mas dentro da janela.
    """
    pos_forte = max((ctx.rfind(k) for k in _CTX_FORTE_EXECUTADO), default=-1)
    pos_rej = max((ctx.rfind(k) for k in _CTX_REJEITAR), default=-1)
    if pos_rej > pos_forte:
        return "rejeitar"
    if pos_forte > -1:
        return "forte"
    return None
 
 
def _extrair_cpf_cnpj(text):
    """
    Extrai o CPF/CNPJ do EXECUTADO (prioriza CNPJ). Devolve str formatado ou None.
    Robustez V8.1: valida DV, aceita CNPJ sem zero à esquerda/sem pontos, aceita
    CPF cru só com rótulo forte, e rejeita por contexto exequente/assinaturas.
    """
    if not text:
        return None
    tn = normalizar(text)  # normalizar() já existe no agente1
    melhor = None  # (chave_ordenacao, tipo, digitos)
 
    for m in _RE_CANDIDATO.finditer(tn):
        token = m.group(0)
        dig = _so_digitos(token)
        formatado = any(c in token for c in "./-")
 
        # Classifica CPF vs CNPJ com validação de DV.
        tipo = None
        if len(dig) == 14 and validar_cnpj(dig):
            tipo = "cnpj"
        elif len(dig) == 13 and validar_cnpj("0" + dig):  # zero à esquerda perdido
            tipo, dig = "cnpj", "0" + dig
        elif len(dig) == 11 and validar_cpf(dig):
            tipo = "cpf"
        if not tipo:
            continue
 
        ctx = tn[max(0, m.start() - _CTX_JANELA): m.start()]
        rotulo = _rotulo_mais_proximo(ctx)
        if rotulo == "rejeitar":
            continue
        forte = (rotulo == "forte")
 
        # Número CRU (sem máscara) sem rótulo forte é arriscado (telefone/
        # registro que passe o DV por acaso) -> descarta.
        if not formatado and not forte:
            continue
 
        chave = (
            (10 if forte else 0) + (3 if formatado else 0),  # 1º: rótulo/forma
            1 if tipo == "cnpj" else 0,                       # 2º: CNPJ > CPF
            -m.start(),                                       # 3º: 1ª ocorrência
        )
        if melhor is None or chave > melhor[0]:
            melhor = (chave, tipo, dig)
 
    if melhor is None:
        return None
    _, tipo, dig = melhor
    return _fmt_cnpj(dig) if tipo == "cnpj" else _fmt_cpf(dig)
 

def _extrair_nome_executado(text):
    """Extrae el nombre/razón social del executado."""
    _PATRONES = [
        r"executad[oa]\s*:\s*([^\n\r]{3,80})",
        r"r[eé]u\s*:\s*([^\n\r]{3,80})",
        r"contribuinte\s*:\s*([^\n\r]{3,80})",
        r"nome[\/\s]?raz[aã]o social\s*:\s*([^\n\r]{3,80})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            nome = m.group(1).strip().rstrip(".,;")
            if len(nome) >= 4 and not nome.replace(" ", "").isdigit():
                return nome[:100]
    return None


def _extrair_nome_exequente(text):
    """Extrae el nombre del exequente. Limpia comillas OCR al inicio."""
    _PATRONES = [
        r"exequente\s*:\s*([^\n\r]{3,80})",
        r"credor\s*:\s*([^\n\r]{3,80})",
        r"autor\s*:\s*([^\n\r]{3,80})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            nome = m.group(1).strip().rstrip(".,;")
            nome = nome.lstrip("'\"\u201c\u201d\u2018\u2019").strip()
            if len(nome) >= 4:
                return nome[:80]
    return None



def _brl_to_float(s):
    """'1.143,74' -> 1143.74. Tolerante a lixo de OCR; nunca lança."""
    s = (s or "").strip().rstrip(".").rstrip(",")
    if not s:
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0
 
 
def _float_to_brl(v):
    """1143.74 -> '1.143,74'."""
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
 
 
_RE_CDA_EXERCICIO   = re.compile(r"certidao de debito n\.?\s*\d{2}\.(\d{4})\.")
_RE_EXERCICIO_LABEL = re.compile(r"exercicio\s+(\d{4})")
_RE_DATA_INSCR      = re.compile(r"data de inscricao\s*:?\s*(\d{2}/\d{2}/\d{4})")
_RE_TOTAL_CDA       = re.compile(r"\btotal\s+r\$\s*([\d.,]+)")
_RE_ORIGINARIO      = re.compile(r"valor originari[oa]\s*r\$?\s*([\d.,]+)")
 
 
def _parse_data(s):
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except (ValueError, TypeError):
        return None
 
 
def _coletar_cdas(tn):
    """
    Divide o texto normalizado em blocos de Certidão de Débito e extrai por
    bloco: exercício, Data de Inscrição, Valor Originário e Total.
    Só retorna blocos que tenham ao menos um dos valores.
    """
    marcadores = [m.start() for m in re.finditer(r"certidao de debito", tn)]
    if not marcadores:
        return []
    marcadores.append(len(tn))
 
    cdas = []
    for i in range(len(marcadores) - 1):
        bloco = tn[marcadores[i]:marcadores[i + 1]]
        m_ex = _RE_CDA_EXERCICIO.search(bloco) or _RE_EXERCICIO_LABEL.search(bloco)
        exercicio = m_ex.group(1) if m_ex else None
        m_dt  = _RE_DATA_INSCR.search(bloco)
        m_tot = _RE_TOTAL_CDA.search(bloco)
        m_ori = _RE_ORIGINARIO.search(bloco)
        if not m_tot and not m_ori:
            continue
        cdas.append({
            # blocos sem exercício NÃO se fundem entre si (chave única por bloco)
            "chave"         : exercicio or f"_bloco{i}",
            "data_inscricao": _parse_data(m_dt.group(1)) if m_dt else None,
            "total"         : _brl_to_float(m_tot.group(1)) if m_tot else None,
            "originario"    : _brl_to_float(m_ori.group(1)) if m_ori else None,
        })
    return cdas
 
 
def _dedup_por_exercicio(cdas):
    """
    Mantém 1 CDA por exercício: a de Data de Inscrição mais recente (a vigente).
    Sem data perde para quem tem data; entre iguais, fica a última vista.
    """
    vigentes = {}
    for c in cdas:
        k = c["chave"]
        atual = vigentes.get(k)
        if atual is None:
            vigentes[k] = c
            continue
        d_novo, d_atual = c["data_inscricao"], atual["data_inscricao"]
        if d_atual is None and d_novo is not None:
            vigentes[k] = c
        elif d_novo is not None and d_atual is not None and d_novo >= d_atual:
            vigentes[k] = c
        elif d_atual is None and d_novo is None:
            vigentes[k] = c
    return list(vigentes.values())
 
 
def _extrair_valor(text):
    """
    Devuelve (valor_original, valor_atualizado) como 'R$ X.XXX,XX' o None.
    Ver cabeçalho do bloco para a regra de soma/dedup por exercício.
    """
    if not text:
        return None, None
    tn = normalizar(text)  # normalizar() já existe no agente1
 
    valor_original = None
    valor_atualizado = None
 
    # --- Fonte primária: Certidões de Débito, 1 por exercício (a vigente) ---
    cdas = _dedup_por_exercicio(_coletar_cdas(tn))
    if cdas:
        totais = [c["total"] for c in cdas if c["total"] is not None]
        origs  = [c["originario"] for c in cdas if c["originario"] is not None]
        if totais:
            valor_atualizado = f"R$ {_float_to_brl(sum(totais))}"
        if origs:  # [SUPOSIÇÃO A REVISAR] remova estas 2 linhas p/ voltar ao 1º só
            valor_original = f"R$ {_float_to_brl(sum(origs))}"
 
    # --- Fallbacks para documentos SEM certidões (outros formatos) ---
    if valor_original is None:
        for pat in [r"valor origin[aá]ri[oa]\s*[r\$:\s]+([\d.,]+)",
                    r"valor original\s*[r\$:\s]+([\d.,]+)",
                    r"vl\.?\s*original\s*[r\$:\s]+([\d.,]+)"]:
            m = re.search(pat, tn)
            if m:
                valor_original = f"R$ {m.group(1).strip()}"
                break
 
    if valor_atualizado is None:
        for pat in [r"valor total da divida parcelada\s*[:\s]*([\d.,]+)",
                    r"total a pagar\s*[:\s]*([\d.,]+)",
                    r"valor atualizado\s*[:\s]*([\d.,]+)",
                    r"valor atual\s*[:\s]*([\d.,]+)",
                    r"total em r\$\s*[:\s]*([\d.,]+)",
                    r"vl\.?\s*corrigido\s*[:\s]*([\d.,]+)",
                    r"total\s*geral\s*[-:>\s]*([\d.,]+)"]:
            m = re.search(pat, tn)
            if m:
                valor_atualizado = f"R$ {m.group(1).strip()}"
                break
 
    if valor_original is None:
        m = re.search(r"r\$\s*([\d.,]+)", tn)
        if m:
            valor_original = f"R$ {m.group(1).strip()}"
 
    return valor_original, valor_atualizado

def _extrair_tipo_tributo(text):
    """Extrae el tipo de tributo de la CDA o del Classe-Assunto."""
    _TRIBUTOS = [
        ("imposto predial territorial urbano",  "IPTU — Imposto Predial e Territorial Urbano"),
        ("imposto predial e territorial urbano", "IPTU — Imposto Predial e Territorial Urbano"),
        ("taxa de licenciamento",               "Taxa de Licenciamento de Estabelecimento"),
        ("taxa de fiscalizacao de funcionamento","TFF — Taxa de Fiscalização de Funcionamento"),
        ("tff",                                  "TFF — Taxa de Fiscalização de Funcionamento"),
        ("cosip",                                "COSIP — Contribuição de Iluminação Pública"),
        ("contribuicao de iluminacao publica",   "COSIP — Contribuição de Iluminação Pública"),
        ("imposto sobre servicos",               "ISS — Imposto sobre Serviços"),
        ("iss",                                  "ISS — Imposto sobre Serviços"),
        ("itbi",                                 "ITBI — Imposto sobre Transmissão de Bens Imóveis"),
        ("imposto sobre transmissao",            "ITBI — Imposto sobre Transmissão de Bens Imóveis"),
        ("iptu",                                 "IPTU — Imposto Predial e Territorial Urbano"),
        ("multa",                                "Multa"),
    ]
    _FRAGMENTOS_IGNORAR = [
        "divida ativa",
        "divida municipal",
        "credito tributario",
        "execucao fiscal",
        "devedor",
        "requerendo",
        "credor",
        "municipio de salvador reu",
        "parte ativa",
    ]
    _RUIDO_LEGAL = ["art.", "lei n", "lei no", "inciso", "paragrafo", "ans."]

    text_norm = normalizar(text)

    for pat in [
        r"esp[eé]cie\s*[:\s]+([^\n\r]{3,60})",
        r"tributo\s*[:\s]+([^\n\r]{3,60})",
        r"classe\s*[-]?\s*assunto\s*[:\s]+([^\n\r]{3,80})",
        r"execu[cç][aã]o fiscal\s*[-\s/]+([^\n\r]{3,30})",
    ]:
        m = re.search(pat, text_norm)
        if m:
            fragmento = normalizar(m.group(1).strip())
            if any(ig in fragmento for ig in _FRAGMENTOS_IGNORAR):
                continue
            for kw, label in _TRIBUTOS:
                if kw in fragmento:
                    return label
            if not any(r in fragmento for r in _RUIDO_LEGAL) and len(fragmento) > 3:
                return fragmento[:60].strip()

    for kw, label in _TRIBUTOS:
        if kw in text_norm:
            return label

    return None


def _extrair_numero_cda(text):
    """Extrae el número de la CDA. Descarta valores < 5 dígitos (evita CGA)."""
    _PATRONES = [
        r"cda\s*n[o°º.]?\s*([\d/.\-]+)",
        r"certid[aã]o de d[ií]vida ativa\s*[-\s]*n[o°º.]?\s*([\d/.\-]+)",
        r"n[o°º.]?\s+da\s+cda\s*[:\s]*([\d/.\-]+)",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            valor = m.group(1).strip().rstrip(".")
            if len(re.sub(r"\D", "", valor)) >= 5:
                return valor
    return None


def _extrair_numero_processo(text):
    """Extrae el número CNJ (NNNNNNN-DD.AAAA.J.TT.OOOO)."""
    m = re.search(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}", text)
    if m:
        return m.group(0)
    return None


def _extrair_vara(text):
    """Extrae la vara/juízo donde tramita el processo."""
    _PATRONES = [
        r"(\d+[aª°]\s*vara\s*da\s*fazenda\s*p[uú]blica[^\n\r]{0,40})",
        r"([oó]rg[aã]o julgador\s*[:\s]+[^\n\r]{5,60})",
        r"(vara\s*[^\n\r]{3,40})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:80]
    return None


def _extrair_exercicio(text):
    """Extrae el año o período fiscal del tributo (prioriza tabla de la CDA)."""
    text_norm = normalizar(text)

    m_cda = re.search(
        r"(?:esp[eé]cie|tributo)[^\n]{0,80}exerc[ií]cio\s*[:\s]*(\d{4}(?:[/]\d{4})?)",
        text_norm
    )
    if not m_cda:
        m_cda = re.search(
            r"exerc[ií]cio\s*[:\s]*(\d{4}(?:[/]\d{4})?)\s+(?:meses|cotas|valor)",
            text_norm
        )
    if not m_cda:
        m_b = re.search(
            r"exerc[ií]cio\s+meses[^\n]*\n\s*(\d{4}(?:[/]\d{4})?)",
            text_norm
        )
        if m_b:
            bloque_tab = text_norm[m_b.start():]
            anos_tab = re.findall(r"\b(20[012]\d)\s+\d{1,2}\s+[\d.,]+", bloque_tab)
            if anos_tab:
                anos_s = sorted(set(anos_tab))
                return anos_s[0] if len(anos_s) == 1 else f"{anos_s[0]}/{anos_s[-1]}"
            return m_b.group(1).strip()
    if not m_cda:
        m_cda = re.search(
            r"exerc[ií]cio\s*:\s*(\d{4})\s*\n\s*(?:meses|cotas)",
            text_norm
        )
    if m_cda:
        return m_cda.group(1).strip()

    m = re.search(r"exerc[ií]cio\s*[:\s]*(\d{4}(?:[/\-]\d{4})?)", text_norm)
    if m:
        return m.group(1).strip().replace("-", "/")

    m2 = re.search(r"exerc[ií]cios?\s+(?:de\s+)?(\d{4})\s+e\s+(\d{4})", text_norm)
    if m2:
        return f"{m2.group(1)}/{m2.group(2)}"

    todos = re.findall(r"exerc[ií]cio\s+(\d{4})", text_norm)
    if todos:
        anos = sorted(set(todos))
        return anos[0] if len(anos) == 1 else f"{anos[0]}/{anos[-1]}"

    m_tabla = re.search(
        r"(20[012]\d)\s+\d{1,2}\s+[\d.,]+",
        text_norm
    )
    if m_tabla:
        todos_tabla = re.findall(r"\b(20[012]\d)\s+\d{1,2}\s+[\d.,]+", text_norm)
        if todos_tabla:
            anos = sorted(set(todos_tabla))
            return anos[0] if len(anos) == 1 else f"{anos[0]}/{anos[-1]}"

    return None


def _extrair_data_inscricao(text):
    """Extrae la fecha de inscripción en la dívida ativa (valida formato)."""
    _PATRONES = [
        r"data\s*d[ae]\s*inscri[cç][aã]o\s*[:\s]*([\d/\.\-]+)",
        r"inscri[cç][aã]o\s+na\s+d[ií]vida\s+ativa\s*[:\s]*([\d/\.\-]+)",
        r"data\s+de\s+emiss[aã]o\s*[:\s]*([\d/\.\-]+)",
    ]
    _PAT_FECHA_VALIDA = re.compile(r"^\d{1,2}[/\.\-]\d{1,2}[/\.\-]\d{2,4}$")
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            candidato = m.group(1).strip()
            if _PAT_FECHA_VALIDA.match(candidato):
                return candidato
    return None


def extract_entidades_agente2(text):
    """
    Punto de entrada de extracción de entidades. Devuelve dict con todos los
    campos; los no encontrados quedan en None (nunca se inventa información).
    """
    campos = [
        "cpf_cnpj", "nome_executado", "nome_exequente",
        "valor_original", "valor_atualizado", "tipo_tributo",
        "numero_cda", "numero_processo", "vara",
        "exercicio", "data_inscricao",
    ]
    if not text:
        return {k: None for k in campos}

    valor_original, valor_atualizado = _extrair_valor(text)

    return {
        "cpf_cnpj"        : _extrair_cpf_cnpj(text),
        "nome_executado"  : _extrair_nome_executado(text),
        "nome_exequente"  : _extrair_nome_exequente(text),
        "valor_original"  : valor_original,
        "valor_atualizado": valor_atualizado,
        "tipo_tributo"    : _extrair_tipo_tributo(text),
        "numero_cda"      : _extrair_numero_cda(text),
        "numero_processo" : _extrair_numero_processo(text),
        "vara"            : _extrair_vara(text),
        "exercicio"       : _extrair_exercicio(text),
        "data_inscricao"  : _extrair_data_inscricao(text),
    }


def extract_parcelamento(text):
    text_norm = normalizar(text)
    if not any(normalizar(k) in text_norm for k in KEYWORDS_PARCELAMENTO_ATIVO):
        return None

    KEYWORDS_ART40_EXCLUSIVOS = [
        "art. 40 da lef",
        "art. 40 da lei 6.830",
        "art. 40 - o juiz suspendera",
        "enquanto nao for localizado o devedor",
        "nao for localizado o devedor",
        "ausencia de localizacao do executado",
        "suspensao do feito pelo prazo de um ano",
        "suspensao pelo prazo de 1 (um) ano",
    ]
    KEYWORDS_PAD_ESPECIFICOS = [
        "art. 151",
        "parcelamento do credito tributario",
        "parcelamento do debito",
        "instrumento de confissao",
        "compromisso de pagamento parcelado",
        "bloq pad",
        "pad n",
        "parcelas mensais e sucessivas",
        "em 38 parcelas",
        "em 12 parcelas",
        "em 24 parcelas",
        "em 36 parcelas",
        "em 48 parcelas",
        "em 60 parcelas",
        "defiro o pedido de suspensao",
        "suspendo/mantenho suspenso",
    ]
    is_art40_context = any(normalizar(k) in text_norm for k in KEYWORDS_ART40_EXCLUSIVOS)
    has_pad_markers  = any(normalizar(k) in text_norm for k in KEYWORDS_PAD_ESPECIFICOS)
    if is_art40_context and not has_pad_markers:
        return None   # art.40 LEF — não confundir com PAD

    KEYWORDS_PAD_HOMOLOGADO = [
        "situacao do parcelamento: homologado",
        "situacao pad: homologado",
        "situacao: homologado",
        "bloq pad",
        "suspendo/mantenho suspenso",
        "suspendo e mantenho suspenso",
    ]
    if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_HOMOLOGADO):
        return "processo suspenso por parcelamento (PAD)"

    KEYWORDS_PAD_ROMPIDO = [
        "situacao do parcelamento: rompido",
        "situacao: rompido",
        "parcelamento rompido",
        "parcelamento cancelado em",
        "cred. ref. ao cancel. do parc",
        "cred. ref. ao cancel. do pad",
        "motivo: pagamento em atraso",
        "data de rompimento:",
        "requerer a citacao da parte executada no seguinte endereco",
        "requer a citacao da parte executada no seguinte endereco",
        "citacao da parte executada no seguinte endereco",
    ]
    if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_ROMPIDO):
        KEYWORDS_PAD_NOVO_ATIVO = [
            "defiro o pedido de suspensao",
            "determino a suspensao do feito",
            "suspendo o feito",
            "suspendo/mantenho suspenso",
            "suspendo e mantenho suspenso",
            "situacao pad: homologado",
            "bloq pad",
        ]
        if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_NOVO_ATIVO):
            return "processo suspenso por parcelamento (PAD)"
        return None

    return "processo suspenso por parcelamento (PAD)"


# ===========================================================================
# HELPERS DE SÍNTESE (usados por Excel, JSON e histórico — fonte única)
# ===========================================================================

def _extrair_sinais_processuais(full_text):
    """
    Reúne os sinais processuais detectados (fatos, não veredito). Fonte única
    usada pela planilha, pelo JSON e pelo histórico para evitar divergência.
    """
    return {
        "extincao"       : extract_extincao(full_text),
        "parcelamento"   : extract_parcelamento(full_text),
        "suspensao_art40": extract_suspensao_art40(full_text),
    }


def _dias_desde(fecha, hoy=None):
    """Días transcurridos desde `fecha` (dato neutral, no es un veredito)."""
    if fecha is None:
        return None
    hoy = hoy or datetime.now()
    return (hoy - fecha).days


# ===========================================================================
# LLM / GEMINI — NOTA DE ARQUITETURA
# ===========================================================================
# O Agente 1 é DETERMINÍSTICO e OFFLINE: usa apenas regex + OCR local. Não faz
# nenhuma chamada a modelo de linguagem, não precisa de chave de API nem de
# acesso à internet em tempo de execução (ver "network_mode: none" no compose).
#
# A etapa de raciocínio jurídico assistido por LLM foi movida para o AGENTE 2.
# É lá que ficará o bloco isolado e claramente marcado:
#
#     # [GEMINI PLACEHOLDER]   (migração OpenAI -> Gemini localizada; a
#                              estrutura de entrada/saída das funções não muda,
#                              só a fonte da resposta.)
#
# Consequência prática p/ LGPD: o container do Agente 1 não embarca SDK de LLM
# nem chave, reduzindo a superfície de dados que sai do servidor da PGMS.
# ===========================================================================


# 6. Resumo informativo (substitui o antigo prompt de GPT; não é veredito)
RESUMO_TEMPLATE = """RESUMO DOS SINAIS DETECTADOS (Agente 1 — extração automática)
Informativo. O Agente 1 NÃO emite juízo APTO/NÃO APTO; apenas reporta o que
foi extraído do processo para revisão humana / Agente 2.

- Última movimentação detectada: {fecha}
- Status da citação detectado:    {citacion}
- Resultado da penhora detectado: {penhora}
"""

def create_prompt(fecha_reciente, citacion, penhora):
    return RESUMO_TEMPLATE.format(
        fecha    = fecha_reciente.strftime("%d/%m/%Y") if fecha_reciente else "Não especificado",
        citacion = citacion or "Não especificado",
        penhora  = penhora or "Não especificado",
    )


# 7. Generar registros para todos los archivos PDF
def generate_prompts(input_dir):
    pdf_files = [f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')]
    prompts = []

    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_dir, pdf_file)

        full_text      = ""
        fecha_reciente = None
        citacion       = None
        penhora        = None
        fecha_orden    = None
        fecha_intento  = None
        fecha_efectiva = None
        prompt         = "Error"
        respuesta_gpt  = None
        ocr_metadata   = {"paginas_ocr": [], "confianza_ocr": {}}
        tipo_processo  = None
        entidades      = None

        try:
            pages_text, ocr_metadata = extract_text_by_page(pdf_path)
            full_text  = " ".join(p for p in pages_text if p)

            if ocr_metadata["paginas_ocr"]:
                n = len(ocr_metadata["paginas_ocr"])
                conf_prom = sum(ocr_metadata["confianza_ocr"].values()) / n
                logging.info(f"  {pdf_file}: {n} página(s) vía OCR, confianza media {conf_prom:.1f}%")

            if not full_text.strip():
                logging.warning(f"PDF sin texto extraíble: {pdf_file}")
                prompts.append((pdf_file, None, None, None, None, None, None,
                                "Erro - PDF sem texto", None, "", ocr_metadata, tipo_processo, None))
                continue

            tipo_processo = detectar_tipo_processo(full_text)
            if (not tipo_processo["es_execucao_fiscal"]) and tipo_processo["confianza"] == "alta":
                logging.warning(f"  {pdf_file}: FORA DE ESCOPO — {tipo_processo['motivo']}")
                prompts.append((
                    pdf_file, None, None, None, None, None, None,
                    "Fora de escopo - não é execução fiscal", None, full_text, ocr_metadata, tipo_processo, None
                ))
                continue

            fecha_reciente = fecha_ultima_movimentacao(full_text)
            citacion       = extract_citacion(full_text)
            penhora        = extract_penhora(full_text)
            fecha_orden, fecha_intento, fecha_efectiva = extraer_fechas_citacion(full_text)

            # Regla: fecha_citacion_efectiva None si status_citacion != "HOUVE CITAÇÃO"
            if not citacion or normalizar(citacion) != normalizar("HOUVE CITAÇÃO"):
                fecha_efectiva = None

            entidades = extract_entidades_agente2(full_text)
            prompt = create_prompt(fecha_reciente, citacion, penhora)

            print(f"\n{'='*60}")
            print(f"ARQUIVO : {pdf_file}")
            print(f"  Última data    : {fecha_reciente.strftime('%Y-%m-%d') if fecha_reciente else 'NÃO ENCONTRADA'}")
            print(f"  Citação        : {citacion}")
            print(f"  Penhora        : {penhora}")
            print(f"  CPF/CNPJ       : {entidades.get('cpf_cnpj') or '—'}")
            print(f"  Executado      : {entidades.get('nome_executado') or '—'}")
            print(f"  Valor orig.    : {entidades.get('valor_original') or '—'}")
            print(f"{'='*60}")

            prompts.append((
                pdf_file, fecha_reciente, citacion,
                fecha_orden, fecha_intento, fecha_efectiva,
                penhora, prompt, respuesta_gpt, full_text, ocr_metadata, tipo_processo, entidades
            ))

        except Exception as e:
            logging.error(f"Erro ao processar {pdf_file}: {e}", exc_info=True)
            prompts.append((
                pdf_file, None, None, None, None, None, None, "Erro", None, full_text, ocr_metadata, tipo_processo, None
            ))

    return prompts


# 8. Guardar resumo de sinais en un archivo de texto (auditoría legible)
def save_prompts_to_file(prompts, output_file):
    with open(output_file, 'w', encoding='utf-8') as file:
        for (pdf_file, _, _, _, _, _, _, prompt, _, _, ocr_metadata, tipo_processo, _) in prompts:
            paginas_ocr = ocr_metadata.get("paginas_ocr", []) if ocr_metadata else []
            ocr_info = f"Páginas via OCR: {paginas_ocr}\n" if paginas_ocr else ""
            tipo_info = f"Tipo de processo: {tipo_processo['motivo']}\n" if tipo_processo else ""
            file.write(f"Arquivo: {pdf_file}\n{tipo_info}{ocr_info}{prompt}\n{'-'*50}\n")
    print(f"Resumo de sinais salvo em {output_file}")


# 9. Procesar registros y generar Excel (extração, SEM veredito APTO/NÃO APTO)
def process_prompts_to_excel(prompts, output_excel):
    import pandas as pd
    resultados = []
    hoy = datetime.now()

    for (pdf_file, fecha_reciente, citacion, fecha_orden, fecha_intento,
         fecha_efectiva, penhora, prompt, respuesta_gpt, full_text,
         ocr_metadata, tipo_processo, entidades) in prompts:

        sinais = _extrair_sinais_processuais(full_text)
        dias = _dias_desde(fecha_reciente, hoy)

        paginas_ocr = ocr_metadata.get("paginas_ocr", []) if ocr_metadata else []
        confianza_ocr_dict = ocr_metadata.get("confianza_ocr", {}) if ocr_metadata else {}
        confianza_ocr_media = (
            round(sum(confianza_ocr_dict.values()) / len(confianza_ocr_dict), 1)
            if confianza_ocr_dict else None
        )
        ent = entidades or {}
        tp  = tipo_processo or {}

        resultados.append({
            # ── Identificação / tipo ──────────────────────────────────────
            "CASO"                          : pdf_file,
            "Número do processo"            : ent.get("numero_processo") or "",
            "É execução fiscal?"            : ("Sim" if tp.get("es_execucao_fiscal") else "Não") if tp else "",
            "Confiança tipo processo"       : tp.get("confianza", "") if tp else "",
            "Classe-Assunto (PJe)"          : (tp.get("classe_assunto") or "(não detectado)") if tp else "",
            # ── Fatos processuais (extraídos, sem veredito) ───────────────
            "Última data de interação"      : fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else "Não especificado",
            "Dias desde última movimentação": dias if dias is not None else "",
            "Status da citação"             : citacion or "Não especificado",
            "Data ordem citação"            : fecha_orden.strftime("%Y-%m-%d")   if fecha_orden    else "Não especificado",
            "Data tentativa citação"        : fecha_intento.strftime("%Y-%m-%d") if fecha_intento  else "Não especificado",
            "Data citação efetiva"          : fecha_efectiva.strftime("%Y-%m-%d") if fecha_efectiva else "Não especificado",
            "Resultado da penhora"          : penhora or "Não especificado",
            "Extinção detectada"            : sinais["extincao"] or "",
            "Parcelamento detectado"        : sinais["parcelamento"] or "",
            "Suspensão art.40 LEF"          : sinais["suspensao_art40"] or "",
            # ── OCR ────────────────────────────────────────────────────────
            "Páginas via OCR"               : ", ".join(map(str, paginas_ocr)) if paginas_ocr else "",
            "Confiança OCR (%)"             : confianza_ocr_media if confianza_ocr_media is not None else "",
            # ── Entidades ──────────────────────────────────────────────────
            "CPF/CNPJ"                      : ent.get("cpf_cnpj") or "",
            "Nome executado"                : ent.get("nome_executado") or "",
            "Nome exequente"                : ent.get("nome_exequente") or "",
            "Tipo de tributo"               : ent.get("tipo_tributo") or "",
            "Exercício"                     : ent.get("exercicio") or "",
            "Número CDA"                    : ent.get("numero_cda") or "",
            "Data inscrição dívida ativa"   : ent.get("data_inscricao") or "",
            "Valor original"                : ent.get("valor_original") or "",
            "Valor atualizado"              : ent.get("valor_atualizado") or "",
            "Vara"                          : ent.get("vara") or "",
        })

    df = pd.DataFrame(resultados)

    cols_texto = ["Número CDA", "CPF/CNPJ", "Número do processo"]
    for col in cols_texto:
        if col in df.columns:
            df[col] = df[col].astype(str).replace({"nan": "", "None": ""})

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extração Agente 1")
        ws = writer.sheets["Extração Agente 1"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        font_header  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        fill_header  = PatternFill("solid", fgColor="1F4E79")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde        = Border(*[Side(style="thin", color="D0D0D0")] * 4)
        fill_fora    = PatternFill("solid", fgColor="D9D9D9")  # cinza: fora de escopo

        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = font_header; c.fill = fill_header
            c.alignment = align_header; c.border = borde

        col_fiscal = (list(df.columns).index("É execução fiscal?") + 1) if "É execução fiscal?" in df.columns else None
        font_dato  = Font(name="Arial", size=10)
        align_dato = Alignment(vertical="top", wrap_text=True)
        for row_idx in range(2, len(df) + 2):
            fora = bool(col_fiscal) and ws.cell(row=row_idx, column=col_fiscal).value == "Não"
            for col_idx in range(1, len(df.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = font_dato; c.border = borde; c.alignment = align_dato
                if fora:
                    c.fill = fill_fora

        for col_name in cols_texto:
            if col_name in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
                for cell in ws[col_letter][1:]:
                    if cell.value:
                        cell.value = str(cell.value)
                        cell.data_type = "s"
                        cell.number_format = "@"

        anchos = {
            "CASO": 34, "Número do processo": 22, "É execução fiscal?": 12,
            "Confiança tipo processo": 12, "Classe-Assunto (PJe)": 30,
            "Última data de interação": 16, "Dias desde última movimentação": 14,
            "Status da citação": 30, "Data ordem citação": 14,
            "Data tentativa citação": 14, "Data citação efetiva": 14,
            "Resultado da penhora": 30, "Extinção detectada": 22,
            "Parcelamento detectado": 28, "Suspensão art.40 LEF": 22,
            "Páginas via OCR": 14, "Confiança OCR (%)": 12,
            "CPF/CNPJ": 20, "Nome executado": 30, "Nome exequente": 30,
            "Tipo de tributo": 20, "Exercício": 12, "Número CDA": 22,
            "Data inscrição dívida ativa": 16, "Valor original": 16,
            "Valor atualizado": 16, "Vara": 28,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(col_name, 18)

        ws.freeze_panes = "A2"
        if len(df.columns):
            ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

    print(f"Planilha Excel gerada: {output_excel}")


# ===========================================================================
# 9.1 — JSON ESTRUTURADO (interface Agente 1 -> Agente 2)
# ===========================================================================

def exportar_json_agente2(prompts, output_json):
    """
    Gera o JSON de interface Agente 1 -> Agente 2.

    [feedback 1] Inclui TODOS os campos que aparecem na planilha.
    [feedback 4] Inclui TODOS os processos (sem filtro por decisão) — o
    Agente 1 não emite mais APTO/NÃO APTO.

    Campos vazios são gravados como null — nunca omitidos.
    """
    import json
    from datetime import datetime as _dt

    VERSION_AGENTE1 = "8.0"

    def _nulo(val):
        if val is None:
            return None
        v = str(val).strip()
        return None if v in ("", "nan", "None", "Não especificado") else v

    hoy = _dt.now()
    processos = []

    for tupla in prompts:
        (pdf_file, fecha_reciente, citacion, fecha_orden, fecha_intento,
         fecha_efectiva, penhora, prompt, respuesta_gpt, full_text,
         ocr_metadata, tipo_processo, entidades) = tupla

        ent = entidades or {}
        tp  = tipo_processo or {}
        sinais = _extrair_sinais_processuais(full_text)

        ocr_meta  = ocr_metadata or {}
        conf_dict = ocr_meta.get("confianza_ocr", {})
        conf_media = round(sum(conf_dict.values()) / len(conf_dict), 1) if conf_dict else None

        processos.append({
            "id_lote"                       : pdf_file,
            "numero_processo"               : _nulo(ent.get("numero_processo")),
            "tipo_processo": {
                "es_execucao_fiscal"        : tp.get("es_execucao_fiscal"),
                "confianca"                 : tp.get("confianza"),
                "classe_assunto"            : _nulo(tp.get("classe_assunto")),
            },
            "ultima_movimentacao"           : fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else None,
            "dias_desde_ultima_movimentacao": _dias_desde(fecha_reciente, hoy),
            "status_citacao"                : _nulo(citacion),
            "data_ordem_citacao"            : fecha_orden.strftime("%Y-%m-%d")   if fecha_orden    else None,
            "data_tentativa_citacao"        : fecha_intento.strftime("%Y-%m-%d") if fecha_intento  else None,
            "data_citacao_efetiva"          : fecha_efectiva.strftime("%Y-%m-%d") if fecha_efectiva else None,
            "resultado_penhora"             : _nulo(penhora),
            "sinais_processuais": {
                "extincao"                  : _nulo(sinais["extincao"]),
                "parcelamento"              : _nulo(sinais["parcelamento"]),
                "suspensao_art40_lef"       : _nulo(sinais["suspensao_art40"]),
            },
            "ocr": {
                "paginas"                   : ocr_meta.get("paginas_ocr", []),
                "confianca_media"           : conf_media,
            },
            "entidades": {
                # numero_processo também dentro de 'entidades' (além do nível
                # superior) — é onde buscar_processo.py e o Agente 2 procuram.
                "numero_processo"           : _nulo(ent.get("numero_processo")),
                "cpf_cnpj"                  : _nulo(ent.get("cpf_cnpj")),
                "nome_executado"            : _nulo(ent.get("nome_executado")),
                "nome_exequente"            : _nulo(ent.get("nome_exequente")),
                "tipo_tributo"              : _nulo(ent.get("tipo_tributo")),
                "exercicio"                 : _nulo(ent.get("exercicio")),
                "numero_cda"                : _nulo(ent.get("numero_cda")),
                "data_inscricao"            : _nulo(ent.get("data_inscricao")),
                "valor_original"            : _nulo(ent.get("valor_original")),
                "valor_atualizado"          : _nulo(ent.get("valor_atualizado")),
                "vara"                      : _nulo(ent.get("vara")),
            },
        })

    payload = {
        "metadata": {
            "generado_em"     : hoy.strftime("%Y-%m-%dT%H:%M:%S"),
            "total_procesados": len(prompts),
            "version_agente1" : VERSION_AGENTE1,
            "observacao"      : "Agente 1 faz apenas extração determinística; não emite juízo APTO/NÃO APTO.",
        },
        "processos": processos,
    }

    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"JSON do Agente 1 gerado: {output_json}")
    print(f"  {len(processos)} processo(s) exportado(s) (todos, sem filtro).")
    return payload


# ===========================================================================
# 9.2 — HISTÓRICO DE EXTRAÇÕES (auditoria acumulativa — append-only)
# ===========================================================================

def exportar_historial_classificacoes(prompts, output_jsonl):
    """
    Registra TODOS os processos do lote num JSONL — uma linha por processo.
    APPEND-ONLY: nunca sobrescreve o rastro anterior. Cada processo vai em
    try/except: um registro com erro é logado e pulado, sem derrubar o lote.
    """
    import json
    from datetime import datetime as _dt

    def _nulo(val):
        if val is None:
            return None
        v = str(val).strip()
        return None if v in ("", "nan", "None", "Não especificado") else v

    hoy = _dt.now()
    gravados, erros = 0, 0

    with open(output_jsonl, "a", encoding="utf-8") as f:
        for tupla in prompts:
            try:
                (pdf_file, fecha_reciente, citacion, fecha_orden, fecha_intento,
                 fecha_efectiva, penhora, prompt, respuesta_gpt, full_text,
                 ocr_metadata, tipo_processo, entidades) = tupla

                ent    = entidades or {}
                tp     = tipo_processo or {}
                sinais = _extrair_sinais_processuais(full_text)

                registro = {
                    "extraido_em"        : hoy.strftime("%Y-%m-%dT%H:%M:%S"),
                    "id_lote"            : pdf_file,
                    "numero_processo"    : _nulo(ent.get("numero_processo")),
                    "es_execucao_fiscal" : tp.get("es_execucao_fiscal"),
                    "ultima_movimentacao": fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else None,
                    "status_citacao"     : _nulo(citacion),
                    "resultado_penhora"  : _nulo(penhora),
                    "extincao"           : _nulo(sinais["extincao"]),
                    "parcelamento"       : _nulo(sinais["parcelamento"]),
                    "suspensao_art40_lef": _nulo(sinais["suspensao_art40"]),
                    "nome_executado"     : _nulo(ent.get("nome_executado")),
                    "cpf_cnpj"           : _nulo(ent.get("cpf_cnpj")),
                }
                f.write(json.dumps(registro, ensure_ascii=False) + "\n")
                gravados += 1
            except Exception as e:
                erros += 1
                _id = tupla[0] if isinstance(tupla, (list, tuple)) and tupla else "desconhecido"
                print(f"[HISTORICO] ERRO ao registrar '{_id}': {e}")

    print(f"Histórico de extrações (append-only): {output_jsonl}")
    print(f"  {gravados} registro(s) gravado(s), {erros} com erro")
    return gravados, erros


# 10. Ejecutar el flujo
if __name__ == "__main__":
    os.makedirs(PASTA_JSON, exist_ok=True)
    os.makedirs(PASTA_RESULTADOS, exist_ok=True)

    _timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    output_file_resumo    = os.path.join(PASTA_JSON, f"resumo_sinais_V8_{_timestamp}.txt")
    output_file_excel     = os.path.join(PASTA_RESULTADOS, f"resultados_V8_{_timestamp}.xlsx")
    output_file_json      = os.path.join(PASTA_JSON, f"saida_agente1_V8_{_timestamp}.json")
    output_file_historico = os.path.join(PASTA_JSON, "historico_extracoes.jsonl")  # append-only

    prompts = generate_prompts(input_directory)
    save_prompts_to_file(prompts, output_file_resumo)
    process_prompts_to_excel(prompts, output_file_excel)
    exportar_json_agente2(prompts, output_file_json)
    exportar_historial_classificacoes(prompts, output_file_historico)