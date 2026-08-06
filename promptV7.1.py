import os
import re
import unicodedata
import pdfplumber
import pandas as pd
from datetime import datetime
import logging
from openai import OpenAI

# --- OCR para páginas escaneadas (sin texto digital) ---
import pytesseract
from pytesseract import Output
from pdf2image import convert_from_path
from concurrent.futures import ThreadPoolExecutor, as_completed


# Función para normalizar texto (remover acentos, convertir a minúsculas)
def normalizar(text):
    if not text:
        return ""
    return unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII').lower()

# Función para extraer fechas de un texto normalizado
def _extraer_fechas_de_texto(text_norm):
    # Patrones para extraer fechas (tanto con meses completos como abreviados)
    _MESES_COMPLETOS = {
        "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
        "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
        "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
    }
    _MESES_ABREV = {
        "jan": 1, "fev": 2, "mar": 3, "abr": 4,
        "mai": 5, "jun": 6, "jul": 7, "ago": 8,
        "set": 9, "out": 10, "nov": 11, "dez": 12
    }
    _PATRON_COMPLETO = r"(\d{1,2}) de (janeiro|fevereiro|marco|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro) de (\d{4})"
    _PATRON_ABREV    = r"(\d{1,2})\s+(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\.?\s+(\d{4})"
    fechas = []
    for dia, mes_txt, anio in re.findall(_PATRON_COMPLETO, text_norm):
        fechas.append(datetime(int(anio), _MESES_COMPLETOS[mes_txt], int(dia)))
    for dia, mes_txt, anio in re.findall(_PATRON_ABREV, text_norm):
        if mes_txt in _MESES_ABREV:
            fechas.append(datetime(int(anio), _MESES_ABREV[mes_txt], int(dia)))
    return fechas


# Esto silencia los logs internos de la librería que genera esos mensajes
logging.getLogger('pdfminer').setLevel(logging.ERROR)
logging.basicConfig(level=logging.INFO)

# Directorio de entrada y salida
from datetime import datetime  # asegurate de tener este import arriba

CURRENT_DIR      = os.path.dirname(os.path.abspath(__file__))

# Las tres carpetas son sobreescribibles por variable de entorno para que la API
# pueda dar a cada lote su propio espacio aislado — sin eso, dos lotes en paralelo
# se pisarían el JSON de traspaso, que tiene nombre fijo. Sin las variables el
# comportamiento es el de siempre: subcarpetas junto al script.
input_directory  = os.environ.get("PASTA_ENTRADA")    or os.path.join(CURRENT_DIR, "processos pra analiser")

# Carpetas de salida
PASTA_JSON       = os.environ.get("PASTA_JSON")       or os.path.join(CURRENT_DIR, "JSON")         # JSON + prompts que escribe el Agente 1
PASTA_RESULTADOS = os.environ.get("PASTA_RESULTADOS") or os.path.join(CURRENT_DIR, "resultados")   # Excel de resultados

os.makedirs(PASTA_JSON, exist_ok=True)
os.makedirs(PASTA_RESULTADOS, exist_ok=True)

# Timestamp compartido por los archivos de esta corrida (formato seguro para nombres de archivo)
_timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")

# Archivos de salida
output_file_prompts = os.path.join(PASTA_JSON, "prompts_generadosV7_1.txt")
output_file_excel   = os.path.join(PASTA_RESULTADOS, f"resultados_V7_1_{_timestamp}.xlsx")
output_file_json    = os.path.join(PASTA_JSON, "resultados_procesosV7_1_agente2.json")



# Plantilla resumida (usada solo para guardar en el .txt)
PROMPT_TEMPLATE = """Você é um assistente jurídico especializado em execuções fiscais brasileiras.
    O sistema automático não conseguiu classificar este processo com certeza. Analise os dados abaixo e tome a decisão final.

    DADOS DO PROCESSO:
    - Última movimentação: {fecha}
    - Status da citação: {citacion}
    - Resultado da penhora: {penhora}

    CRITÉRIOS DE CLASSIFICAÇÃO:

    1. Se a última movimentação ocorreu há MENOS de 1 ano → NÃO APTO
    2. Se a última movimentação ocorreu há MAIS de 1 ano, avalie citação e penhora:
        a) CITAÇÃO AUSENTE ou FALHA → APTO
        b) CITAÇÃO VÁLIDA → avalie penhora:
            - Sem penhora / tentativa sem resultado → APTO
            - Penhora EFETIVADA (bacenjud/sisbajud, imóvel, faturamento, quotas, créditos, RENAJUD, CNIB) → NÃO APTO
    3. Se os dados forem insuficientes → INFORMAÇÃO INSUFICIENTE

    FORMATO DE RESPOSTA (responda APENAS neste formato, sem texto adicional):
    DECISÃO: <APTO | NÃO APTO | INFORMAÇÃO INSUFICIENTE>
    MOTIVO: <justificativa em uma frase>
    """
# Template para GPT com texto completo do PDF   
PROMPT_TEMPLATE_FULL = """Você é um assistente jurídico especializado em execuções fiscais brasileiras.
    O sistema automático não conseguiu classificar este processo com certeza.
    Analise o TEXTO COMPLETO do processo abaixo e determine a classificação.

    CRITÉRIOS:
    1. Última movimentação há MENOS de 1 ano → NÃO APTO
    2. Última movimentação há MAIS de 1 ano:
        a) Executado NÃO foi citado (ou citação falhou) → APTO
        b) Citação VÁLIDA → avalie penhora:
            - Sem penhora / apenas tentativa → APTO
            - Penhora EFETIVADA (BacenJud/SisBajud com saldo bloqueado, imóvel, 
              faturamento, quotas, RENAJUD, CNIB) → NÃO APTO
    3. Dados insuficientes → INFORMAÇÃO INSUFICIENTE

    ⚠️ ATENÇÃO — DISTINÇÃO CRÍTICA SOBRE SISBAJUD/BACENJUD:
    - Uma PETIÇÃO do exequente SOLICITANDO o bloqueio via SISBAJUD/BacenJud NÃO é penhora efetivada.
    - Para ser NÃO APTO, deve existir no processo um DOCUMENTO DE RESULTADO: 
      extrato de bloqueio, comprovante de valores bloqueados, resposta dos bancos, 
      ou despacho confirmando o bloqueio com saldo.
    - Se só há petição solicitando, sem documento de resultado → classifique como APTO.

    DADOS EXTRAÍDOS PELO SISTEMA (use como referência, mas confie no texto original):
    - Última movimentação detectada: {fecha}
    - Status da citação detectado: {citacion}
    - Status da penhora detectado: {penhora}

    TEXTO COMPLETO DO PROCESSO:
    {full_text}

    FORMATO DE RESPOSTA (responda APENAS neste formato, sem texto adicional):
    DECISÃO: <APTO | NÃO APTO | INFORMAÇÃO INSUFICIENTE>
    MOTIVO: <justificativa em uma frase>
    STATUS CITAÇÃO: <resumo do que encontrou sobre citação>
    STATUS PENHORA: <resumo do que encontrou sobre penhora>
    """

# --- Configuración de OCR (fallback para páginas escaneadas) ---
OCR_DPI = 300
# [AJUSTABLE] DPI más bajo = OCR más rápido. Medido: 300dpi≈1.27s/página,
# 200dpi≈0.73s/página, 150dpi≈0.52s/página, mismo texto correcto en
# pruebas sintéticas. Antes de bajarlo en producción, comparar la
# "Confiança OCR (%)" en el Excel contra una muestra de páginas reales
# escaneadas (no las sintéticas usadas para validar esto) — documentos
# con sellos/firmas/ruído de escaneo pueden perder precisión a menor DPI
# de un modo que el texto sintético no refleja.
OCR_MIN_CHARS = 20      # umbral empírico — ajustable según casos reales
OCR_IDIOMA = 'por'

# [NUEVO] Agrupamiento de páginas OCR en lotes, para reducir la cantidad
# de llamadas a Poppler (cada llamada reabre y parsea el PDF) y para
# habilitar paralelismo real entre lotes en máquinas multi-núcleo.
OCR_CLUSTER_GAP = 5     # páginas que necesitan OCR separadas por <= N páginas se fusionan en un solo lote
OCR_MAX_LOTE = 40        # tamaño máximo de un lote — evita que un cluster gigante se vuelva 1 sola tarea sin paralelismo
OCR_MAX_WORKERS = 4      # lotes procesados en paralelo — subir si tu máquina tiene más núcleos disponibles


def _pagina_necesita_ocr(texto_pagina, min_chars=OCR_MIN_CHARS):
    """
    pdfplumber devuelve None o una cadena muy corta cuando la página
    es una imagen escaneada sin capa de texto digital.
    """
    if texto_pagina is None:
        return True
    return len(texto_pagina.strip()) < min_chars


def _agrupar_paginas_en_lotes(paginas_ocr, gap_maximo=OCR_CLUSTER_GAP, max_lote=OCR_MAX_LOTE):
    """
    Agrupa números de página que necesitan OCR en rangos (inicio, fin)
    inclusive, para minimizar la cantidad de llamadas a Poppler.

    Páginas separadas por <= gap_maximo páginas se fusionan en un mismo
    rango (se renderizan también las páginas intermedias que no
    necesitan OCR, pero sale más barato que abrir un proceso de Poppler
    nuevo por cada página individual).

    Rangos más grandes que max_lote se subdividen, para que un cluster
    grande no termine siendo una sola tarea gigante sin paralelismo.
    """
    if not paginas_ocr:
        return []
    paginas_ordenadas = sorted(paginas_ocr)
    lotes_brutos = []
    inicio = anterior = paginas_ordenadas[0]
    for p in paginas_ordenadas[1:]:
        if p - anterior > gap_maximo:
            lotes_brutos.append((inicio, anterior))
            inicio = p
        anterior = p
    lotes_brutos.append((inicio, anterior))

    lotes_finales = []
    for ini, fin in lotes_brutos:
        cursor = ini
        while cursor <= fin:
            sub_fin = min(cursor + max_lote - 1, fin)
            lotes_finales.append((cursor, sub_fin))
            cursor = sub_fin + 1
    return lotes_finales


def _ocr_lote(pdf_path, inicio, fin, paginas_necesarias, dpi=OCR_DPI, idioma=OCR_IDIOMA):
    """
    Renderiza el rango [inicio, fin] en UNA sola llamada a Poppler y
    aplica OCR solo a las páginas de `paginas_necesarias` dentro de ese
    rango (las demás páginas del rango ya tenían texto digital).
    Devuelve {numero_pagina: (texto_extraido, confianza_0_a_100)}.
    """
    resultados = {}
    try:
        imagenes = convert_from_path(pdf_path, dpi=dpi, first_page=inicio, last_page=fin)
    except Exception as e:
        logging.error(f"Error renderizando lote {inicio}-{fin} de {pdf_path}: {e}", exc_info=True)
        for p in paginas_necesarias:
            resultados[p] = ("", 0.0)
        return resultados

    for offset, imagen in enumerate(imagenes):
        numero_pagina = inicio + offset
        if numero_pagina not in paginas_necesarias:
            continue  # esta página del rango ya tenía texto digital
        try:
            datos = pytesseract.image_to_data(imagen, lang=idioma, output_type=Output.DICT)
            palabras, confianzas = [], []
            for texto, conf in zip(datos['text'], datos['conf']):
                if texto.strip():
                    palabras.append(texto)
                    if int(conf) >= 0:   # tesseract usa -1 cuando no calcula confianza
                        confianzas.append(int(conf))
            texto_final = ' '.join(palabras)
            confianza = sum(confianzas) / len(confianzas) if confianzas else 0.0
            resultados[numero_pagina] = (texto_final, confianza)
        except Exception as e:
            logging.error(f"Error OCR en página {numero_pagina} de {pdf_path}: {e}", exc_info=True)
            resultados[numero_pagina] = ("", 0.0)
    return resultados


# --- Filtro de tipo de documento: ¿es una execução fiscal de dívida ativa? ---
# [NUEVO] El Agente 1 solo está diseñado para clasificar execuções fiscais
# de dívida ativa. Sin este filtro, cualquier PDF judicial (ej. ações de
# saúde, mandados de segurança) recibe igualmente un DECISÃO/MOTIVO con el
# mismo formato prolijo, aunque la decisión no tenga sentido para ese caso.

KEYWORDS_CLASSE_EXECUCAO_FISCAL = [
    "execucao fiscal",
    "execucoes fiscais",
    "cobranca da divida ativa",
    "cobranca de divida ativa",
    "execucao da divida ativa",
]

KEYWORDS_CLASSE_NAO_FISCAL = [
    "obrigacao de fazer",
    "mandado de seguranca",
    "acao civil publica",
    "fornecimento de medicamento",
    "sem registro na anvisa",
    "indenizacao por dano",
    "alvara judicial",
    "usucapiao",
    "interdicao",
    "acao popular",
    "desapropriacao",
]

# Marcadores de respaldo en todo el texto, usados solo cuando no se
# encuentra (o es ambiguo) el campo Classe-Assunto del PJe
KEYWORDS_TEXTO_EXECUCAO_FISCAL = [
    "execucao fiscal",
    "exequente",
    "certidao de divida ativa",
    "cda no", "cda n.", "cda no.",
    "lei 6.830", "lei n. 6.830", "lei no 6.830",
    "divida ativa",
    "embargos a execucao fiscal",
    "penhora",
]

KEYWORDS_TEXTO_NAO_FISCAL = [
    "obrigacao de fazer",
    "fornecimento de medicamento",
    "tratamento medico",
    "anvisa",
    "sistema unico de saude",
    "alimentos",
    "guarda do menor",
    "uniao estavel",
    "usucapiao",
    "mandado de seguranca",
]


def _extraer_classe_assunto(text_norm):
    """
    Busca el campo 'Classe - Assunto:' típico de las capas administrativas
    del PJe. Usa re.DOTALL porque el PDF puede quebrar la línea a mitad de
    frase (ej. '[Obrigação de' \\n 'Fazer / Não Fazer...]'), y colapsa
    espacios/saltos de línea antes de devolver el fragmento.
    """
    m = re.search(r"classe\s*[-/]?\s*assunto\s*:?\s*(.{0,300})", text_norm, re.DOTALL)
    if not m:
        return None
    fragmento = re.sub(r"\s+", " ", m.group(1))
    corte = re.split(
        r"reclamante|requerente|autor|orgao julgador|exequente|executado|reclamado",
        fragmento
    )[0]
    return corte.strip()


def detectar_tipo_processo(text):
    """
    Determina si el texto corresponde a una execução fiscal de dívida ativa.

    Señal primaria: campo 'Classe - Assunto' del PJe (cuando aparece, es la
    fuente más confiable porque viene de la propia clasificación del tribunal).
    Señal de respaldo: conteo de keywords en todo el texto, usado solo cuando
    el campo Classe-Assunto no aparece o no es concluyente.

    Solo retorna confianza='alta' cuando el campo Classe-Assunto es explícito
    en una u otra dirección; la heurística por conteo nunca llega a 'alta',
    a propósito, para evitar bloquear documentos por una heurística todavía
    no validada contra suficientes casos reales de la PGMS.

    Devuelve:
      {
        "es_execucao_fiscal": bool,
        "confianza": "alta" | "media" | "baixa",
        "motivo": str,
        "classe_assunto": str | None,
      }
    """
    text_norm = normalizar(text)
    classe_assunto = _extraer_classe_assunto(text_norm)

    if classe_assunto:
        if any(normalizar(k) in classe_assunto for k in KEYWORDS_CLASSE_NAO_FISCAL):
            return {
                "es_execucao_fiscal": False,
                "confianza": "alta",
                "motivo": f"Classe-Assunto indica outro tipo de ação: '{classe_assunto[:150]}'",
                "classe_assunto": classe_assunto,
            }
        if any(normalizar(k) in classe_assunto for k in KEYWORDS_CLASSE_EXECUCAO_FISCAL):
            return {
                "es_execucao_fiscal": True,
                "confianza": "alta",
                "motivo": f"Classe-Assunto confirma execução fiscal: '{classe_assunto[:150]}'",
                "classe_assunto": classe_assunto,
            }
        # Campo encontrado pero ambiguo: cae a la heurística por conteo

    score_fiscal = sum(1 for k in KEYWORDS_TEXTO_EXECUCAO_FISCAL if normalizar(k) in text_norm)
    score_nao_fiscal = sum(1 for k in KEYWORDS_TEXTO_NAO_FISCAL if normalizar(k) in text_norm)

    if score_fiscal == 0 and score_nao_fiscal == 0:
        return {
            "es_execucao_fiscal": True,
            "confianza": "baixa",
            "motivo": "Nenhum marcador de tipo de ação encontrado — recomenda-se revisão manual",
            "classe_assunto": classe_assunto,
        }

    es_fiscal = score_fiscal >= score_nao_fiscal
    return {
        "es_execucao_fiscal": es_fiscal,
        "confianza": "media",
        "motivo": f"Heurística por contagem de palavras-chave: fiscal={score_fiscal}, não-fiscal={score_nao_fiscal}",
        "classe_assunto": classe_assunto,
    }


# 1. Extraer texto de PDF por páginas
def extract_text_by_page(pdf_path):
    """
    Extrae texto de cada página. Páginas sin texto digital suficiente se
    agrupan en lotes (rangos contiguos o casi contiguos) y se procesan
    con OCR en paralelo, en vez de abrir un proceso de Poppler nuevo por
    cada página individual.

    Devuelve:
      textos:   list[str]  — mismo formato que antes, compatible con el resto del pipeline
      metadata: dict       — {'paginas_ocr': [...], 'confianza_ocr': {pagina: score}}
    """
    textos = {}
    with pdfplumber.open(pdf_path) as pdf:
        total_paginas = len(pdf.pages)
        paginas_que_necesitan_ocr = []
        for i, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text()
            if _pagina_necesita_ocr(texto):
                paginas_que_necesitan_ocr.append(i)
                textos[i] = None  # se completa en la fase de OCR
            else:
                textos[i] = texto

    metadata = {"paginas_ocr": [], "confianza_ocr": {}}

    if paginas_que_necesitan_ocr:
        lotes = _agrupar_paginas_en_lotes(paginas_que_necesitan_ocr)
        necesarias_por_lote = {
            (ini, fin): set(p for p in paginas_que_necesitan_ocr if ini <= p <= fin)
            for ini, fin in lotes
        }
        logging.info(
            f"  {os.path.basename(pdf_path)}: {len(paginas_que_necesitan_ocr)} página(s) "
            f"necesitam OCR, agrupadas em {len(lotes)} lote(s) "
            f"(até {OCR_MAX_WORKERS} em paralelo)"
        )
        with ThreadPoolExecutor(max_workers=OCR_MAX_WORKERS) as executor:
            futuros = {
                executor.submit(_ocr_lote, pdf_path, ini, fin, necesarias_por_lote[(ini, fin)]): (ini, fin)
                for ini, fin in lotes
            }
            concluidos = 0
            for futuro in as_completed(futuros):
                ini, fin = futuros[futuro]
                concluidos += 1
                try:
                    resultado_lote = futuro.result()
                    for numero_pagina, (texto_ocr, confianza) in resultado_lote.items():
                        textos[numero_pagina] = texto_ocr
                        metadata["paginas_ocr"].append(numero_pagina)
                        metadata["confianza_ocr"][numero_pagina] = round(confianza, 1)
                    logging.info(f"  Lote {concluidos}/{len(lotes)} concluído (páginas {ini}-{fin})")
                except Exception as e:
                    logging.error(f"Lote {ini}-{fin} de {pdf_path} falhou: {e}", exc_info=True)

    metadata["paginas_ocr"].sort()
    lista_textos = [textos[i] for i in range(1, total_paginas + 1)]
    return lista_textos, metadata

# 2. Filtrar texto relevante con palabras clave
def filter_text_by_keywords(text, keywords):
    relevant_lines = []
    norm_keywords = [normalizar(k) for k in keywords]
    for line in text.splitlines():
        if any(kw in normalizar(line) for kw in norm_keywords):
            relevant_lines.append(line)
    return " ".join(relevant_lines)

# 3. Obtener la fecha más reciente en el texto
# Keywords que indican una movimentación procesal real
KEYWORDS_MOVIMENTACAO_PROCESSUAL = [
    # Despachos y decisiones
    "despacho", "decisao interlocutoria", "decisao",
    "sentenca", "acordao",
    "conferi.", "digitei, eu",
    # Certidões processuales
    "certidao de publicacao de relacao",
    "certidao de remessa da intimacao",
    "certidao de intimacao",
    "ciencia da intimacao",
    "certidao de publicacao",
    # Peticiones del exequente  [FIX Bug G]
    "pede deferimento",
    "pede juntada",
    "vem requerer",
    "vem expor e requerer",
    "vem, por seu procurador",
    "vem, perante",
    "vem, respeitosamente",
    "nestes termos, pede deferimento",
    "requer a v. exa",
    "requer a vossa excelencia",
    "por seu procurador infrafirmado",
    "por sua procuradora infrafirmada",
    "por seu procurador ao fim assinado",
    "vem aduzir e requerer",
    "reiterar pedido",
    # Certidões do oficial de justiça
    "certifico que",
    "certifico, para os devidos fins",
    "o referido e verdade e dou fe",
    "o referido e verdade",
    # Atos cartorários
    "ato ordinatorio",
    "cumpra-se",
    "publique-se. intime-se",
    # Localización de fecha en documentos judiciales
    "salvador (ba),",
    "salvador, ba,",
    # Intimaciones electrónicas
    "data da intimacao",
    "encaminhado para intimacao no portal eletronico",
]

def fecha_ultima_movimentacao(text):
    """
    Reemplaza fecha_mas_reciente().
    Solo considera fechas cercanas a keywords de movimentación processual,
    ignorando fechas de fichas cadastrais, extratos y consultas CNPJ.
    Ahora:
    - usa correctamente el texto normalizado para cortar fragmentos (evita desalineamiento de índices)
    - descarta fechas que aparecen dentro de secciones de exclusión (fichas/extratos)
    """
    text_norm = normalizar(text)
    fechas = []

    # Marcadores que delimitan secciones no processuales (misma lista que el fallback)
    MARCADORES_EXCLUSION = [
        "ficha cadastral",
        "extrato fiscal",
        "consulta de dados via cpf",
        "dados do cnpj",
        "dados de empresa via cnpj",
        "posicao de debito",
        "certidao de divida ativa",
        "termo de confissao de divida",
    ]
    excl_norm = [normalizar(m) for m in MARCADORES_EXCLUSION]

    for keyword in KEYWORDS_MOVIMENTACAO_PROCESSUAL:
        keyword_norm = normalizar(keyword)
        for match in re.finditer(re.escape(keyword_norm), text_norm):
            start = max(0, match.start() - 300)
            end   = match.end() + 300
            # Usar el texto normalizado (text_norm) para extraer el fragmento: evita desalineamiento
            fragmento_norm = text_norm[start:end]

            # Si el fragmento contiene marcadores de exclusión, ignorarlo
            if any(marker in fragmento_norm for marker in excl_norm):
                continue

            fechas_encontradas = _extraer_fechas_de_texto(fragmento_norm)
            if fechas_encontradas:
                fechas.extend(fechas_encontradas)

    if not fechas:
        # Fallback: si no encontró nada con keywords,
        # usar el método global pero excluyendo secciones de fichas y extratos
        return _fecha_mas_reciente_fallback(text)

    return max(fechas)


def _fecha_mas_reciente_fallback(text):
    """
    Fallback: escanea todo el texto pero excluye secciones
    de fichas cadastrais, extratos y consultas CNPJ.
    """
    # Marcadores que delimitan secciones no processuales
    MARCADORES_EXCLUSION = [
        "ficha cadastral",
        "extrato fiscal",
        "consulta de dados via cpf",
        "dados do cnpj",
        "dados de empresa via cnpj",
        "posicao de debito",
        "certidao de divida ativa",
        "termo de confissao de divida",
    ]

    text_norm = normalizar(text)
    lineas = text.split("\n")
    lineas_filtradas = []
    excluir = False

    for linea in lineas:
        linea_norm = normalizar(linea)
        # Activar exclusión si encontramos un marcador
        if any(normalizar(m) in linea_norm for m in MARCADORES_EXCLUSION):
            excluir = True
        # Desactivar exclusión al encontrar inicio de documento judicial
        if any(kw in linea_norm for kw in ["poder judiciario", "comarca de salvador", "despacho", "decisao", "certidao"]):
            excluir = False
        if not excluir:
            lineas_filtradas.append(linea)

    texto_filtrado = "\n".join(lineas_filtradas)
    fechas = _extraer_fechas_de_texto(normalizar(texto_filtrado))
    return max(fechas) if fechas else None
    
# Buscador de fechas cercanas a palabras clave específicas (como las relacionadas con citación)
def _limpar_referencias_legais(text_norm):
    """
    [FIX fecha_intento] Remove datas de referencias legais (decretos, leis, resolucoes)
    para evitar que sejam capturadas como datas de citacao.
    Ex: "Decreto Judiciario nr 638, de 17 de setembro de 2018" → removido
        "Lei nº 6.830, de 22 de setembro de 1980"             → removido
    """
    # Padrao generico: "palavra-chave legal ... numero ..., de DD de mes de AAAA"
    padroes = [
        # Decretos, leis, resolucoes, portarias
        r'decreto\s+\w+\s+n\w*\s*\d+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'decreto\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'lei\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'resolucao\s+n\w*\s*\d+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'instrucao\s+normativa\s+\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'portaria\s+n\w*\s*[\d\.]+,?\s*de\s+\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        # [FIX RS Gesso Bug1] Datas de assinatura de Certidoes de Divida Ativa
        # Ex: "Certifico que se acha inscrito na Divida Ativa... Salvador, 18 de setembro de 2012"
        # Essas datas aparecem 4x (uma por CDA) e contaminam a fecha_intento
        r'certifico que se acha inscrito[^S]{0,400}?salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        r'estes sao os elementos contidos[^S]{0,200}?salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}',
        # Assinatura de coordenador/procurador na CDA (logo após a data)
        r'salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}\s+raimundo\s+cordeiro',
        r'salvador,?\s*\d{1,2}\s+de\s+\w+\s+de\s+\d{4}\s+\w+\s+cordeiro',
        r'coordenador\s+da\s+cda',  # marca o contexto CDA
    ]
    result = text_norm
    for pat in padroes:
        result = re.sub(pat, ' [REF_LEGAL] ', result)
    return result


def extraer_fecha_cercana(text, keywords, ventana=500, prefer='latest', min_year=1990):
    """
    Busca fechas cerca de las ocurrencias de `keywords` en `text`.
    - ventana: número de caracteres antes/después para buscar fechas.
    - prefer: 'latest' (por defecto) devuelve la fecha más reciente entre las encontradas;
              'earliest' devuelve la primera (la más antigua).
    Esto permite priorizar la primera carta expedida (usar prefer='earliest').
    """
    # [FIX] limpar referencias legais antes de buscar datas
    text_norm = _limpar_referencias_legais(normalizar(text))
    fechas = []
    for keyword in keywords:
        kw_norm = normalizar(keyword)
        for match in re.finditer(re.escape(kw_norm), text_norm):
            start = max(0, match.start() - ventana)
            end = match.end() + ventana
            fechas_local = _extraer_fechas_de_texto(text_norm[start:end])
            fechas_local = [f for f in fechas_local if f.year >= min_year]
            if fechas_local:
                fechas.extend(fechas_local)
    if not fechas:
        return None
    if prefer == 'earliest':
        return min(fechas)
    return max(fechas)

# Función específica para extraer fechas relacionadas con citación, usando diferentes conjuntos de keywords para intentar capturar distintos tipos de órdenes y resultados de citación.
def extraer_fechas_citacion(text):
    """
    Extrae tres fechas relacionadas a la citación:
      - fecha_orden: fecha de la primera orden/carta expedida (priorizar la primera)
      - fecha_intento: fecha de la primera tentativa/AR
      - fecha_efectiva: fecha próxima a evidencias de entrega/assinatura (firma/AR)
    Separando la lógica se evita confundir una tentativa con una citación efectiva.
    """
    KEYWORDS_ORDEN_ESPECIFICAS = [
        "determino a citação",
        "ordeno a citação",
        "expeça-se o competente mandado de citação",
        "expeça-se carta de citação",
        "expeça-se mandado de citação",
        "proceda-se à citação",
        "promova-se a citação",
        "expedir carta de citação",
        "diligencie-se para citação",
        "ao oficial de justiça, cite",
        "cite-se por edital", "citação por edital",
        "converta-se em mandado de citação",
    ]
    KEYWORDS_ORDEN_GENERICAS = [
        "expeça-se citação",
        "seja citado", "sejam citados",
        "citem-se",
        "cite-se",
        "cite(m)-se",      # [FIX] variante com (m): despachos usam "Cite(m)-se."
        "proceda-se a citacao",
        "proceda-se a citação",
        # [FIX Marcio] Despachos longos (~1000 chars): a data aparece após
        # "Registre-se" ou "o presente despacho servira como mandado",
        # keywords muito mais próximos da data do que "proceda-se a citacao"
        "o presente despacho servira como mandado",
        "registre-se.",
        "publique-se. registre-se",
    ]
    KEYWORDS_ORDEN = KEYWORDS_ORDEN_ESPECIFICAS + KEYWORDS_ORDEN_GENERICAS

    # Keywords que aparecem ANTES da data (buscar apenas APÓS o keyword)
    KEYWORDS_INTENTO_APOS = [
        "conferi.",          # carta: "conferi. 23 de maio de 2013"
        "digitei, eu",       # mandado: "digitei, ... conferi. DD de mes de AAAA"
    ]
    # Keywords que aparecem em contexto amplo (janela bidirecional)
    KEYWORDS_INTENTO = [
        "carta com ar", "aviso de recebimento", "tentativa",
        "aviso de recebimento negativo",
    ]

    KEYWORDS_EFECTIVA = [
        "recebido", "assinado", "assinatura", "assinatura do recebedor", "recebido em", "recebido com"
    ]

    # Buscar la PRIMERA carta/orden expedida (earliest) para fecha_orden
    # [FIX RS Gesso Bug2] ventana reduzida 1500->400: certidoes de publicacao
    # republiquem o despacho sem a data original, ficando a ~1000 chars do keyword.
    # Com ventana=400 capturamos o despacho real sem pegar as certidoes.
    # [FIX Ministério Obra Santa] ventana 400→800: despachos longos (>500 chars)
    # têm a data no final; 800 cobre sem pegar certidoes a 1000+ chars.
    fecha_orden = extraer_fecha_cercana(text, KEYWORDS_ORDEN, ventana=800, prefer='earliest')

    # Separar la lógica de intento: priorizar la primera tentativa/ar encontrada
    # [FIX RS Gesso] Dois passes para fecha_intento:
    # 1) Keywords que precedem a data (ex: "conferi.") → buscar só APÓS o keyword
    #    usando ventana_antes=10 para não capturar datas anteriores no mesmo bloco
    import re as _re
    def _fecha_intento_apos(text_norm, keywords, min_year=1990):
        """Busca data apenas APÓS o keyword (ventana_antes=10)."""
        fechas = []
        for kw in keywords:
            kw_norm = normalizar(kw)
            for m in _re.finditer(_re.escape(kw_norm), text_norm):
                # janela: apenas 10 chars antes (para alguma margem) e 300 após
                trecho = text_norm[max(0, m.start()-10): m.end()+300]
                fs = [f for f in _extraer_fechas_de_texto(trecho) if f.year >= min_year]
                fechas.extend(fs)
        return min(fechas) if fechas else None

    text_norm_clean = _limpar_referencias_legais(normalizar(text))
    fecha_intento = _fecha_intento_apos(text_norm_clean, KEYWORDS_INTENTO_APOS)
    if not fecha_intento:
        fecha_intento = extraer_fecha_cercana(text, KEYWORDS_INTENTO, ventana=500, prefer='earliest')

    # Fecha efectiva: buscar indicios de entrega/firmas (usar ventana menor)
    fecha_efectiva = extraer_fecha_cercana(text, KEYWORDS_EFECTIVA, ventana=400, prefer='earliest')

    # Depuración: mostrar las fechas detectadas por tipo
    logging.debug(f"  fecha_orden (primera):    {fecha_orden}")
    logging.debug(f"  fecha_intento (primera):  {fecha_intento}")
    logging.debug(f"  fecha_efectiva (posible): {fecha_efectiva}")

    return fecha_orden, fecha_intento, fecha_efectiva

# [FIX Bug A] KEYWORDS_AR_ENTREGUE removido — essas strings aparecem como campos
# em branco em formularios AR nao entregues e causavam falso positivo sistematico.
# A deteccao de citacao efetiva agora usa apenas KEYWORDS_CITACION_OK em
# extract_citacion(), que so inclui expressoes que provam entrega real.


# 4. Extraer estado de citación
def extract_citacion(text):
    # [FIX Bug A] Keywords que SO aparecem em certidoes de citacao efetiva.
    # NAO incluir "assinatura do recebedor", "nome legivel do recebedor",
    # "data de entrega" — sao boilerplate impresso em branco em todo formulario AR.
    KEYWORDS_CITACION_OK = [
        # Certidao do oficial de justica confirmando citacao pessoal
        "certifico que procedi a citacao",
        "certifico que o executado foi citado",
        "certifico que citei",
        "certifico ter realizado a citacao",
        "fica citado",
        "devidamente citado",
        "ar positivo",
        # [FIX Maria Gloria] Certidão de Decurso de Prazo = citação válida provada
        # Só há prazo para correr se o executado foi efetivamente citado
        "certidao de decurso de prazo",
        "decorreu o prazo legal sem qualquer manifestacao",
        "decorreu o prazo legal",
        "decurso de prazo",
        "nao se manifestou quanto ao pagamento",
        # Defesa / embargos (executado compareceu)
        #"embargos a execucao",
        "apresentou embargos",
        "embargos foram opostos",
        "citacao valida",
        # Citacao espontanea — confissao de divida / parcelamento assinado
        "instrumento de confissao de divida e compromisso de pagamento parcelado",
        "instrumento de confissao de divida",
        "confissao de divida e compromisso",
        "exarou o ciente",
        "aceitou a contrafe que lhe foi oferecida",
        "ele aceitou a contrafe",
        "citado nos autos",
        # PAD assinado = ciencia inequivoca
        "parcelamento de debitos",
    ]

    # [FIX Bug A] Inclui motivos de devolucao do AR e certidoes negativas do OJ
    KEYWORDS_CITACION_NAO_OK = [
        # Declaracoes judiciais explicitas
        "aviso de recebimento negativo",
        "intime-se a fazenda publica para que adote as providencias cabiveis",
        "o reu nao foi citado",
        "executado nao foi citado",
        "sem citacao do executado",
        "nao houve citacao",
        "ausencia de citacao",
        "nao logrando exito na citacao",
        "nao foi possivel realizar a citacao",
        "a parte executada nao foi citada",
        # Certidao do oficial de justica de tentativa frustrada
        "deixei de proceder a citacao",
        "deixei de citar",
        "nao encontrado o executado",
        "nao foi localizado o executado",
        "nao reside no endereco",
        "nao mora no endereco",
        "nao conhece o executado",
        "nao sabe informar o seu paradeiro",
        # Motivos de devolucao marcados no AR pelos Correios
        "motivos de devolucao",   # cabecalho da secao = AR devolvido
        "mudou-se",
        "nao procurado",
        "nao existe o numero",
        "endereco insuficiente",
        "desconhecido",
        "falecido",
    ]

    text_norm = normalizar(text)

    # Verificar negativo primeiro
    if any(normalizar(k) in text_norm for k in KEYWORDS_CITACION_NAO_OK):
        # Se ha tambem evidencia positiva forte (ex: confissao assinada apos
        # tentativa frustrada), o positivo prevalece
        for k_ok in KEYWORDS_CITACION_OK:
            if normalizar(k_ok) in text_norm:
                return "HOUVE CITAÇÃO"
        return "NÃO HOUVE ou TENTATIVA FALHA"

    # Verificar positivo
    if any(normalizar(k) in text_norm for k in KEYWORDS_CITACION_OK):
        return "HOUVE CITAÇÃO"

    # [FIX Bug A] REMOVIDO: bloco de verificacao por "assinatura do recebedor" /
    # "nome legivel do recebedor" — essas strings aparecem impressas em branco em
    # todo formulario AR nao entregue e causavam falso positivo em 81% dos casos.

    return "Citação não encontrado"

# 5. Extraer resultado de la penhora
def extract_penhora(text):
    # Adicionamos muitas variações aos filtros.
    # --- BACENJUD / SISBAJUD (bloqueio bancário) ---
    # Penhora bancária confirmada (despacho autorizando OU comprovante de bloqueio)
    KEYWORDS_BACENJUD_POSITIVO = [
        "bloqueio bacenjud", "penhora bacenjud",
        "bloqueio sisbajud", "penhora sisbajud",
        "penhora on-line", "penhora online",
    ]
    # Solicited via petition — result not yet confirmed in the document
    KEYWORDS_BACENJUD_SOLICITADO = [
        "via sistema sisbajud",
        "via sistema bacenjud",
        "via bacenjud",
        "via sisbajud",
        "sistema sisbajud",
        "sistema bacenjud",
        "bloqueio de dinheiro/ativos financeiros",
        "bloqueio de ativos financeiros",
        "requer o bloqueio de dinheiro",
        # [FIX] variantes com espaço: "BACEN JUD", "SISBAJUD", "SIS BAJUD"
        "bacen jud",
        "sis bajud",
        "nos moldes do bacen jud",
        "nos moldes do bacenjud",
    ]
    # Confirmed successful bloqueio (result document present)
    KEYWORDS_BACENJUD_CONFIRMADO = [
        "valores bloqueados",
        "bloqueio efetuado",
        "bloqueio realizado",
        "dinheiro bloqueado",
        "penhora on-line efetuada",
        "penhora online efetuada",
        "extrato de bloqueio",
        "certidao de bloqueio",
        "comprovante de bloqueio",
    ]
    KEYWORDS_BACENJUD_NEGATIVO = [
        "resultado negativo da diligencia bacenjud",
        "resultado negativo da diligencia sisbajud",
        "suspensao pelo art. 40 da lef",
        "sem saldo positivo",
        "nao ha saldo",
        "bacenjud sem exito",
        "sisbajud sem exito",
        "diligencia bacenjud restou infrutifera",
        "diligencia sisbajud restou infrutifera",
        "bloqueio desbloqueado",
        "desbloqueio do valor",
        # [FIX Maria Gloria] SisBajud resultado negativo na decisão judicial
        "nao possui relacionamento com as instituicoes financeiras",
        "cpf indicado nao possui relacionamento",
        "cnpj indicado nao possui relacionamento",
        # [FIX Ministério Obra Santa] "negativa Bacenjud" em decisão art.40 LEF
        "negativa bacenjud",
        "ar negativo e/ou negativa bacenjud",
        "negativa do bacenjud",
        "bacenjud negativo",
        "resultado bacenjud negativo",
        "nao possui relacionamentos com",
        "cpf nao possui relacionamento",
        "tentativa de penhora on-line",
    ]

    # --- RENAJUD (bloqueio veicular) ---
    KEYWORDS_RENAJUD_ATIVO = [
        "comprovante de inclusao de restricao veicular",
        "insercao de restricao veicular",
        "bloqueio renajud",
        "restricao renajud",
        "penhora de veiculo",
        "auto de penhora de veiculo",
    ]
    KEYWORDS_RENAJUD_NEGATIVO = [
        "renajud sem exito",
        "restricao renajud cancelada",
        "nao foram localizados veiculos",
        "pesquisa renajud negativa",
    ]

    # --- Penhora de imóvel ---
    KEYWORDS_PENHORA_IMOVEL = [
        "penhora de imovel",
        "penhora do imovel",
        "registro de penhora",
        "matricula do imovel penhorado",
        "imovel penhorado",
        "termo de penhora de imovel",
        "penhora sobre imovel",
    ]

    # --- Penhora de faturamento ---
    KEYWORDS_PENHORA_FATURAMENTO = [
        "penhora de faturamento",
        "penhora sobre faturamento",
        "penhora sobre o faturamento",
        "deposito de percentual do faturamento",
    ]

    # --- Indisponibilidade de bens (CNIB) ---
    KEYWORDS_INDISPONIBILIDADE = [
        "indisponibilidade de bens",
        "decretada a indisponibilidade",
        "cnib",
        "cadastro nacional de indisponibilidade",
        # [FIX Felipao] "bloqueio de bens" REMOVIDO — aparece em
        # decisões de RENAJUD ("imediato bloqueio de bens eventualmente
        # encontrados"), causando falso positivo CNIB
    ]

    # --- Penhora de quotas / participação societária ---
    KEYWORDS_PENHORA_QUOTAS = [
        "penhora de quotas",
        "penhora de cotas",
        "penhora de participacao societaria",
        "penhora de acoes",
    ]

    # --- Penhora de créditos / precatório / direitos ---
    KEYWORDS_PENHORA_CREDITOS = [
        "penhora de creditos",
        "penhora de precatorio",
        "penhora de direitos",
        "penhora de direitos hereditarios",
        "penhora de aplicacao financeira",
    ]

    # --- Tentativas genéricas (sem resultado confirmado) ---
    # REMOVED "intimacao da penhora" — aparece como boilerplate em cartas de citação:
    #   "...ou da intimação da penhora (art. 16 da Lei n. 6.830/80)"
    # Use "intimado da penhora" (particípio passado = notificação já ocorrida) em vez disso.
    KEYWORDS_TENTATIVA_PENHORA = [
        "termo de penhora",
        "auto de penhora e avaliacao",
        "diligencia de penhora",
        "tentativa de penhora",
        "intimado da penhora",
        "intimados da penhora",
        "oficial de justica nao localizou bens",
        "nao foram localizados bens",
        "nao encontrou bens",
        "sem bens a penhorar",
        "bens insuficientes",
    ]
    text_norm = normalizar(text)

    # Indisponibilidade CNIB — bloqueio administrativo amplo
    # [FIX Felipao] RENAJUD verificado ANTES de CNIB —
    # decisões RENAJUD usam "bloqueio de bens" que causava falso positivo CNIB
    if any(normalizar(k) in text_norm for k in KEYWORDS_RENAJUD_ATIVO):
        if any(normalizar(k) in text_norm for k in KEYWORDS_RENAJUD_NEGATIVO):
            return "tentativa de penhora renajud (negativa)"
        return "bloqueio renajud ativo"

    if any(normalizar(k) in text_norm for k in KEYWORDS_INDISPONIBILIDADE):
        return "indisponibilidade de bens (CNIB)"

    # Penhora de imóvel
    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_IMOVEL):
        return "penhora de imóvel"

    # Penhora de faturamento
    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_FATURAMENTO):
        return "penhora de faturamento"

    # Penhora de quotas societárias
    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_QUOTAS):
        return "penhora de quotas/ações"

    # Penhora de créditos / precatório / direitos
    if any(normalizar(k) in text_norm for k in KEYWORDS_PENHORA_CREDITOS):
        return "penhora de créditos/direitos"

    # [FIX 3a] BACENJUD_NEGATIVO verificado de forma INDEPENDENTE:
    # Cobre casos onde o juiz indefere o bloqueio (sem keyword positivo)
    # Ex: "CPF não possui relacionamento com as instituições financeiras"
    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_NEGATIVO):
        return "tentativa de penhora bacenjud (negativa)"

    # BACENJUD / SISBAJUD — confirmed positive keywords (comprovante/despacho de bloqueio)
    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_POSITIVO):
        return "penhora bacenjud/sisbajud"

    # BACENJUD / SISBAJUD — only a petition requesting it; result not in document
    if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_SOLICITADO):
        if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_NEGATIVO):
            return "tentativa de penhora bacenjud (negativa)"
        if any(normalizar(k) in text_norm for k in KEYWORDS_BACENJUD_CONFIRMADO):
            return "penhora bacenjud/sisbajud"
        return "sisbajud/bacenjud solicitado (resultado desconhecido)"

    # Tentativas genéricas sem penhora efetivada
    if any(normalizar(k) in text_norm for k in KEYWORDS_TENTATIVA_PENHORA):
        return "tentativa de penhora (sem resultado)"

    if "penhora nao realizada" in text_norm:
        return "penhora não realizada"

    return "Penhora não encontrado"
# Nueva función para detectar suspensión por parcelamento
KEYWORDS_PARCELAMENTO_ATIVO = [
    # ── DECISÕES DO JUIZ deferindo suspensão por parcelamento ──────────────
    "defiro o pedido de suspensao",        # "Defiro o pedido de suspensão pelo prazo requerido"
    "defiro a suspensao",
    "determino a suspensao do feito",
    "determino a suspensao da execucao",
    "suspendo o feito",
    "suspendo o curso do feito",
    "suspendo/mantenho suspenso",
    "suspensao do feito",                  # em petições e decisões
    "suspensao da execucao",
    "suspensao do processo",

    # ── PETIÇÕES requerendo suspensão por parcelamento ──────────────────────
    "requerer a suspensao do feito",       # "vem requerer a suspensão do Feito"
    "requer a suspensao do feito",
    "requerer a suspensao da execucao",
    "requer a suspensao da execucao",
    "suspensao pelo parcelamento",
    "pleiteou administrativamente o parcelamento",  # Antonildo
    "adesao ao parcelamento",
    "aderiu ao parcelamento",
    "aderiu o executado",                  # "parcelamento a que aderiu o Executado"
    "parcelamento a que aderiu",

    # ── OBJETO do parcelamento ───────────────────────────────────────────────
    "parcelamento do debito exequendo",
    "parcelamento do credito tributario",
    "parcelamento do debito tributario",
    "parcelamento administrativo de debitos",
    "parcelamento administrativo",
    "parcelamento em",                     # "parcelamento em 38 parcelas"
    "parcelas mensais e sucessivas",
    "em 38 parcelas",                      # número explícito de parcelas
    "em 12 parcelas",
    "em 24 parcelas",
    "em 36 parcelas",
    "em 48 parcelas",
    "em 60 parcelas",

    # ── REFERÊNCIAS AO PAD (número ou sigla) ────────────────────────────────
    "adesao ao pad",
    "pad homologado",
    "bloq pad",
    "pad n",                               # "PAD n. 968002" / "PAD nº 3070"
    "pad no",
    "pad nr",
    "parcelamento pad",
    "pad internet",                        # "Tipo de Adesão: PAD Internet"

    # ── INSTRUMENTO DE CONFISSÃO DE DÍVIDA ──────────────────────────────────
    "instrumento de confissao de divida e compromisso de pagamento parcelado",
    "compromisso de pagamento parcelado",
    "instrumento de confissao",            # OCR pode garbling "de divida"

    # ── ARTIGOS DO CTN (inciso I e VI) — várias grafias ─────────────────────
    # Inciso VI — moratória / parcelamento
    "art. 151, vi, do ctn",
    "art. 151, inc. vi",
    "art. 151, inciso vi",
    "art. 151, inc. vi,",
    # Inciso I — parcelamento (menos comum, mas usado em Antonildo)
    "art. 151, i, do ctn",
    "art. 151, inc. i, do ctn",
    "art. 151, inc. i,",
    "art. 151, inciso i",
    # Numeral arábico "1" em vez de letra "i" (OCR / digitação)
    "art. 151, inc. 1,",
    "art. 151, 1, do ctn",
    "art. 151, inc. 1, do",
    "151, inc. 1,",
    # Artigo processual de suspensão
    "art. 313, ii, do cpc",
    "art. 265, inc. ii, do cpc",           # CPC/1973 — código antigo
    "art. 265, inc. ii",
]

KEYWORDS_SUSPENSAO_ART40 = [
    "resultado negativo da diligencia bacenjud",
    "resultado negativo da diligencia sisbajud",
    "suspendo a presente execucao fiscal",
    "art. 40, caput, da lei n. 6.830",
    "arquive-se o feito nos termos do art. 40",
    "suspensao da execucao fiscal pelo prazo de 01",
]

def extract_suspensao_art40(text):
    text_norm = normalizar(text)
    if any(normalizar(k) in text_norm for k in KEYWORDS_SUSPENSAO_ART40):
        return "processo suspenso — art. 40 LEF (bens não localizados)"
    return None


def extract_extincao(text):
    """
    [FIX Maria Gloria] Detecta se o processo foi extinto/sentenciado.
    Retorna string descritiva ou None.

    Lógica em três camadas:

    1. REVERSÃO: se o texto contém sinais de que uma sentença de extinção
       foi REFORMADA por acórdão de apelação, retorna None imediatamente —
       o processo foi reativado e não deve ser marcado como extinto.
       Exemplos reais: acórdão TJBA dando provimento ao Município e mandando
       os autos de volta ao juízo de origem para tramitação regular.

    2. KEYWORDS FORTES: muito específicas de extinção — sozinhas bastam para
       confirmar que o processo foi extinto. Só aparecem em sentenças/decisões
       que efetivamente extinguem a execução.

    3. KEYWORDS FRACAS: genéricas, aparecem em qualquer processo incluindo em
       acórdãos que *reformam* sentenças de extinção (ex: "arquivem-se",
       "transito em julgado", "dando-se baixa"). Só disparam extinção se
       acompanhadas de pelo menos uma keyword forte — nunca sozinhas.

    [FIX bug latente identificado em chat anterior]
    "transito em julgado" e "arquivem-se" aparecem no fechamento de qualquer
    processo judicial brasileiro; remover como disparadores isolados evita falso
    positivo em processos com sentença de extinção REFORMADA por apelação.
    """
    # ── 1. SINAIS DE REVERSÃO (sentença de extinção reformada por apelação) ──
    # Se qualquer um destes aparecer, a extinção foi revertida → não é extinto.
    KEYWORDS_REVERSAO = [
        # Acórdão dando provimento ao exequente (reforma da extinção)
        "dar provimento",
        "dou provimento",
        "da-se provimento",
        "recurso provido",
        "provimento ao recurso",
        "reforma a sentenca",
        "reforma-se a sentenca",
        "reformando a sentenca",
        "anulando a sentenca",
        "anulo a sentenca",
        # Retorno ao juízo de origem para tramitação (após reforma)
        "retorno dos autos ao primeiro grau",
        "retornem os autos ao juizo de origem",
        "retornem-se os autos",
        "retorno ao juizo de origem",
        "regular tramitacao",
        "para regular tramitacao",
        # Redirecionamento da execução (sócio-gerente) — processo ativo
        "redirecionamento da execucao",
        "redirecionar a execucao",
        "desconsideracao da personalidade juridica",
        "incluir o socio",
        "citar o socio",
    ]

    # ── 2. KEYWORDS FORTES (extinção confirmada, sozinhas bastam) ──
    KEYWORDS_FORTES = [
        "processo ja se encontra sentenciado",
        "ja se encontra sentenciado",
        "processo sentenciado",
        "extingo o processo com resolucao do merito",
        "declaro a prescricao",
        "declaro extinto o processo",
        "julgo extinta a execucao",
        "julgo extinto o feito",
        "extingo a execucao",
        "extincao da execucao",
        "extincao do processo",
        "sentenca de extincao",
        "extinto por prescricao",
        "extinta por prescricao",
        # "processo extinto" mantido como forte — é conclusivo quando isolado
        "processo extinto",
    ]

    # ── 3. KEYWORDS FRACAS (precisam de ao menos uma forte para disparar) ──
    # Estas aparecem em qualquer encerramento processual, inclusive em acórdãos
    # que reformam sentenças de extinção. Nunca disparam sozinhas.
    KEYWORDS_FRACAS = [
        "sentenca transitada",
        "transitada em julgado",
        "transito em julgado",      # [FIX] aparece em acórdãos de reforma também
        "processo arquivado",
        "arquivem-se",              # [FIX] encerramento genérico, não só extinção
        "dando-se baixa",           # [FIX] idem
        "extinção da execução",
        "extinção do processo",
    ]

    text_norm = normalizar(text)

    # Passo 1: checar reversão — se houver, retorna None diretamente
    if any(normalizar(k) in text_norm for k in KEYWORDS_REVERSAO):
        return None

    # Passo 2: checar keywords fortes — qualquer uma dispara extinção
    tem_forte = any(normalizar(k) in text_norm for k in KEYWORDS_FORTES)
    if tem_forte:
        return "processo extinto/sentenciado"

    # Passo 3: keywords fracas só disparam se acompanhadas de ao menos uma forte
    # (que já verificamos acima — se chegou aqui, não há forte)
    # → keywords fracas sozinhas NÃO disparam extinção
    return None


# ===========================================================================
# EXTRACCIÓN DE ENTIDADES PARA EL AGENTE 2
# ===========================================================================
#
# Propósito: extraer del texto del processo los datos estructurados que el
# Agente 2 necesitará para análisis jurídico-fiscal y priorización de cobro.
#
# Diseño deliberadamente conservador:
#   - Cada campo devuelve el valor extraído O None (nunca inventa).
#   - Todos los campos son opcionales: si no se encuentra, el downstream
#     decide qué hacer con None.
#   - Se puede ampliar campo a campo sin tocar el resto del pipeline.
#
# Campos actuales (base funcional v1):
#   cpf_cnpj         — CPF o CNPJ del executado (primer encontrado)
#   nome_executado   — Razón social / nombre del executado
#   nome_exequente   — Municipio / ente exequente
#   valor_original   — Valor original de la deuda (R$)
#   valor_atualizado — Valor actualizado de la deuda (R$), si figura
#   tipo_tributo     — ISS, IPTU, TFF, etc.
#   numero_cda       — Número de la Certidão de Dívida Ativa
#   numero_processo  — Número CNJ del processo
#   data_inscricao   — Fecha de inscripción en la dívida ativa
#   vara             — Vara/Juízo (ej. "9ª Vara da Fazenda Pública")
#   exercicio        — Año(s) del tributo (ej. "2006", "2014/2015")
#
# Para agregar un campo nuevo: definir su regex/lógica como función privada
# _extrair_CAMPO() y llamarla dentro de extract_entidades_agente2().
# ===========================================================================

def _extrair_cpf_cnpj(text):
    """
    Extrae el CPF o CNPJ del executado.
    Prioriza CNPJ (14 dígitos) sobre CPF (11 dígitos).
    Busca en contexto de labels típicos del processo para evitar falsos positivos
    (ej. CNPJ del propio Municipio que aparece como exequente).
    Devuelve el primer CNPJ/CPF encontrado asociado al executado, o None.
    """
    # Primero intentamos con contexto "executado" / "contribuinte" / "réu"
    _PATRON_CNPJ = r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[\-\s]?\d{2}"
    _PATRON_CPF  = r"\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[\-\s]?\d{2}"

    # Buscar en contexto del executado/réu/contribuinte (primeras apariciones)
    contextos_executado = [
        r"executad[oa][:\s]+[^\n]{0,200}",
        r"r[eé]u[:\s]+[^\n]{0,200}",
        r"contribuinte[:\s]+[^\n]{0,200}",
        r"cpf[\/]?cnpj[:\s]+[^\n]{0,100}",
        r"cnpj[\/]?cpf[:\s]+[^\n]{0,100}",
        r"inscri[cç][aã]o[:\s]+[^\n]{0,100}",
    ]
    text_norm = normalizar(text)
    for ctx_pat in contextos_executado:
        for ctx_match in re.finditer(ctx_pat, text_norm):
            fragmento = ctx_match.group(0)
            # Preferir CNPJ
            m = re.search(_PATRON_CNPJ, fragmento)
            if m:
                raw = re.sub(r"[\s]", "", m.group(0))
                digitos = re.sub(r"\D", "", raw)
                if len(digitos) == 14:
                    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
            # Fallback CPF
            m = re.search(_PATRON_CPF, fragmento)
            if m:
                raw = re.sub(r"[\s]", "", m.group(0))
                digitos = re.sub(r"\D", "", raw)
                if len(digitos) == 11:
                    return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    return None


def _extrair_nome_executado(text):
    """
    Extrae el nombre/razón social del executado.
    Busca en labels típicos del cabeçalho del processo.
    """
    _PATRONES = [
        r"executad[oa]\s*:\s*([^\n\r]{3,80})",
        r"r[eé]u\s*:\s*([^\n\r]{3,80})",
        r"contribuinte\s*:\s*([^\n\r]{3,80})",
        r"nome[\/\s]?raz[aã]o social\s*:\s*([^\n\r]{3,80})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            nome = m.group(1).strip().rstrip(".,;")
            # Filtrar resultados demasiado cortos o que son solo números
            if len(nome) >= 4 and not nome.replace(" ", "").isdigit():
                return nome[:100]
    return None


def _extrair_nome_exequente(text):
    """
    Extrae el nombre del exequente (normalmente 'Município de Salvador' o similar).
    [FIX] Limpia artefactos de OCR como comillas dobles al inicio ("''Município...").
    """
    _PATRONES = [
        r"exequente\s*:\s*([^\n\r]{3,80})",
        r"credor\s*:\s*([^\n\r]{3,80})",
        r"autor\s*:\s*([^\n\r]{3,80})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            nome = m.group(1).strip().rstrip(".,;")
            # Limpiar comillas OCR al inicio
            nome = nome.lstrip("'\"\u201c\u201d\u2018\u2019").strip()
            if len(nome) >= 4:
                return nome[:80]
    return None


def _extrair_valor(text):
    """
    Extrae valor original y valor actualizado de la deuda.
    Devuelve (valor_original, valor_atualizado) como strings "R$ X.XXX,XX" o None.

    Estrategia:
    - "Valor Originário" o "Valor Original" → valor_original
    - "Total em R$" / "Total a Pagar" / "Valor Atual" → valor_atualizado
    - Si no hay labels, busca la primera mención de valor en R$
    """
    # Patrón de valor monetario brasileño
    _PAT_VALOR = r"r\$\s*[\d.,]+"

    def _limpiar_valor(raw):
        """Normaliza el valor a formato 'R$ X.XXX,XX'"""
        m = re.search(r"[\d.,]+", raw)
        if not m:
            return None
        return f"R$ {m.group(0).strip()}"

    text_norm = normalizar(text)

    valor_original   = None
    valor_atualizado = None

    # Buscar valor originário
    for pat in [
        r"valor origin[aá]ri[oa]\s*[r\$:\s]+([\d.,]+)",
        r"valor original\s*[r\$:\s]+([\d.,]+)",
        r"vl\.?\s*original\s*[r\$:\s]+([\d.,]+)",
    ]:
        m = re.search(pat, text_norm)
        if m:
            valor_original = f"R$ {m.group(1).strip()}"
            break

    # Buscar valor atualizado / total a pagar
    for pat in [
        r"total em r\$\s*[:\s]*([\d.,]+)",
        r"total a pagar\s*[:\s]*([\d.,]+)",
        r"valor atual\s*[:\s]*([\d.,]+)",
        r"valor atualizado\s*[:\s]*([\d.,]+)",
        r"vl\.?\s*corrigido\s*[:\s]*([\d.,]+)",
    ]:
        m = re.search(pat, text_norm)
        if m:
            valor_atualizado = f"R$ {m.group(1).strip()}"
            break

    # Fallback: si no hay labels, buscar primer valor en R$ en el texto
    if not valor_original:
        m = re.search(r"r\$\s*([\d.,]+)", text_norm)
        if m:
            valor_original = f"R$ {m.group(1).strip()}"

    return valor_original, valor_atualizado


def _extrair_tipo_tributo(text):
    """
    Extrae el tipo de tributo de la execução fiscal.
    Busca en el campo 'Espécie' de la CDA o en el Classe-Assunto.
    """
    # [FIX] Lista ordenada de más específico a menos (evita matches parciales)
    _TRIBUTOS = [
        ("imposto predial territorial urbano",  "IPTU — Imposto Predial e Territorial Urbano"),
        ("imposto predial e territorial urbano", "IPTU — Imposto Predial e Territorial Urbano"),
        ("taxa de licenciamento",               "Taxa de Licenciamento de Estabelecimento"),
        ("taxa de fiscalizacao de funcionamento","TFF — Taxa de Fiscalização de Funcionamento"),
        ("tff",                                  "TFF — Taxa de Fiscalização de Funcionamento"),
        ("cosip",                                "COSIP — Contribuição de Iluminação Pública"),
        ("contribuicao de iluminacao publica",   "COSIP — Contribuição de Iluminação Pública"),
        ("imposto sobre servicos",               "ISS — Imposto sobre Serviços"),
        ("iss",                                  "ISS — Imposto sobre Serviços"),
        ("itbi",                                 "ITBI — Imposto sobre Transmissão de Bens Imóveis"),
        ("imposto sobre transmissao",            "ITBI — Imposto sobre Transmissão de Bens Imóveis"),
        ("iptu",                                 "IPTU — Imposto Predial e Territorial Urbano"),
        ("multa",                                "Multa"),
    ]
    # Fragmentos processuais que NO identifican tributo
    # [FIX] Fragmentos que son categorías processuais o texto de petição, no tributos
    _FRAGMENTOS_IGNORAR = [
        "divida ativa",
        "divida municipal",
        "credito tributario",
        "execucao fiscal",     # solo "execução fiscal" sin tributo especificado
        "devedor",             # fragmentos de texto de petição inicial
        "requerendo",
        "credor",
        "municipio de salvador reu",  # cabeçalho do Classe-Assunto do PJe
        "parte ativa",
    ]
    _RUIDO_LEGAL = ["art.", "lei n", "lei no", "inciso", "paragrafo", "ans."]

    text_norm = normalizar(text)

    for pat in [
        r"esp[eé]cie\s*[:\s]+([^\n\r]{3,60})",
        r"tributo\s*[:\s]+([^\n\r]{3,60})",
        r"classe\s*[-]?\s*assunto\s*[:\s]+([^\n\r]{3,80})",
        r"execu[cç][aã]o fiscal\s*[-\s/]+([^\n\r]{3,30})",
    ]:
        m = re.search(pat, text_norm)
        if m:
            fragmento = normalizar(m.group(1).strip())
            if any(ig in fragmento for ig in _FRAGMENTOS_IGNORAR):
                continue
            for kw, label in _TRIBUTOS:
                if kw in fragmento:
                    return label
            if not any(r in fragmento for r in _RUIDO_LEGAL) and len(fragmento) > 3:
                return fragmento[:60].strip()

    for kw, label in _TRIBUTOS:
        if kw in text_norm:
            return label

    return None


def _extrair_numero_cda(text):
    """
    Extrae el número de la Certidão de Dívida Ativa (CDA).

    Formatos reales observados en PDFs de la PGMS:
    - "CDA n. 12012243416"              → número largo en la CDA emitida por la PGM
    - "CDA nº 65.2014.001323.05155"     → formato con año y livro/folha
    - "Certidão de Dívida Ativa - nº 12012243416"

    [FIX] El número de inscrição municipal (CGA) como "Inscrição 230536" NO es el
    número de CDA — es el registro interno del contribuinte na PGMS/SEFAZ.
    Antes lo capturaba erróneamente. Ahora solo retorna números de CDA reales.
    El número de inscrição se guarda separado si se necesita en el futuro.
    """
    _PATRONES = [
        r"cda\s*n[o°º.]?\s*([\d/.\-]+)",
        r"certid[aã]o de d[ií]vida ativa\s*[-\s]*n[o°º.]?\s*([\d/.\-]+)",
        r"n[o°º.]?\s+da\s+cda\s*[:\s]*([\d/.\-]+)",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            valor = m.group(1).strip().rstrip(".")
            # Descartar si el valor es muy corto (< 5 dígitos)
            if len(re.sub(r"\D", "", valor)) >= 5:
                return valor
    return None


def _extrair_numero_processo(text):
    """
    Extrae el número CNJ del processo (formato NNNNNNN-DD.AAAA.J.TT.OOOO).
    """
    m = re.search(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}", text)
    if m:
        return m.group(0)
    return None


def _extrair_vara(text):
    """
    Extrae la vara/juízo donde tramita el processo.
    """
    _PATRONES = [
        r"(\d+[aª°]\s*vara\s*da\s*fazenda\s*p[uú]blica[^\n\r]{0,40})",
        r"([oó]rg[aã]o julgador\s*[:\s]+[^\n\r]{5,60})",
        r"(vara\s*[^\n\r]{3,40})",
    ]
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:80]
    return None


def _extrair_exercicio(text):
    """
    Extrae el año o período fiscal del tributo.
    [FIX] Prioriza el contexto de la tabla de CDA (Espécie + Exercício en misma línea)
    para evitar capturar años de acórdãos o jurisprudência que aparecen antes.
    """
    text_norm = normalizar(text)

    # Paso 1: contexto de tabla de CDA de la PGMS
    # Formato A: "Espécie: ISS  Exercício: 2006  Meses: 10"  (año en misma línea)
    # Formato B: header "Exercício  Meses  Valor\n2006  10  540,00" (año en línea siguiente)
    m_cda = re.search(
        r"(?:esp[eé]cie|tributo)[^\n]{0,80}exerc[ií]cio\s*[:\s]*(\d{4}(?:[/]\d{4})?)",
        text_norm
    )
    if not m_cda:
        # "Exercício AAAA Meses" — año en misma línea que header
        m_cda = re.search(
            r"exerc[ií]cio\s*[:\s]*(\d{4}(?:[/]\d{4})?)\s+(?:meses|cotas|valor)",
            text_norm
        )
    if not m_cda:
        # Formato B: "exercício  meses  valor do debito\n2006  10  540,00"
        # Si hay múltiples años en filas siguientes, devolver rango.
        m_b = re.search(
            r"exerc[ií]cio\s+meses[^\n]*\n\s*(\d{4}(?:[/]\d{4})?)",
            text_norm
        )
        if m_b:
            bloque_tab = text_norm[m_b.start():]
            anos_tab = re.findall(r"\b(20[012]\d)\s+\d{1,2}\s+[\d.,]+", bloque_tab)
            if anos_tab:
                anos_s = sorted(set(anos_tab))
                return anos_s[0] if len(anos_s) == 1 else f"{anos_s[0]}/{anos_s[-1]}"
            return m_b.group(1).strip()
    if not m_cda:
        # Formato alternativo PGMS: "Exercício: 2006\nMeses: 10\nValor"
        m_cda = re.search(
            r"exerc[ií]cio\s*:\s*(\d{4})\s*\n\s*(?:meses|cotas)",
            text_norm
        )
    if m_cda:
        return m_cda.group(1).strip()

    # Paso 2: "Exercício: 2006" o "Exercício: 2014/2015" standalone
    m = re.search(r"exerc[ií]cio\s*[:\s]*(\d{4}(?:[/\-]\d{4})?)", text_norm)
    if m:
        return m.group(1).strip().replace("-", "/")

    # Paso 3: "exercícios de 2005 e 2006"
    m2 = re.search(r"exerc[ií]cios?\s+(?:de\s+)?(\d{4})\s+e\s+(\d{4})", text_norm)
    if m2:
        return f"{m2.group(1)}/{m2.group(2)}"

    # Paso 4: todos los años en contexto de "exercício" standalone
    todos = re.findall(r"exerc[ií]cio\s+(\d{4})", text_norm)
    if todos:
        anos = sorted(set(todos))
        return anos[0] if len(anos) == 1 else f"{anos[0]}/{anos[-1]}"

    # Paso 5: tabla de CDA — filas con año + meses + valor (sin label "Exercício")
    # Formato: "2010  10  50.000,00" o "2010/2011" en columna de año
    # Buscar años válidos (2000-2030) en contexto de tabla fiscal
    m_tabla = re.search(
        r"(20[012]\d)\s+\d{1,2}\s+[\d.,]+",  # año  meses  valor
        text_norm
    )
    if m_tabla:
        # Buscar todos los años con ese patrón para devolver rango si hay varios
        todos_tabla = re.findall(r"\b(20[012]\d)\s+\d{1,2}\s+[\d.,]+", text_norm)
        if todos_tabla:
            anos = sorted(set(todos_tabla))
            return anos[0] if len(anos) == 1 else f"{anos[0]}/{anos[-1]}"

    return None
def _extrair_data_inscricao(text):
    """
    Extrae la fecha de inscripción en la dívida ativa.

    [FIX] Valida que el resultado sea una fecha plausible (DD/MM/AAAA o DD/MM/AA).
    Descarta resultados garbled de OCR como "06/0821" que tienen longitud anómala.
    """
    _PATRONES = [
        r"data\s*d[ae]\s*inscri[cç][aã]o\s*[:\s]*([\d/\.\-]+)",
        r"inscri[cç][aã]o\s+na\s+d[ií]vida\s+ativa\s*[:\s]*([\d/\.\-]+)",
        r"data\s+de\s+emiss[aã]o\s*[:\s]*([\d/\.\-]+)",
    ]
    _PAT_FECHA_VALIDA = re.compile(
        r"^\d{1,2}[/\.\-]\d{1,2}[/\.\-]\d{2,4}$"
    )
    for pat in _PATRONES:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            candidato = m.group(1).strip()
            # Validar que tenga estructura de fecha (d/m/a o d.m.a)
            if _PAT_FECHA_VALIDA.match(candidato):
                return candidato
    return None


def extract_entidades_agente2(text):
    """
    Punto de entrada principal para extracción de entidades.

    Extrae del texto completo del processo los campos estructurados
    necesarios para el Agente 2 (análisis jurídico-fiscal).

    Devuelve un dict con todos los campos. Los campos no encontrados
    tienen valor None — nunca se inventa información.

    Campos:
        cpf_cnpj         (str|None)  CPF o CNPJ del executado
        nome_executado   (str|None)  Nombre/razón social del executado
        nome_exequente   (str|None)  Nombre del exequente
        valor_original   (str|None)  Valor original de la deuda
        valor_atualizado (str|None)  Valor actualizado de la deuda
        tipo_tributo     (str|None)  Tipo de tributo (ISS, IPTU, TFF, etc.)
        numero_cda       (str|None)  Número de la CDA
        numero_processo  (str|None)  Número CNJ del processo
        vara             (str|None)  Vara/Juízo
        exercicio        (str|None)  Año fiscal del tributo
        data_inscricao   (str|None)  Fecha de inscripción en dívida ativa
    """
    if not text:
        return {k: None for k in [
            "cpf_cnpj", "nome_executado", "nome_exequente",
            "valor_original", "valor_atualizado", "tipo_tributo",
            "numero_cda", "numero_processo", "vara",
            "exercicio", "data_inscricao",
        ]}

    valor_original, valor_atualizado = _extrair_valor(text)

    return {
        "cpf_cnpj"        : _extrair_cpf_cnpj(text),
        "nome_executado"  : _extrair_nome_executado(text),
        "nome_exequente"  : _extrair_nome_exequente(text),
        "valor_original"  : valor_original,
        "valor_atualizado": valor_atualizado,
        "tipo_tributo"    : _extrair_tipo_tributo(text),
        "numero_cda"      : _extrair_numero_cda(text),
        "numero_processo" : _extrair_numero_processo(text),
        "vara"            : _extrair_vara(text),
        "exercicio"       : _extrair_exercicio(text),
        "data_inscricao"  : _extrair_data_inscricao(text),
    }


def extract_parcelamento(text):
    text_norm = normalizar(text)
    if not any(normalizar(k) in text_norm for k in KEYWORDS_PARCELAMENTO_ATIVO):
        return None

    # [FIX Felipao Bug2] Verificar se o PAD está ROMPIDO/CANCELADO.
    # Extrato PAD com "Situação: Rompido" ou extrato fiscal com
    # "Cred. ref. ao cancel. do PAD" confirmam que o PAD foi encerrado.
    # [FIX Ministério Obra Santa] Checar se "suspensão do feito" vem de
    # contexto art.40 LEF (sem bens/devedor), NÃO de PAD.
    # Art.40 usa "suspensão do feito" sem os marcadores específicos de PAD.
    KEYWORDS_ART40_EXCLUSIVOS = [
        "art. 40 da lef",
        "art. 40 da lei 6.830",
        "art. 40 - o juiz suspendera",
        "enquanto nao for localizado o devedor",
        "nao for localizado o devedor",
        "ausencia de localizacao do executado",
        "suspensao do feito pelo prazo de um ano",
        "suspensao pelo prazo de 1 (um) ano",
    ]
    KEYWORDS_PAD_ESPECIFICOS = [
        "art. 151",
        "parcelamento do credito tributario",
        "parcelamento do debito",
        "instrumento de confissao",
        "compromisso de pagamento parcelado",
        "bloq pad",
        "pad n",
        "parcelas mensais e sucessivas",
        "em 38 parcelas",
        "em 12 parcelas",
        "em 24 parcelas",
        "em 36 parcelas",
        "em 48 parcelas",
        "em 60 parcelas",
        "defiro o pedido de suspensao",    # juiz defere PAD explicitamente
        "suspendo/mantenho suspenso",
    ]
    # Se o contexto é art.40 E não há marcadores específicos de PAD → não é PAD
    is_art40_context = any(normalizar(k) in text_norm for k in KEYWORDS_ART40_EXCLUSIVOS)
    has_pad_markers  = any(normalizar(k) in text_norm for k in KEYWORDS_PAD_ESPECIFICOS)
    if is_art40_context and not has_pad_markers:
        return None   # art.40 LEF — não confundir com PAD

    # [FIX JJFM] Verificar primeiro se PAD está ATIVO (Homologado/Bloqueado).
    # Se sim, retornar imediatamente sem checar rompido.
    KEYWORDS_PAD_HOMOLOGADO = [
        "situacao do parcelamento: homologado",
        "situacao pad: homologado",
        "situacao: homologado",
        "bloq pad",                         # extrato: débito bloqueado pelo PAD
        "suspendo/mantenho suspenso",        # decisão judicial deferindo PAD ativo
        "suspendo e mantenho suspenso",
    ]
    if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_HOMOLOGADO):
        return "processo suspenso por parcelamento (PAD)"

    KEYWORDS_PAD_ROMPIDO = [
        "situacao do parcelamento: rompido",
        "situacao: rompido",
        "parcelamento rompido",
        # [FIX JJFM] "parcelamento cancelado" era genérico — batia em
        # PADs antigos mesmo com novo PAD ativo. Usar padrão com data.
        "parcelamento cancelado em",
        "cred. ref. ao cancel. do parc",
        "cred. ref. ao cancel. do pad",
        "motivo: pagamento em atraso",
        "data de rompimento:",
        # [FIX Marcio] Petição de NOVA CITAÇÃO após PAD assinado indica
        # que parcelamento foi cancelado por inadimplência e processo reativado
        "requerer a citacao da parte executada no seguinte endereco",
        "requer a citacao da parte executada no seguinte endereco",
        "citacao da parte executada no seguinte endereco",
    ]
    if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_ROMPIDO):
        # PAD rompido — checar se há NOVO PAD ativo (decisão judicial recente)
        KEYWORDS_PAD_NOVO_ATIVO = [
            "defiro o pedido de suspensao",
            "determino a suspensao do feito",
            "suspendo o feito",
            "suspendo/mantenho suspenso",   # [FIX JJFM] grafia alternativa
            "suspendo e mantenho suspenso",
            "situacao pad: homologado",
            "bloq pad",
        ]
        if any(normalizar(k) in text_norm for k in KEYWORDS_PAD_NOVO_ATIVO):
            return "processo suspenso por parcelamento (PAD)"
        # PAD rompido sem novo PAD ativo — não é situação especial ativa
        return None

    return "processo suspenso por parcelamento (PAD)"
# 6. Crear el prompt
def create_prompt(fecha_reciente, citacion, penhora):
    return PROMPT_TEMPLATE.format(
        fecha    = fecha_reciente.strftime("%d/%m/%Y") if fecha_reciente else "Não especificado",
        citacion=citacion or "Não especificado",
        penhora=penhora or "Não especificado"
    )

# 7. Generar prompts para todos los archivos PDF
def generate_prompts(input_dir):
    pdf_files = [f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')]
    prompts = []

    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_dir, pdf_file)

        # Inicializar ANTES del try
        full_text      = ""
        fecha_reciente = None
        citacion       = None
        penhora        = None
        fecha_orden    = None
        fecha_intento  = None
        fecha_efectiva = None
        prompt         = "Error"
        respuesta_gpt  = None
        ocr_metadata   = {"paginas_ocr": [], "confianza_ocr": {}}
        tipo_processo  = None
        entidades      = None   # dict con entidades para el Agente 2

        try:
            pages_text, ocr_metadata = extract_text_by_page(pdf_path)
            full_text  = " ".join(p for p in pages_text if p)

            if ocr_metadata["paginas_ocr"]:
                n = len(ocr_metadata["paginas_ocr"])
                conf_prom = sum(ocr_metadata["confianza_ocr"].values()) / n
                logging.info(f"  {pdf_file}: {n} página(s) vía OCR, confianza media {conf_prom:.1f}%")

            if not full_text.strip():
                logging.warning(f"PDF sin texto extraíble: {pdf_file}")
                prompts.append((pdf_file, None, None, None, None, None, None, "Error - PDF sin texto", "Error", "", ocr_metadata, tipo_processo, None))
                continue

            tipo_processo = detectar_tipo_processo(full_text)
            if (not tipo_processo["es_execucao_fiscal"]) and tipo_processo["confianza"] == "alta":
                logging.warning(f"  {pdf_file}: FORA DE ESCOPO — {tipo_processo['motivo']}")
                prompts.append((
                    pdf_file, None, None, None, None, None, None,
                    "Fora de escopo - não é execução fiscal", None, full_text, ocr_metadata, tipo_processo, None
                ))
                continue

            fecha_reciente = fecha_ultima_movimentacao(full_text)
            citacion       = extract_citacion(full_text)
            penhora        = extract_penhora(full_text)
            fecha_orden, fecha_intento, fecha_efectiva = extraer_fechas_citacion(full_text)

            # Regla: fecha_citacion_efectiva debe ser None siempre que status_citacion != "HOUVE CITAÇÃO"
            if not citacion or normalizar(citacion) != normalizar("HOUVE CITAÇÃO"):
                fecha_efectiva = None

            # Extracción de entidades para el Agente 2
            entidades = extract_entidades_agente2(full_text)

            prompt = create_prompt(fecha_reciente, citacion, penhora)

            print(f"\n{'='*60}")
            print(f"ARQUIVO : {pdf_file}")
            print(f"  Última fecha   : {fecha_reciente.strftime('%Y-%m-%d') if fecha_reciente else 'NO ENCONTRADA'}")
            print(f"  Citação        : {citacion}")
            print(f"  Penhora        : {penhora}")
            print(f"  CPF/CNPJ       : {entidades.get('cpf_cnpj') or '—'}")
            print(f"  Executado      : {entidades.get('nome_executado') or '—'}")
            print(f"  Valor orig.    : {entidades.get('valor_original') or '—'}")
            print(f"{'='*60}")

            prompts.append((
                pdf_file, fecha_reciente, citacion,
                fecha_orden, fecha_intento, fecha_efectiva,
                penhora, prompt, respuesta_gpt, full_text, ocr_metadata, tipo_processo, entidades
            ))

        except Exception as e:
            logging.error(f"Error al procesar {pdf_file}: {e}", exc_info=True)
            prompts.append((
                pdf_file, None, None, None, None, None, None, "Error", "Error", full_text, ocr_metadata, tipo_processo, None
            ))

    return prompts
client = OpenAI()
MAX_CHARS_FILTRADO = 40_000   # ~10k tokens — intento rápido y barato
MAX_CHARS_FULL     = 400_000  # fallback con texto completo
MAX_CHARS_GPT = 400_000  # ~100k tokens, seguro para gpt-4o-mini (128k ctx)
# Secciones relevantes para filtrar el texto antes de enviarlo a GPT
KEYWORDS_FILTRO_GPT = [
    # Citación
    "citação", "citacao", "cite-se", "citar", "carta", "aviso de recebimento",
    "oficial de justiça", "mandado",
    # Penhora
    "penhora", "bacenjud", "sisbajud", "renajud", "cnib", "bloqueio",
    "indisponibilidade", "arresto", "constrição",
    # Decisiones clave
    "despacho", "decisão", "sentença", "determino", "defiro",
    # Movimentación
    "suspensão", "arquivamento", "extinção", "prescrição"
]


def _filtrar_texto_relevante(full_text):
    """Extrae solo los párrafos que contienen keywords relevantes."""
    keywords_norm = [normalizar(k) for k in KEYWORDS_FILTRO_GPT]
    parrafos = []
    for parrafo in full_text.split("\n\n"):
        parrafo_norm = normalizar(parrafo)
        if any(kw in parrafo_norm for kw in keywords_norm):
            parrafos.append(parrafo.strip())
    return "\n\n".join(parrafos)

def _parsear_respuesta_gpt(respuesta):
    """Extrae decisión y motivo de la respuesta estructurada del GPT."""
    decision = motivo = None
    for linea in respuesta.splitlines():
        if linea.startswith("DECISÃO:"):
            decision = linea.split(":", 1)[1].strip()
        elif linea.startswith("MOTIVO:"):
            motivo = linea.split(":", 1)[1].strip()
    return decision, motivo

def call_chatgpt(full_text, fecha, citacion, penhora):

    """

    GPT deshabilitado temporalmente.

    Mantiene la misma interfaz para no modificar el resto del código.

    """

    print("  [GPT] Deshabilitado.")

    return (

        "DECISÃO: CHAMADA PENDENTE\n"

        "JUSTIFICATIVA: GPT desabilitado no momento."

    )# 8. Guardar prompts en un archivo de texto
def save_prompts_to_file(prompts, output_file):
    with open(output_file, 'w', encoding='utf-8') as file:
        for pdf_file, _, _, _, _, _, _, prompt, respuesta, _, ocr_metadata, tipo_processo, _ in prompts:
            paginas_ocr = ocr_metadata.get("paginas_ocr", []) if ocr_metadata else []
            ocr_info = f"Páginas via OCR: {paginas_ocr}\n" if paginas_ocr else ""
            tipo_info = f"Tipo de processo: {tipo_processo['motivo']}\n" if tipo_processo else ""
            file.write(f"Archivo: {pdf_file}\n{tipo_info}{ocr_info}Prompt:\n{prompt}\nRespuesta GPT:\n{respuesta}\n{'-'*50}\n")
    print(f"Prompts guardados en {output_file}")

# 9. Procesar prompts y generar Excel
def process_prompts_to_excel(prompts, output_excel):
    resultados = []
    hoy = datetime.now()

    for pdf_file, fecha_reciente, citacion, fecha_orden, fecha_intento, fecha_efectiva, penhora, prompt, respuesta_gpt, full_text, ocr_metadata, tipo_processo, entidades in prompts:
        extincao      = extract_extincao(full_text)      # [FIX] Passo 2: extinção
        parcelamento  = extract_parcelamento(full_text)
        suspensao_art40 = extract_suspensao_art40(full_text)
        # Inicializar siempre antes del bloque de decisión
        respuesta_gpt = None
        fonte = "Sistema automático"

        # [NUEVO] Corte temprano: si el filtro de tipo de documento ya
        # detectó con alta confianza que esto no es una execução fiscal,
        # no tiene sentido aplicar el resto del árbol (citação/penhora)
        # ni escalar a GPT.
        fora_de_escopo = (
            tipo_processo is not None
            and not tipo_processo["es_execucao_fiscal"]
            and tipo_processo["confianza"] == "alta"
        )

        if fora_de_escopo:
            decision = "FORA DE ESCOPO"
            motivo = tipo_processo["motivo"]
            fonte = "Filtro de tipo de documento"
        elif fecha_reciente is None:
            decision = "Informação insuficiente"
            motivo = "Data da última movimentação não informada"
        elif (hoy - fecha_reciente).days < 365:
            decision = "NO APTO"
            motivo = "Movimentação recente (< 1 ano)"
        elif extincao:                              # [FIX] Passo 2: processo extinto
            decision = "NO APTO"
            motivo = f"{extincao} — execução encerrada por sentença ou prescrição"
        elif parcelamento:
            decision = "NO APTO"
            motivo = f"Processo suspenso — {parcelamento}. Verificar status atual do PAD."
        elif suspensao_art40:
            decision = "APTO"
            motivo   = f"{suspensao_art40} — sem bens localizados para penhora"
        else:
            citacao_ausente = (
                not citacion
                or "não encontrado" in citacion.lower()
                or "nao houve" in normalizar(citacion)
                or "tentativa falha" in normalizar(citacion)
            )
            if citacao_ausente:
                decision = "APTO"
                motivo = "Citação ausente ou não localizada"
            else:
                _p = penhora.lower() if penhora else ""
                if not penhora or "não encontrado" in _p or "nao encontrado" in _p:
                    decision = "APTO"
                    motivo = "Ausência de penhora"
                elif "penhora não realizada" in _p or "penhora nao realizada" in _p:
                    decision = "APTO"
                    motivo = "Penhora não realizada"
                elif "tentativa" in _p:
                    decision = "APTO"
                    motivo = f"Apenas tentativa de penhora: {penhora}"
                elif "indisponibilidade de bens" in _p:
                    decision = "NO APTO"
                    motivo = "Indisponibilidade de bens (CNIB) decretada"
                elif "bloqueio renajud ativo" in _p:
                    decision = "NO APTO"
                    motivo = "Bloqueio RENAJUD ativo — penhora de veículo em andamento"
                elif "penhora de imóvel" in _p or "penhora de imovel" in _p:
                    decision = "NO APTO"
                    motivo = "Penhora de imóvel efetivada"
                elif "penhora de faturamento" in _p:
                    decision = "NO APTO"
                    motivo = "Penhora de faturamento efetivada"
                elif "penhora de quotas" in _p or "penhora de acoes" in _p:
                    decision = "NO APTO"
                    motivo = "Penhora de quotas/ações efetivada"
                elif "penhora de créditos" in _p or "penhora de creditos" in _p:
                    decision = "NO APTO"
                    motivo = "Penhora de créditos/direitos efetivada"
                elif "penhora bacenjud" in _p or "penhora sisbajud" in _p:
                    decision = "NO APTO"
                    motivo = "Penhora efetivada via BacenJud/SisBajud"
                elif "solicitado (resultado desconhecido)" in _p:
                    decision = "APTO"
                    motivo = "SISBAJUD/BacenJud apenas solicitado em petição — sem resultado confirmado no processo"
                else:
                    decision = "Informação insuficiente"
                    motivo = "Não foi possível determinar regra aplicável"

        # Llamar a GPT solo si el sistema no pudo decidir
        if decision == "Informação insuficiente":
            fecha_str = fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else "Não especificado"
            try:
                respuesta_gpt = call_chatgpt(
                    full_text,
                    fecha_str,
                    citacion or "Não especificado",
                    penhora  or "Não especificado"
                )
                # Parsear la respuesta para actualizar decisión y motivo
                decision_gpt, motivo_gpt = _parsear_respuesta_gpt(respuesta_gpt)
                if decision_gpt:
                    decision = decision_gpt
                    motivo   = motivo_gpt or motivo
                    fonte    = "GPT (paso 1 - filtrado)" if "paso 1" in respuesta_gpt else "GPT (paso 2 - completo)"

            except Exception as e:
                logging.error(f"Error GPT para {pdf_file}: {e}")
                respuesta_gpt = "Error en API"

        paginas_ocr = ocr_metadata.get("paginas_ocr", []) if ocr_metadata else []
        confianza_ocr_dict = ocr_metadata.get("confianza_ocr", {}) if ocr_metadata else {}
        confianza_ocr_media = (
            round(sum(confianza_ocr_dict.values()) / len(confianza_ocr_dict), 1)
            if confianza_ocr_dict else None
        )

        # Extraer campos de entidades (con fallback si entidades es None)
        ent = entidades or {}

        resultados.append({
            # ── Campos originales del Agente 1 ──────────────────────────────
            "CASO"                    : pdf_file,
            "Última data de interação": fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else "Não especificado",
            "Status da citação"       : citacion or "Não especificado",
            "Fecha orden citación"    : fecha_orden.strftime("%Y-%m-%d")   if fecha_orden    else "Não especificado",
            "Fecha intento citación"  : fecha_intento.strftime("%Y-%m-%d") if fecha_intento  else "Não especificado",
            "Fecha citación efectiva" : fecha_efectiva.strftime("%Y-%m-%d") if fecha_efectiva else "Não especificado",
            "Resultado da penhora"    : penhora or "Não especificado",
            "Decisión"                : decision,
            "Motivo"                  : motivo,
            "Fonte da decisão"        : fonte,
            "Respuesta GPT"           : respuesta_gpt or "",
            "Páginas via OCR"         : ", ".join(map(str, paginas_ocr)) if paginas_ocr else "",
            "Confiança OCR (%)"       : confianza_ocr_media if confianza_ocr_media is not None else "",
            "Tipo de Processo"        : (tipo_processo["classe_assunto"] or "(não detectado)") if tipo_processo else "",
            "Confiança Tipo Processo" : tipo_processo["confianza"] if tipo_processo else "",
            # ── Entidades para el Agente 2 ───────────────────────────────────
            "A2_numero_processo"      : ent.get("numero_processo") or "",
            "A2_cpf_cnpj"            : ent.get("cpf_cnpj") or "",
            "A2_nome_executado"      : ent.get("nome_executado") or "",
            "A2_nome_exequente"      : ent.get("nome_exequente") or "",
            "A2_tipo_tributo"        : ent.get("tipo_tributo") or "",
            "A2_exercicio"           : ent.get("exercicio") or "",
            "A2_numero_cda"          : ent.get("numero_cda") or "",
            "A2_data_inscricao"      : ent.get("data_inscricao") or "",
            "A2_valor_original"      : ent.get("valor_original") or "",
            "A2_valor_atualizado"    : ent.get("valor_atualizado") or "",
            "A2_vara"                : ent.get("vara") or "",
        })

    df = pd.DataFrame(resultados)
    # [FIX] Columnas que deben guardarse como texto puro en Excel
    # Estrategia: escribir con openpyxl y forzar data_type='s' (string)
    # en las celdas de esas columnas, evitando que Excel las interprete
    # como números aunque sean dígitos puros.
    cols_texto = ["A2_numero_cda", "A2_cpf_cnpj", "A2_numero_processo"]
    for col in cols_texto:
        if col in df.columns:
            df[col] = df[col].astype(str).replace({"nan": "", "None": ""})

    # ── Ordenar por decisión (APTO primero) para lectura rápida ──
    # Mismo criterio que el reporte del Agente 2 (ALTA→MEDIA→BAIXA):
    # lo accionable arriba. Orden estable: dentro de cada grupo se
    # mantiene el orden de procesamiento de los PDFs.
    _ORDEN_DECISION = {
        "APTO": 0, "Informação insuficiente": 1, "NO APTO": 2, "FORA DE ESCOPO": 3,
    }
    if not df.empty and "Decisión" in df.columns:
        df["_ord_dec"] = df["Decisión"].map(lambda d: _ORDEN_DECISION.get(d, 9))
        df = df.sort_values(by="_ord_dec", kind="stable").drop(columns=["_ord_dec"])

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resultados")
        ws = writer.sheets["Resultados"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ── Estilos (mismo layout del reporte del Agente 2) ──
        font_header  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        fill_header  = PatternFill("solid", fgColor="1F4E79")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde        = Border(*[Side(style="thin", color="D0D0D0")] * 4)

        fills_decision = {
            "APTO"                   : PatternFill("solid", fgColor="C6E0B4"),  # verde suave
            "Informação insuficiente": PatternFill("solid", fgColor="FFE699"),  # amarillo suave
            "NO APTO"                : PatternFill("solid", fgColor="F8CBAD"),  # rojo suave
            "FORA DE ESCOPO"         : PatternFill("solid", fgColor="D9D9D9"),  # gris
        }

        # Header
        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font      = font_header
            c.fill      = fill_header
            c.alignment = align_header
            c.border    = borde

        # Filas de datos
        col_decision = (list(df.columns).index("Decisión") + 1) if "Decisión" in df.columns else None
        font_dato    = Font(name="Arial", size=10)
        align_dato   = Alignment(vertical="top", wrap_text=True)
        for row_idx in range(2, len(df) + 2):
            fill = None
            if col_decision:
                fill = fills_decision.get(ws.cell(row=row_idx, column=col_decision).value)
            for col_idx in range(1, len(df.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font      = font_dato
                c.border    = borde
                c.alignment = align_dato
                if fill and col_idx == col_decision:
                    c.fill = fill

        # Forzar texto puro en las columnas de identificadores
        for col_name in cols_texto:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter][1:]:  # saltar header
                    if cell.value:
                        # Forzar tipo string explícitamente en openpyxl
                        cell.value = str(cell.value)
                        cell.data_type = "s"
                        cell.number_format = "@"

        # Anchos de columna
        anchos = {
            "CASO": 34, "Última data de interação": 16, "Status da citação": 30,
            "Fecha orden citación": 14, "Fecha intento citación": 14,
            "Fecha citación efectiva": 14, "Resultado da penhora": 30,
            "Decisión": 16, "Motivo": 45, "Fonte da decisão": 20,
            "Respuesta GPT": 50, "Páginas via OCR": 14, "Confiança OCR (%)": 12,
            "Tipo de Processo": 28, "Confiança Tipo Processo": 12,
            "A2_numero_processo": 22, "A2_cpf_cnpj": 20, "A2_nome_executado": 30,
            "A2_nome_exequente": 30, "A2_tipo_tributo": 18, "A2_exercicio": 12,
            "A2_numero_cda": 22, "A2_data_inscricao": 14,
            "A2_valor_original": 16, "A2_valor_atualizado": 16, "A2_vara": 28,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = anchos.get(col_name, 18)

        # Congelar header y activar autofiltro
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

    print(f"Planilla Excel generada: {output_excel}")
# ===========================================================================
# 7.1 — SALIDA ESTRUCTURADA PARA EL AGENTE 2
# ===========================================================================

def exportar_json_agente2(prompts, resultados_excel, output_json):
    """
    Genera el JSON de interfaz entre Agente 1 y Agente 2.
    Solo incluye procesos con Decisión == "APTO".
    Los campos vacíos se guardan como null — nunca se omiten.
    """
    import json
    from datetime import datetime as _dt

    VERSION_AGENTE1 = "7.1"

    idx = {r["CASO"]: r for r in resultados_excel} if resultados_excel else {}
    processos_apto = []

    for tupla in prompts:
        (pdf_file, fecha_reciente, citacion, fecha_orden, fecha_intento,
         fecha_efectiva, penhora, prompt, respuesta_gpt, full_text,
         ocr_metadata, tipo_processo, entidades) = tupla

        res      = idx.get(pdf_file, {})
        decision = res.get("Decisión", "")
        motivo   = res.get("Motivo", "")
        fonte    = res.get("Fonte da decisão", "")

        if decision.upper() != "APTO":
            continue

        ent = entidades or {}

        def _nulo(val):
            if val is None:
                return None
            v = str(val).strip()
            return None if v in ("", "nan", "None", "Não especificado") else v

        ocr_meta   = ocr_metadata or {}
        pags_ocr   = ocr_meta.get("paginas_ocr", [])
        conf_dict  = ocr_meta.get("confianza_ocr", {})
        conf_media = (
            round(sum(conf_dict.values()) / len(conf_dict), 1)
            if conf_dict else None
        )

        processos_apto.append({
            "id_lote"                : pdf_file,
            "decisao_agente1"        : decision,
            "motivo_agente1"         : _nulo(motivo),
            "fonte_decisao"          : _nulo(fonte),
            "ultima_movimentacao"    : fecha_reciente.strftime("%Y-%m-%d") if fecha_reciente else None,
            "status_citacao"         : _nulo(citacion),
            "resultado_penhora"      : _nulo(penhora),
            "paginas_ocr"            : pags_ocr,
            "confianca_ocr_media"    : conf_media,
            "tipo_processo_confianca": tipo_processo.get("confianza") if tipo_processo else None,
            "entidades": {
                "numero_processo"  : _nulo(ent.get("numero_processo")),
                "cpf_cnpj"         : _nulo(ent.get("cpf_cnpj")),
                "nome_executado"   : _nulo(ent.get("nome_executado")),
                "nome_exequente"   : _nulo(ent.get("nome_exequente")),
                "tipo_tributo"     : _nulo(ent.get("tipo_tributo")),
                "exercicio"        : _nulo(ent.get("exercicio")),
                "numero_cda"       : _nulo(ent.get("numero_cda")),
                "data_inscricao"   : _nulo(ent.get("data_inscricao")),
                "valor_original"   : _nulo(ent.get("valor_original")),
                "valor_atualizado" : _nulo(ent.get("valor_atualizado")),
                "vara"             : _nulo(ent.get("vara")),
            }
        })

    payload = {
        "metadata": {
            "generado_em"      : _dt.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "total_procesados" : len(prompts),
            "total_aptos"      : len(processos_apto),
            "version_agente1"  : VERSION_AGENTE1,
        },
        "processos": processos_apto,
    }

    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"JSON Agente 2 generado: {output_json}")
    print(f"  {len(processos_apto)} processo(s) APTO(s) de {len(prompts)} procesados")
    return payload
# 10. Ejecutar el flujo
if __name__ == "__main__":
    prompts = generate_prompts(input_directory)
    save_prompts_to_file(prompts, output_file_prompts)

    # Generar Excel (Agente 1 — revisión humana)
    process_prompts_to_excel(prompts, output_file_excel)

    # [7.1] Generar JSON para el Agente 2 (solo procesos APTO)
    import pandas as _pd
    _df = _pd.read_excel(output_file_excel)
    resultados_para_json = _df.to_dict(orient="records")
    exportar_json_agente2(prompts, resultados_para_json, output_file_json)