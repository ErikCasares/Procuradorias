"""
Isolamento entre consumidores nos ARTEFATOS — o que carrega nome, CPF/CNPJ e
valor de dívida. É a regressão que o contrato com a PGMS não pode ter.
"""
import io
import json
from datetime import datetime

import pytest

# Lotes semeados com a data de AGORA. Com data fixa no passado, a retenção
# (que roda ao fim de cada lote processado) os apagava no meio do teste — e a
# falha aparecia só quando um upload de outro arquivo de teste terminava junto.
AGORA = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def semear(app_mod, lote_id, origem, numero, nome_exec, cpf):
    """Injeta um lote concluído com histórico, sem precisar rodar os agentes."""
    app_mod._lotes[lote_id] = {
        "id": lote_id, "status": "concluido", "origem": origem,
        "criado_em": AGORA, "iniciado_em": AGORA,
        "concluido_em": AGORA, "arquivos": [lote_id + ".pdf"],
        "etapa": "concluído", "log": [], "analises": [], "avisos": [],
        "resumo": None, "erro": None, "totais": None, "planilha": None,
    }
    app_mod.PASTA_JSON.mkdir(parents=True, exist_ok=True)

    linha = {
        "numero_processo": numero,
        "nome_executado": nome_exec,
        "id_lote": lote_id + ".pdf",
        "origem_lote": "lote_" + lote_id + "_agente2.json",
        "processado_em": "2026-01-01T00:00:02",
        "analise": {
            "prioridade": "ALTA",
            "acao_recomendada": "Executar",
            "justificativa": "valor da dívida: R$ 1.000,00",
            "observacoes": [],
        },
        "agente1": {"entidades": {"numero_processo": numero, "cpf_cnpj": cpf,
                                  "nome_executado": nome_exec}},
    }
    with open(app_mod.PASTA_JSON / "historial_agente2.jsonl", "a", encoding="utf-8") as f:
        f.write(json.dumps(linha, ensure_ascii=False))
        f.write("\n")

    proc = {"arquivo": lote_id + ".pdf",
            "entidades": {"numero_processo": numero, "cpf_cnpj": cpf,
                          "nome_executado": nome_exec}}
    alvo = app_mod.PASTA_JSON / ("lote_" + lote_id + "_agente2.json")
    with open(alvo, "w", encoding="utf-8") as f:
        json.dump({"processos": [proc]}, f, ensure_ascii=False)

    reg = {"extraido_em": "2026-01-01T00:00:01", "arquivo": lote_id + ".pdf",
           "numero_processo": numero, "nome_executado": nome_exec, "cpf_cnpj": cpf}
    with open(app_mod.PASTA_JSON / "historico_extracoes.jsonl", "a", encoding="utf-8") as f:
        f.write(json.dumps(reg, ensure_ascii=False))
        f.write("\n")


@pytest.fixture(scope="module", autouse=True)
def dados(app_mod):
    semear(app_mod, "AAA111", "siap",  "1111111-11.2011.8.05.0001", "EMPRESA ALFA LTDA", "11.111.111/0001-11")
    semear(app_mod, "BBB222", "outro", "2222222-22.2022.8.05.0002", "EMPRESA BETA SA",   "22.222.222/0002-22")


NUM_A = "1111111-11.2011.8.05.0001"
NUM_B = "2222222-22.2022.8.05.0002"


def test_cada_consumidor_acha_o_proprio(client, auth, auth2):
    assert client.get("/api/v1/processos", params={"numero": NUM_A}, headers=auth).status_code == 200
    assert client.get("/api/v1/processos", params={"numero": NUM_B}, headers=auth2).status_code == 200


def test_consulta_nao_alcanca_processo_alheio(client, auth, auth2):
    r = client.get("/api/v1/processos", params={"numero": NUM_B}, headers=auth)
    assert r.status_code == 404, "VAZOU processo de outro consumidor:\n" + r.text[:400]
    assert "BETA" not in r.text and "22.222.222" not in r.text

    r = client.get("/api/v1/processos", params={"numero": NUM_A}, headers=auth2)
    assert r.status_code == 404, "VAZOU processo de outro consumidor:\n" + r.text[:400]
    assert "ALFA" not in r.text and "11.111.111" not in r.text


def test_planilha_agente2_recortada_ao_consumidor(client, auth, auth2):
    import openpyxl

    def texto_da_planilha(resp):
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        return "\n".join(
            str(c.value) for ws in wb.worksheets
            for row in ws.iter_rows() for c in row if c.value is not None
        )

    r = client.get("/api/v1/lotes/AAA111/planilha/agente2", headers=auth)
    assert r.status_code == 200, r.text
    t = texto_da_planilha(r)
    assert "ALFA" in t, "faltou o processo do próprio consumidor"
    assert "BETA" not in t, "planilha de 'siap' contém processo de 'outro'"

    r = client.get("/api/v1/lotes/BBB222/planilha/agente2", headers=auth2)
    assert r.status_code == 200, r.text
    t = texto_da_planilha(r)
    assert "BETA" in t
    assert "ALFA" not in t, "planilha de 'outro' contém processo de 'siap'"


def test_download_por_tipo_tambem_recortado(client, auth):
    import openpyxl
    r = client.get("/api/v1/lotes/AAA111/arquivos/agente2_planilha", headers=auth)
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    t = "\n".join(str(c.value) for ws in wb.worksheets
                  for row in ws.iter_rows() for c in row if c.value is not None)
    assert "BETA" not in t, "rota /arquivos/{tipo} vaza processo alheio"


def test_planilha_deduplicada(client, app_mod, auth):
    """Reprocessar o mesmo processo não pode duplicar a linha no relatório."""
    import openpyxl
    semear(app_mod, "AAA111", "siap", NUM_A, "EMPRESA ALFA LTDA", "11.111.111/0001-11")
    semear(app_mod, "AAA111", "siap", NUM_A, "EMPRESA ALFA LTDA", "11.111.111/0001-11")

    r = client.get("/api/v1/lotes/AAA111/planilha/agente2", headers=auth)
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
    numeros = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
    assert numeros.count(NUM_A) == 1, f"processo duplicado no relatório: {numeros}"
