"""
AGENTE 2 — Raciocínio Jurídico-Fiscal
HERA Tecnologia / PGMS — Contrato nº 01/2026

Responsabilidade:
    Recebe TODOS os processos do Agente 1 via JSON e executa análise
    jurídico-fiscal para priorização de cobrança.
    (v0.3: sem triagem APTO/NÃO APTO — a análise roda sobre todos.)

Modo de operação:
    File watcher — monitora a pasta de saída do Agente 1 e processa
    automaticamente qualquer arquivo *_agente2.json novo ou modificado.

Uso:
    # Modo watcher (produção) — fica rodando em background
    python agente2.py --watch

    # Modo pontual — processa um JSON específico e encerra
    python agente2.py --arquivo resultados_procesosV7_agente2.json

    # Modo watcher com pasta explícita
    python agente2.py --watch --pasta /caminho/para/outputs/

Dependências:
    Apenas stdlib — sem dependências externas para o watcher.
    (watchdog opcional para produção de alta performance)

Migração a Gemini:
    A função analisar_processo() tem un bloco marcado [GEMINI PLACEHOLDER].
    Quando chegarem as credenciais de SEMIT, só substituir esse bloco.
    A estrutura de entrada/saída não muda.
"""

import os
import re
import sys
import json
import time
import logging
import argparse
from datetime import datetime
from pathlib import Path

# ── Logging ──────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [Agente2] %(levelname)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("agente2")

# ── Configuración ─────────────────────────────────────────────────
CURRENT_DIR       = os.path.dirname(os.path.abspath(__file__))

# Sobreescribibles por variable de entorno — ver la nota equivalente en promptV7.1.py.
# El historial acumulativo vive en PASTA_JSON, así que la API deja esta apuntando
# siempre a la carpeta compartida: es lo que mantiene el Excel del procurador
# acumulando entre lotes en vez de reiniciarse en cada uno.
PASTA_JSON       = os.environ.get("PASTA_JSON")       or os.path.join(CURRENT_DIR, "JSON")         # carpeta donde el Agente 1 escribe los JSON
PASTA_RESULTADOS  = os.environ.get("PASTA_RESULTADOS") or os.path.join(CURRENT_DIR, "resultados")   # carpeta de salida del Agente 2
SUFIJO_INPUT      = "_agente2.json"       # archivos que genera el Agente 1
SUFIJO_OUTPUT     = "_agente2_resultado.json"   # resultado por lote
HISTORIAL_JSONL   = "historial_agente2.jsonl"   # registro acumulativo — una línea por processo
REPORTE_XLSX      = "historial_agente2.xlsx"    # reporte Excel acumulativo para el procurador
INTERVALO_WATCH   = 10                          # segundos entre chequeos del watcher

# ── Limiares de prioridade OPERACIONAL ────────────────────────────────
# [FIX] Estas constantes eram usadas em _calcular_prioridad() mas nunca
#       tinham sido definidas → NameError ('LIMIAR_PRIORIDADE_ALTA' is not
#       defined) que derrubava a análise de todo processo sem risco de
#       prescrição. Valores conforme o docstring de _calcular_prioridad.
#
# ASSUNÇÃO PROVISÓRIA (definição HERA, NÃO jurídica): R$ 5.000 / R$ 1.000
# foram escolhidos pela equipe apenas para ORDENAR o trabalho do procurador;
# não derivam de norma nem definem ajuizamento. Substituir pelo valor mínimo
# de ajuizamento oficial da PGMS quando definido — basta trocar a variável de
# ambiente, sem tocar no fluxo. >>> CONFIRMAR limiares com a PGMS. <<<
LIMIAR_PRIORIDADE_ALTA  = float(os.environ.get("LIMIAR_PRIORIDADE_ALTA",  "5000"))
LIMIAR_PRIORIDADE_MEDIA = float(os.environ.get("LIMIAR_PRIORIDADE_MEDIA", "1000"))

VERSION_AGENTE2 = "0.3"


# ════════════════════════════════════════════════════════════════════
# HISTORIAL JSONL — registro acumulativo sin duplicados
# ════════════════════════════════════════════════════════════════════
#
# El archivo historial_agente2.jsonl vive en la misma carpeta que
# el script. Cada línea es un JSON completo de un processo analizado.
# Formato de cada línea:
# {"numero_processo":"...","id_lote":"...","analise":{...},"processado_em":"..."}
#
# Deduplicación: se usa numero_processo (número CNJ) como clave única.
# Si llega un processo ya registrado, se actualiza su línea en lugar
# de agregar una nueva — esto evita duplicados entre lotes.
#
# Lectura del historial: cada línea es JSON válido independiente,
# se puede leer con: for line in f: record = json.loads(line)
# ════════════════════════════════════════════════════════════════════

def _ruta_historial(pasta: str) -> str:
    """
    Devuelve la ruta absoluta del archivo historial JSONL
    dentro de la carpeta resultados/ (se crea si no existe).
    El parámetro pasta se ignora — siempre usa PASTA_RESULTADOS.
    """
    os.makedirs(PASTA_JSON, exist_ok=True)
    return os.path.join(PASTA_JSON, HISTORIAL_JSONL)

def _cargar_processos_registrados(path_jsonl: str) -> dict:
    """
    Lee el historial JSONL y devuelve un dict {numero_processo: número_de_línea}.
    Si el archivo no existe, devuelve dict vacío.
    Líneas corruptas se saltan con warning.
    """
    registrados = {}
    if not os.path.exists(path_jsonl):
        return registrados
    with open(path_jsonl, encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                np = rec.get("numero_processo")
                if np:
                    registrados[np] = i
            except json.JSONDecodeError:
                log.warning(f"Linha {i} corrompida no histórico — ignorada")
    return registrados

def _escribir_historial(path_jsonl: str, resultados: list, origen: str):
    """
    Anexa ao historial JSONL uma linha por processo analisado — APPEND-ONLY.
    NÃO reescreve linhas antigas: guarda a história completa das análises.
 
    O contador vez_analisada reflete quantas vezes o mesmo numero_processo
    já foi analisado. Útil quando a recomendação muda entre análises — por
    evolução do processo (documentos juntados) ou por não-determinismo do
    LLM quando o Gemini entrar.
 
    A deduplicação ('último estado por processo') deixou de acontecer aqui:
    passou a ser uma VISTA calculada na geração do Excel — ver
    _ultimo_estado_por_processo().
 
    Devolve (novos, reanalises) — conteo para o log.
    (Antes devolvia (nuevos, actualizados); ajustar o texto no call site.)
    """
    # Contagem prévia por numero_processo (para vez_analisada)
    contagens = {}
    if os.path.exists(path_jsonl):
        with open(path_jsonl, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue  # linha corrompida — ignorada na contagem
                np = rec.get("numero_processo")
                if np:
                    contagens[np] = contagens.get(np, 0) + 1
 
    novos      = 0
    reanalises = 0
    linhas     = []
    contador   = dict(contagens)   # contagem acumulada (disco + esta corrida)
    vistos     = set()             # numero_processo já vistos nesta corrida
 
    for r in resultados:
        np = (r.get("entidades") or {}).get("numero_processo") or r.get("numero_processo")
        if not np:
            log.warning(
                f"  Processo sem número — agregado sem deduplicar: "
                f"{r.get('id_lote') or '?'}"
            )
            np = f"SEM_NUMERO_{r.get('id_lote') or 'origem_desconhecida'}"
 
        vez = contador.get(np, 0) + 1
        contador[np] = vez
 
        registro = {
            "numero_processo" : np,
            "id_lote"         : r.get("id_lote"),
            "nome_executado"  : r.get("nome_executado"),
            "agente1"         : r.get("agente1"),
            "vez_analisada"   : vez,
            "analise"         : r.get("analise", {}),
            "processado_em"   : r.get("processado_em"),
            "origem_lote"     : origen,
        }
        linhas.append(json.dumps(registro, ensure_ascii=False))
 
        if np in contagens or np in vistos:
            reanalises += 1
        else:
            novos += 1
        vistos.add(np)
 
    if linhas:
        # Append em bloco, modo 'a' — nunca toca linhas antigas.
        try:
            with open(path_jsonl, "a", encoding="utf-8") as f:
                f.write("\n".join(linhas) + "\n")
        except OSError as e:
            # Falha VISÍVEL: loga claro e propaga. O caller decide.
            log.error(f"Falha ao gravar o histórico {path_jsonl}: {e}")
            raise
 
    return novos, reanalises

def _ultimo_estado_por_processo(path_jsonl: str) -> list:
    """
    Lê o historial APPEND-ONLY e devolve a VISTA deduplicada:
    um registro por numero_processo (o MAIS RECENTE), com o campo
    'vez_analisada' = total de análises desse processo no historial.
 
    Mantém a ordem de primeira aparição. É o que alimenta o Excel:
    uma linha por processo, mostrando o último estado + quantas vezes
    foi reanalisado. Linhas corrompidas são puladas com warning.
    """
    ultimo = {}   # numero_processo -> registro mais recente
    total  = {}   # numero_processo -> contagem de análises
    orden  = []   # ordem de primeira aparição
 
    if not os.path.exists(path_jsonl):
        return []
 
    with open(path_jsonl, encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                log.warning(f"Histórico: linha {i} corrompida — ignorada")
                continue
            np = rec.get("numero_processo")
            if not np:
                continue
            if np not in total:
                orden.append(np)
            total[np] = total.get(np, 0) + 1
            ultimo[np] = rec
 
    saida = []
    for np in orden:
        rec = dict(ultimo[np])
        rec["vez_analisada"] = total[np]   # garante o total real, não o da última linha
        saida.append(rec)
    return saida

# ════════════════════════════════════════════════════════════════════
# NÚCLEO DE ANÁLISIS
# ════════════════════════════════════════════════════════════════════

def analisar_processo(processo: dict) -> dict:
    """
    Analisa um processo do Agente 1 e devolve recomendações jurídico-fiscais.

    A partir da v0.3 roda sobre TODOS os processos do Agente 1 — não há mais
    triagem APTO/NÃO APTO. O esquema de entrada não traz decisao_agente1 nem
    motivo_agente1; os sinais de estratégia vêm dos próprios campos do
    processo (status_citacao, resultado_penhora, sinais_processuais).

    Input  — dict com a estrutura do JSON do Agente 1.
    Output — dict com análise, prioridade, ação recomendada e alertas.

    [GEMINI PLACEHOLDER]
    El bloque marcado abajo será reemplazado por la llamada a Gemini
    cuando lleguen las credenciales de SEMIT. La estructura de retorno
    no cambia.
    """
    ent  = processo.get("entidades", {})
    tipo = processo.get("tipo_processo", {})
    ocr  = processo.get("ocr", {})

    # OCR: no esquema novo a confiança vem aninhada em ocr.confianca_media
    # (antes era top-level confianca_ocr_media). Fallback por compatibilidade.
    confianca_ocr = processo.get("confianca_ocr_media")
    if confianca_ocr is None:
        confianca_ocr = ocr.get("confianca_media")

    # Snapshot dos dados do Agente 1 — arrastado ao historial para que a
    # busca mostre a triagem/entidades mesmo se o JSON do Agente 1 já tiver
    # sido sobrescrito por uma corrida posterior.
    snapshot_a1 = {
        "tipo_processo"      : tipo or None,
        "status_citacao"     : processo.get("status_citacao"),
        "resultado_penhora"  : processo.get("resultado_penhora"),
        "sinais_processuais" : processo.get("sinais_processuais", {}),
        "ultima_movimentacao": processo.get("ultima_movimentacao"),
        "dias_desde_ultima_movimentacao": processo.get("dias_desde_ultima_movimentacao"),
        "confianca_ocr_media": confianca_ocr,
        "entidades"          : ent,
    }

    # Análise jurídico-fiscal — roda para TODOS os processos.
    prioridad    = _calcular_prioridad(processo)
    accion       = _recomendar_accion(processo)
    alerta_presc = _verificar_prescricao(processo)
    observacoes  = _generar_observacoes(processo)

    # ── [GEMINI PLACEHOLDER] ──────────────────────────────────────
    # Reemplazar este bloque cuando lleguen credenciales SEMIT:
    #
    # import vertexai
    # from vertexai.generative_models import GenerativeModel
    # vertexai.init(project=GEMINI_PROJECT_ID, location=GEMINI_REGION)
    # model = GenerativeModel("gemini-pro")
    # prompt = _construir_prompt_gemini(processo, prioridad, accion)
    # response = model.generate_content(prompt)
    # analise_llm = _parsear_resposta_gemini(response.text)
    # accion      = analise_llm.get("acao_recomendada", accion)
    # observacoes = analise_llm.get("observacoes", observacoes)
    # ──────────────────────────────────────────────────────────────

    return {
        # 'arquivo' é como o Agente 1 chama o PDF de origem; 'id_lote' era o
        # nome antigo. Sem este fallback a procedência some do relatório.
        "id_lote"          : processo.get("arquivo") or processo.get("id_lote"),
        "numero_processo"  : ent.get("numero_processo") or processo.get("numero_processo"),
        "nome_executado"   : ent.get("nome_executado"),
        "agente1"          : snapshot_a1,
        "analise": {
            "prioridade"        : prioridad,
            "acao_recomendada"  : accion,
            "justificativa"     : _justificativa(prioridad, processo),
            "alerta_prescricao" : alerta_presc,
            "observacoes"       : observacoes,
        },
        "processado_em": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    }


# ── Funciones de análisis determinístico ──────────────────────────

def _valor_numerico(valor_str):
    """Convierte 'R$ 228.433,16' a 228433.16."""
    if not valor_str:
        return 0.0
    try:
        limpio = str(valor_str).replace("R$","").replace(".","").replace(",",".").strip()
        return float(limpio)
    except ValueError:
        return 0.0



def _calcular_prioridad(processo: dict) -> str:
    """
    Prioridade OPERACIONAL (ordem em que o procurador analisa os casos).
    NÃO é critério de ajuizamento: BAIXA nunca significa "não cobrar" — o
    processo segue tramitando normalmente.
 
    Critérios (provisórios — ver limiares e notas abaixo):
      ALTA  → valor >= LIMIAR_PRIORIDADE_ALTA (R$ 5.000)
              OU risco de prescrição (art. 174 CTN)
              OU sem movimentação >= 5 anos  [ver ASSUNÇÃO nº 3]
      MEDIA → valor entre LIMIAR_PRIORIDADE_MEDIA (R$ 1.000) e ALTA
      BAIXA → valor < LIMIAR_PRIORIDADE_MEDIA (R$ 1.000)
 
    Dois eixos combinados: RECUPERABILIDADE (valor) e URGÊNCIA/RISCO
    (prescrição). Um crédito pequeno prestes a prescrever entra em ALTA
    porque perdê-lo é perda definitiva do crédito.
 
    ASSUNÇÃO PROVISÓRIA (definição HERA, NÃO jurídica): os limiares
    R$ 5.000 / R$ 1.000 foram escolhidos pela equipe apenas para ordenar o
    trabalho; não derivam de norma. Substituir pelo valor mínimo de
    ajuizamento oficial da PGMS quando for definido — basta trocar a variável
    de ambiente, sem tocar no fluxo.
 
    ASSUNÇÃO nº 3 (PENDENTE de confirmação PGMS): "sem movimentação >= 5 anos"
    pode configurar prescrição intercorrente (art. 40 §4º da LEF, Súmula 314
    do STJ). Gatilho de ALTA mantido idêntico à 0.1 por ora; revisar quando a
    prescrição intercorrente for modelada.
    """
    ent   = processo.get("entidades", {})
    valor = _valor_numerico(ent.get("valor_atualizado")) or \
            _valor_numerico(ent.get("valor_original"))
 
    # Eixo urgência/risco: prescrição empurra para ALTA, independente do valor.
    if _verificar_prescricao(processo):
        return "ALTA"
 
    anos_parado = 0
    ultima = processo.get("ultima_movimentacao")
    if ultima:
        try:
            dt = datetime.strptime(ultima, "%Y-%m-%d")
            anos_parado = (datetime.now() - dt).days / 365
        except ValueError:
            pass
 
    if valor >= LIMIAR_PRIORIDADE_ALTA or anos_parado >= 5:
        return "ALTA"
    elif valor >= LIMIAR_PRIORIDADE_MEDIA:
        return "MEDIA"
    else:
        return "BAIXA"


def _recomendar_accion(processo: dict) -> str:
    """
    Ação recomendada a partir dos sinais do PRÓPRIO processo.
    Já não depende de motivo_agente1 (removido junto com a triagem APTO):
    usa status_citacao, resultado_penhora e sinais_processuais.
    """
    citacao         = (processo.get("status_citacao") or "").upper()
    penhora         = (processo.get("resultado_penhora") or "").lower()
    citacao_efetiva = processo.get("data_citacao_efetiva")
    sinais          = processo.get("sinais_processuais") or {}

    # Sinais que mudam a estratégia antes de qualquer coisa.
    if sinais.get("extincao"):
        return "Sinal de extinção — verificar sentença e arquivar/baixar se transitado em julgado"
    if sinais.get("parcelamento"):
        return "Parcelamento identificado — suspender cobrança e acompanhar adimplemento das parcelas"
    if sinais.get("suspensao_art40_lef"):
        return "Suspenso pelo art. 40 da LEF — controlar prazo da prescrição intercorrente (Súmula 314 STJ)"

    # Citação não efetivada → nova tentativa.
    if not citacao_efetiva and ("NÃO HOUVE" in citacao or "FALHA" in citacao or "AUSENTE" in citacao):
        return "Realizar nova tentativa de citação — verificar endereços alternativos (SISBAJUD, RENAJUD, edital)"

    # Penhora não localizada → penhora online.
    if "não" in penhora or "nao" in penhora:
        return "Requerer penhora online via SISBAJUD / RENAJUD"

    return "Analisar histórico completo e definir estratégia de cobrança"


def _verificar_prescricao(processo: dict) -> bool:
    """
    Alerta de prescripción quinquenal (CTN art. 174):
    si han pasado >= 5 años desde el ejercicio fiscal sin citação válida.
    """
    ent      = processo.get("entidades", {})
    exercicio = ent.get("exercicio") or ""
    citacao   = (processo.get("status_citacao") or "").lower()

    # "não houve citação" contiene "houve citação" como substring — verificar negación
    citacao_valida = (
        "houve citação" in citacao
        and "não houve" not in citacao
        and "não ocorreu" not in citacao
    )
    if citacao_valida:
        return False  # citación válida interrumpió prescripción

    try:
        ano_base = int(str(exercicio)[:4])
        return (datetime.now().year - ano_base) >= 5
    except (ValueError, TypeError):
        return False


def _generar_observacoes(processo: dict) -> list:
    """Observaciones relevantes para el procurador."""
    obs = []
    ent = processo.get("entidades", {})

    if not ent.get("cpf_cnpj"):
        obs.append("CPF/CNPJ não identificado — verificar manualmente no SIAP")
    if not ent.get("numero_cda"):
        obs.append("Número de CDA não extraído do PDF — consultar sistema PGMS")
    if not ent.get("valor_atualizado"):
        obs.append("Valor atualizado não disponível — solicitar cálculo atualizado à PGMS")

    conf_ocr = processo.get("confianca_ocr_media")
    if conf_ocr is None:
        conf_ocr = (processo.get("ocr") or {}).get("confianca_media")
    if conf_ocr is not None and conf_ocr < 50:
        obs.append(f"Qualidade de OCR baixa ({conf_ocr}%) — dados podem conter erros")

    if _verificar_prescricao(processo):
        obs.append("⚠️  Risco de prescrição quinquenal — verificar interrupção do prazo (CTN art. 174)")

    if not obs:
        obs.append("Processo sem observações adicionais")

    return obs


def _justificativa(prioridad: str, processo: dict) -> str:
    ent    = processo.get("entidades", {})
    tipo   = processo.get("tipo_processo", {})
    valor  = ent.get("valor_atualizado") or ent.get("valor_original") or "não informado"
    classe = tipo.get("classe_assunto") or "—"
    # Mantém o trecho "valor da dívida: {valor}" — o Excel o extrai por regex
    # para ordenar por valor (ver gerar_reporte_xlsx).
    return f"Prioridade {prioridad} — valor da dívida: {valor}. Classe/assunto: {classe}."


# ════════════════════════════════════════════════════════════════════
# PROCESAMIENTO DE LOTE
# ════════════════════════════════════════════════════════════════════

def processar_lote(path_json: str):
    """
    Lee un JSON del Agente 1, analiza cada proceso y escribe resultado.
    Devuelve la ruta del archivo de resultado, o None si hubo error.
    """
    log.info(f"Processando lote: {os.path.basename(path_json)}")

    try:
        with open(path_json, encoding="utf-8") as f:
            payload = json.load(f)
    except Exception as e:
        log.error(f"Erro ao ler {path_json}: {e}")
        return None

    processos  = payload.get("processos", [])
    meta_a1    = payload.get("metadata", {})
    resultados = []

    for proc in processos:
        try:
            resultado = analisar_processo(proc)
            resultados.append(resultado)
            analise = resultado["analise"]
            _prio = analise.get("prioridade") or analise.get("status_triagem") or "—"
            _acao = analise.get("acao_recomendada") or "—"
            log.info(
                f"  [{_prio}] "
                f"{resultado.get('nome_executado','—')} — "
                f"{_acao[:55]}"
            )
        except Exception as e:
            log.error(f"  Erro em {proc.get('arquivo') or proc.get('id_lote') or '?'}: {e}")
            resultados.append({
                "id_lote"      : proc.get("arquivo") or proc.get("id_lote"),
                "erro"         : str(e),
                "processado_em": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            })

    # Conteo de prioridades
    prioridades = {"ALTA": 0, "MEDIA": 0, "BAIXA": 0}
    for r in resultados:
        p = r.get("analise", {}).get("prioridade")
        if p in prioridades:
            prioridades[p] += 1

    output_payload = {
        "metadata": {
            "gerado_em"        : datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "versao_agente2"   : VERSION_AGENTE2,
            "total_analisados" : len(resultados),
            "prioridades"      : prioridades,
            "origem_agente1"   : {
                "arquivo"         : os.path.basename(path_json),
                "gerado_em"        : meta_a1.get("gerado_em") or meta_a1.get("generado_em"),
                "total_processados": meta_a1.get("total_processados") or meta_a1.get("total_procesados"),
                "versao_agente1"   : meta_a1.get("versao_agente1") or meta_a1.get("version_agente1"),
            }
        },
        "analises": resultados,
    }

    # Asegurar que la carpeta resultados/ existe
    os.makedirs(PASTA_RESULTADOS, exist_ok=True)

    # Resultado por lote → *_agente2_resultado.json
    #
    # str.replace() não muda nada quando o sufixo não está no nome, e o
    # resultado ia gravado EM CIMA do arquivo de entrada — destruindo o JSON do
    # Agente 1 em qualquer chamada '--arquivo' com outro nome (o do próprio
    # Agente 1, 'saida_agente1_V8.json', é justamente um desses).
    base = os.path.basename(path_json)
    if base.endswith(SUFIJO_INPUT):
        nome_output = base[:-len(SUFIJO_INPUT)] + SUFIJO_OUTPUT
    else:
        nome_output = os.path.splitext(base)[0] + SUFIJO_OUTPUT
    path_output = os.path.join(PASTA_JSON, nome_output)

    if os.path.abspath(path_output) == os.path.abspath(path_json):
        log.error(
            f"O resultado sairia em cima da entrada ({base}) — abortando para "
            "não destruir o JSON do Agente 1."
        )
        return None

    try:
        # Resultado por lote (JSON único por lote)
        with open(path_output, "w", encoding="utf-8") as f:
            json.dump(output_payload, f, ensure_ascii=False, indent=2)
        log.info(f"  Resultado do lote: {os.path.basename(path_output)}")

        # Historial JSONL acumulativo → resultados/historial_agente2.jsonl
        pasta = os.path.dirname(os.path.abspath(path_json))
        path_jsonl = _ruta_historial(pasta)
        novos, atualizados = _escribir_historial(
            path_jsonl, resultados, os.path.basename(path_json)
        )
        log.info(
            f"  Histórico: {os.path.basename(path_jsonl)} — "
            f"{novos} novo(s), {atualizados} atualizado(s)"
        )

        # Generar reporte Excel acumulativo a partir del historial
        path_xlsx = os.path.join(PASTA_RESULTADOS, REPORTE_XLSX)
        gerar_reporte_xlsx(path_jsonl, path_xlsx)

        log.info(
            f"  Prioridades — "
            f"ALTA={prioridades['ALTA']} "
            f"MEDIA={prioridades['MEDIA']} "
            f"BAIXA={prioridades['BAIXA']}"
        )
        return path_output
    except Exception as e:
        log.error(f"Erro ao gravar o resultado: {e}")
        return None


# ════════════════════════════════════════════════════════════════════
# REPORTE EXCEL — acumulativo, para el procurador
# ════════════════════════════════════════════════════════════════════

def _valor_para_ordenar(valor_str):
    """Convierte 'R$ 228.433,16' a float para ordenar. Vacío → 0."""
    if not valor_str:
        return 0.0
    try:
        limpio = str(valor_str).replace("R$", "").replace(".", "").replace(",", ".").strip()
        return float(limpio)
    except (ValueError, AttributeError):
        return 0.0


def gerar_reporte_xlsx(path_jsonl: str, path_xlsx: str, origens=None):
    """
    Gera o relatório Excel de priorização a partir do histórico JSONL.

    Lê a vista DEDUPLICADA do histórico — uma linha por processo, com o estado
    mais recente e a contagem de quantas vezes ele foi reanalisado. Antes lia o
    JSONL append-only cru, e cada reprocessamento do mesmo processo virava uma
    linha nova: o procurador via o mesmo devedor repetido várias vezes.

    origens: conjunto de nomes de arquivo de origem (campo 'origem_lote').
             Quando informado, só entram os processos vindos desses lotes — é
             assim que a API entrega a cada consumidor um relatório com os
             lotes DELE. None = todos, que é o relatório do procurador.

    Ordena por prioridade (ALTA → MEDIA → BAIXA) e, dentro de cada nível, por
    valor da dívida decrescente.

    Requer: pandas, openpyxl.
    """
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        log.error(f"Falta dependência para gerar o relatório Excel: {e}")
        log.error("Instale com: pip install pandas openpyxl")
        return None

    if not os.path.exists(path_jsonl):
        log.warning(f"Sem histórico JSONL para gerar o relatório: {path_jsonl}")
        return None

    # ── Uma linha por processo: o estado mais recente ──
    filas = []
    for rec in _ultimo_estado_por_processo(path_jsonl):
        if origens is not None and rec.get("origem_lote") not in origens:
            continue
        analise = rec.get("analise", {})
        filas.append({
            "Prioridade"       : analise.get("prioridade", ""),
            "Nº Processo"      : rec.get("numero_processo", ""),
            "Executado"        : rec.get("nome_executado", ""),
            "Ação Recomendada" : analise.get("acao_recomendada", ""),
            "Alerta Prescrição": "SIM" if analise.get("alerta_prescricao") else "",
            "Justificativa"    : analise.get("justificativa", ""),
            "Observações"      : " | ".join(analise.get("observacoes", [])),
            "Arquivo de origem": rec.get("id_lote", "") or "",
            "Origem (lote)"    : rec.get("origem_lote", ""),
            "Processado em"    : rec.get("processado_em", ""),
            "Vezes analisado"  : rec.get("vez_analisada", 1),
        })

    if not filas:
        log.warning("Histórico vazio — o relatório Excel não foi gerado")
        return None

    df = pd.DataFrame(filas)

    # ── Ordenar: prioridade (ALTA→MEDIA→BAIXA) e valor decrescente ──
    # O valor não está no JSONL como número; sai da justificativa, que o
    # _justificar() escreve como "valor da dívida: R$ X".
    _ORDEM_PRIORIDADE = {"ALTA": 0, "MEDIA": 1, "BAIXA": 2, "": 3}
    df["_ord_prio"] = df["Prioridade"].map(lambda p: _ORDEM_PRIORIDADE.get(p, 3))

    def _extrair_valor_justif(j):
        m = re.search(r"valor da d[ií]vida:\s*(R\$[\d.,]+)", str(j))
        return _valor_para_ordenar(m.group(1)) if m else 0.0
    df["_ord_valor"] = df["Justificativa"].map(_extrair_valor_justif)

    df = df.sort_values(
        by=["_ord_prio", "_ord_valor"],
        ascending=[True, False]
    ).drop(columns=["_ord_prio", "_ord_valor"])

    # ── Escrever com formato ──
    # Grava num temporário e só então move por cima do arquivo final: o
    # relatório é baixável pela web enquanto o Agente 2 roda, e escrever no
    # lugar entregava um .xlsx truncado a quem baixasse no meio.
    destino = os.path.dirname(os.path.abspath(path_xlsx))
    os.makedirs(destino, exist_ok=True)
    # O pandas valida a extensão do arquivo de saída, então o temporário
    # também precisa terminar em .xlsx.
    path_tmp = path_xlsx[:-5] + ".parcial.xlsx" if path_xlsx.endswith(".xlsx") else path_xlsx + ".parcial.xlsx"

    with pd.ExcelWriter(path_tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Priorização")
        ws = writer.sheets["Priorização"]

        # Estilos
        font_header = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        fill_header = PatternFill("solid", fgColor="1F4E79")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde = Border(*[Side(style="thin", color="D0D0D0")] * 4)

        fills_prioridade = {
            "ALTA" : PatternFill("solid", fgColor="F8CBAD"),   # vermelho suave
            "MEDIA": PatternFill("solid", fgColor="FFE699"),   # amarelo suave
            "BAIXA": PatternFill("solid", fgColor="C6E0B4"),   # verde suave
        }

        # Cabeçalho
        for col_idx, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_header
            c.border = borde

        # Linhas de dados
        col_prioridade = list(df.columns).index("Prioridade") + 1
        for row_idx in range(2, len(df) + 2):
            prio = ws.cell(row=row_idx, column=col_prioridade).value
            fill = fills_prioridade.get(prio)
            for col_idx in range(1, len(df.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = Font(name="Arial", size=10)
                c.border = borde
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if fill and col_idx == col_prioridade:
                    c.fill = fill

        # Largura das colunas
        larguras = {
            "Prioridade": 11, "Nº Processo": 22, "Executado": 30,
            "Ação Recomendada": 38, "Alerta Prescrição": 10,
            "Justificativa": 50, "Observações": 40,
            "Arquivo de origem": 26, "Origem (lote)": 26,
            "Processado em": 20, "Vezes analisado": 10,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_name, 18)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

    os.replace(path_tmp, path_xlsx)
    log.info(f"  Relatório Excel: {os.path.basename(path_xlsx)} ({len(df)} processo(s))")
    return path_xlsx


# ════════════════════════════════════════════════════════════════════
# FILE WATCHER (stdlib pura — sin watchdog)
# ════════════════════════════════════════════════════════════════════

def _ya_procesado(path_json: str) -> bool:
    """True si el resultado ya existe en resultados/ y es más nuevo que el input."""
    nombre = os.path.basename(path_json).replace(SUFIJO_INPUT, SUFIJO_OUTPUT)
    path_resultado = os.path.join(PASTA_JSON, nombre)
    if not os.path.exists(path_resultado):
        return False
    return os.path.getmtime(path_resultado) >= os.path.getmtime(path_json)


def watcher(pasta: str, intervalo: int = INTERVALO_WATCH):
    """
    Monitorea la carpeta y procesa automáticamente cualquier
    *_agente2.json nuevo o modificado. Termina con Ctrl+C.
    """
    log.info(f"=== Agente 2 iniciado — v{VERSION_AGENTE2} ===")
    log.info(f"Monitorando: {pasta}")
    log.info(f"Sufixo procurado: *{SUFIJO_INPUT}")
    log.info(f"Intervalo: {intervalo}s — Ctrl+C para parar\n")

    procesados = set()  # evitar reprocesar en el mismo ciclo

    try:
        while True:
            jsons = sorted(Path(pasta).glob(f"*{SUFIJO_INPUT}"))

            for path_json in jsons:
                path_str = str(path_json)
                if not _ya_procesado(path_str):
                    if path_str not in procesados:
                        log.info(f"Novo arquivo detectado: {path_json.name}")
                    processar_lote(path_str)
                    procesados.add(path_str)
                elif path_str in procesados:
                    # Archivo ya procesado en esta sesión — ignorar silencioso
                    pass

            time.sleep(intervalo)

    except KeyboardInterrupt:
        log.info("Agente 2 encerrado.")


# ════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Agente 2 — Raciocínio Jurídico-Fiscal (PGMS/HERA)"
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument(
        "--watch", action="store_true",
        help="Modo watcher: monitorea la carpeta continuamente"
    )
    grupo.add_argument(
        "--arquivo", metavar="PATH",
        help="Modo puntual: procesa un JSON específico y termina"
    )
    parser.add_argument(
        "--pasta", metavar="DIR", default=PASTA_JSON,
        help=f"Carpeta a monitorear (default: directorio del script)"
    )
    parser.add_argument(
        "--intervalo", type=int, default=INTERVALO_WATCH,
        help=f"Segundos entre chequeos en modo watcher (default: {INTERVALO_WATCH})"
    )

    args = parser.parse_args()

    if args.watch:
        watcher(args.pasta, args.intervalo)
    else:
        if not os.path.exists(args.arquivo):
            log.error(f"Arquivo não encontrado: {args.arquivo}")
            sys.exit(1)
        resultado = processar_lote(args.arquivo)
        sys.exit(0 if resultado else 1)


if __name__ == "__main__":
    main()